from __future__ import annotations

import csv
import hashlib
import io
import json
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

import xlrd
from backend.app.repositories.ledger_import_repo import LedgerImportRepository
from backend.app.schemas.ledger_import import (
    LedgerImportBatchSummary,
    LedgerImportListItem,
)
from openpyxl import load_workbook
from xlrd import xldate_as_datetime

RULE_VERSION = "position_key_contract_v1"
NULL_TEXT = "__NULL__"
SUPPORTED_SUFFIXES = {".xls", ".xlsx", ".csv"}
LEDGER_SHEET_NAME = "ZQTZSHOW"


@dataclass(frozen=True, slots=True)
class FieldSpec:
    source_field: str
    standard_field: str
    value_kind: str


FIELD_SPECS = (
    FieldSpec("债券代号", "bond_code", "text"),
    FieldSpec("债券名称", "bond_name", "text"),
    FieldSpec("交易对手CIF号", "counterparty_cif_no", "text"),
    FieldSpec("投资组合", "portfolio", "text"),
    FieldSpec("日期", "as_of_date", "date"),
    FieldSpec("业务种类", "business_type", "text"),
    FieldSpec("授信客户属性", "credit_customer_attribute", "text"),
    FieldSpec("业务种类1", "business_type_1", "text"),
    FieldSpec("账户类别", "account_category_std", "text"),
    FieldSpec("成本中心", "cost_center", "text"),
    FieldSpec("资产分类", "asset_class_std", "text"),
    FieldSpec("风险缓释", "risk_mitigation", "text"),
    FieldSpec("面值", "face_amount", "decimal"),
    FieldSpec("公允价值", "fair_value", "decimal"),
    FieldSpec("摊余成本", "amortized_cost", "decimal"),
    FieldSpec("应计利息", "accrued_interest", "decimal"),
    FieldSpec("计息方式", "interest_method", "text"),
    FieldSpec("利率", "coupon_rate", "decimal"),
    FieldSpec("起息日", "interest_start_date", "date"),
    FieldSpec("到期日", "maturity_date", "date"),
    FieldSpec("利率浮动代码", "interest_rate_benchmark_code", "text"),
    FieldSpec("利率浮动频率", "interest_rate_reset_frequency", "text"),
    FieldSpec("交易对手行业大类", "counterparty_industry", "text"),
    FieldSpec("交易对手客户名称中文", "counterparty_name_cn", "text"),
    FieldSpec("授信客户证件编号", "credit_customer_id", "text"),
    FieldSpec("授信客户客户号", "credit_customer_no", "text"),
    FieldSpec("授信客户债券评级", "credit_customer_rating", "text"),
    FieldSpec("授信客户行业大类", "credit_customer_industry", "text"),
    FieldSpec("应收/应付利息", "interest_receivable_payable", "decimal"),
    FieldSpec("币种", "currency", "text"),
    FieldSpec("授信客户名称", "credit_customer_name", "text"),
    FieldSpec("减值数据（手工补录）", "manual_impairment_adjustment", "mixed"),
    FieldSpec("渠道", "channel", "text"),
    FieldSpec("法人客户名称", "legal_customer_name", "text"),
    FieldSpec("法人客户证件号", "legal_customer_id", "text"),
    FieldSpec("集团客户名称", "group_customer_name", "text"),
    FieldSpec("集团客户证件号", "group_customer_id", "text"),
    FieldSpec("还本逾期", "principal_overdue_flag", "mixed"),
    FieldSpec("收息逾期", "interest_overdue_flag", "mixed"),
    FieldSpec("数量", "quantity", "decimal"),
    FieldSpec("最新面值", "latest_face_value", "decimal"),
    FieldSpec("本金逾期天数", "principal_overdue_days", "integer"),
    FieldSpec("利息逾期天数", "interest_overdue_days", "integer"),
    FieldSpec("到期收益率", "yield_to_maturity", "decimal"),
    FieldSpec("下一行权日/逾期资产到期日", "option_or_special_maturity_date", "date"),
)

KEY_FIELDS = (
    "bond_code",
    "bond_name",
    "portfolio",
    "account_category_std",
    "cost_center",
    "asset_class_std",
    "currency",
    "counterparty_cif_no",
    "credit_customer_id",
    "legal_customer_id",
    "interest_start_date",
    "maturity_date",
    "channel",
)


@dataclass(slots=True)
class ParsedLedgerFile:
    file_name: str
    file_hash: str
    source_version: str
    as_of_date: str
    rows: list[dict[str, object]]


class LedgerImportService:
    def __init__(self, duckdb_path: str) -> None:
        self.repo = LedgerImportRepository(duckdb_path)

    def import_file(self, *, file_name: str, content: bytes) -> dict[str, object]:
        parsed = parse_ledger_file(file_name=file_name, content=content)
        summary = self.repo.insert_import(
            file_name=parsed.file_name,
            file_hash=parsed.file_hash,
            as_of_date=parsed.as_of_date,
            rows=parsed.rows,
            source_version=parsed.source_version,
            rule_version=RULE_VERSION,
        )
        duplicate = str(summary["status"]) == "duplicate"
        data = _batch_summary_data(summary)
        if duplicate:
            data["status"] = "duplicate"
        return {
            "data": data,
            **(
                {
                    "error": {
                        "code": "LEDGER_IMPORT_DUPLICATE",
                        "message": "Ledger file has already been imported.",
                        "retryable": False,
                    }
                }
                if duplicate
                else {}
            ),
            "metadata": _metadata(
                batch_id=int(summary["batch_id"]),
                source_version=str(summary["source_version"]),
                no_data=False,
            ),
            "trace": _trace(
                source_file_hash=parsed.file_hash,
                batch_id=int(summary["batch_id"]),
                duplicate_of_batch_id=(
                    int(summary["duplicate_of_batch_id"])
                    if summary.get("duplicate_of_batch_id") is not None
                    else None
                ),
            ),
        }

    def list_imports(self) -> dict[str, object]:
        items = [
            LedgerImportListItem(**item).model_dump(mode="json")
            for item in self.repo.list_batches()
        ]
        latest = items[0] if items else None
        return {
            "data": {"items": items, "total": len(items)},
            "metadata": _metadata(
                batch_id=int(latest["batch_id"]) if latest else None,
                source_version=str(latest["source_version"]) if latest else None,
                no_data=not bool(items),
            ),
            "trace": _trace(
                batch_id=int(latest["batch_id"]) if latest else None,
                source_file_hash=None,
            ),
        }


def parse_ledger_file(*, file_name: str, content: bytes) -> ParsedLedgerFile:
    suffix = Path(file_name).suffix.lower()
    if suffix not in SUPPORTED_SUFFIXES:
        raise ValueError(f"Unsupported ledger import file type: {suffix or '<none>'}")
    file_digest = hashlib.sha256(content).hexdigest()
    file_hash = f"sha256:{file_digest}"
    source_version = f"sv_ledger_{file_digest[:12]}"
    raw_rows = _read_raw_rows(file_name=file_name, content=content, suffix=suffix)
    if not raw_rows:
        raise ValueError("Ledger import file contains no data rows.")

    rows: list[dict[str, object]] = []
    as_of_dates: set[str] = set()
    for row_no, raw_row in enumerate(raw_rows, start=1):
        row = _standardize_row(
            row_no=row_no,
            raw_row=raw_row,
            source_version=source_version,
        )
        as_of_date = str(row.get("as_of_date") or "")
        if not as_of_date:
            raise ValueError(f"Ledger row {row_no} is missing required as_of_date.")
        as_of_dates.add(as_of_date)
        rows.append(row)

    if len(as_of_dates) != 1:
        raise ValueError(f"Ledger import contains mixed as_of_date values: {sorted(as_of_dates)}")

    return ParsedLedgerFile(
        file_name=file_name,
        file_hash=file_hash,
        source_version=source_version,
        as_of_date=next(iter(as_of_dates)),
        rows=rows,
    )


def _read_raw_rows(*, file_name: str, content: bytes, suffix: str) -> list[dict[str, object]]:
    if suffix == ".xls":
        return _read_xls_rows(content)
    if suffix == ".xlsx":
        return _read_xlsx_rows(content)
    if suffix == ".csv":
        return _read_csv_rows(content)
    raise ValueError(f"Unsupported ledger import file type: {suffix}")


def _read_xls_rows(content: bytes) -> list[dict[str, object]]:
    workbook = xlrd.open_workbook(file_contents=content)
    if LEDGER_SHEET_NAME not in workbook.sheet_names():
        raise ValueError(f"Ledger workbook must contain sheet {LEDGER_SHEET_NAME}.")
    sheet = workbook.sheet_by_name(LEDGER_SHEET_NAME)
    if sheet.nrows < 3:
        return []
    headers = [_header_text(sheet.cell_value(1, column)) for column in range(sheet.ncols)]
    rows: list[dict[str, object]] = []
    for row_index in range(2, sheet.nrows):
        values = [
            _xls_cell_value(sheet.cell(row_index, column), datemode=workbook.datemode)
            for column in range(sheet.ncols)
        ]
        row = _raw_row(headers, values)
        if not _is_blank_row(row):
            rows.append(row)
    return rows


def _read_xlsx_rows(content: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        if LEDGER_SHEET_NAME not in workbook.sheetnames:
            raise ValueError(f"Ledger workbook must contain sheet {LEDGER_SHEET_NAME}.")
        worksheet = workbook[LEDGER_SHEET_NAME]
        header_values = next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if header_values is None:
            return []
        headers = [_header_text(value) for value in header_values]
        rows: list[dict[str, object]] = []
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            row = _raw_row(headers, list(values))
            if not _is_blank_row(row):
                rows.append(row)
        return rows
    finally:
        workbook.close()


def _read_csv_rows(content: bytes) -> list[dict[str, object]]:
    text = _decode_csv_content(content)
    reader = list(csv.reader(io.StringIO(text)))
    if len(reader) < 3:
        return []
    headers = [_header_text(value) for value in reader[1]]
    rows: list[dict[str, object]] = []
    for values in reader[2:]:
        row = _raw_row(headers, values)
        if not _is_blank_row(row):
            rows.append(row)
    return rows


def _standardize_row(
    *,
    row_no: int,
    raw_row: dict[str, object],
    source_version: str,
) -> dict[str, object]:
    standard: dict[str, object] = {
        "row_no": row_no,
        "source_version": source_version,
        "rule_version": RULE_VERSION,
    }
    for spec in FIELD_SPECS:
        raw_value = raw_row.get(spec.source_field)
        if spec.value_kind == "text":
            value = _normalize_text(raw_value)
        elif spec.value_kind == "date":
            value = _normalize_date(raw_value)
        elif spec.value_kind == "decimal":
            value = _normalize_decimal(raw_value)
        elif spec.value_kind == "integer":
            value = _normalize_integer(raw_value)
        elif spec.value_kind == "mixed":
            value = _normalize_mixed(raw_value)
        else:
            raise ValueError(f"Unsupported field kind: {spec.value_kind}")
        if spec.standard_field in {"account_category_std", "asset_class_std"}:
            value = _normalize_issuance_alias(value)
        standard[spec.standard_field] = value

    standard["direction"] = _direction(standard)
    standard["position_key"] = _position_key(standard)
    standard["raw_json"] = json.dumps(
        {key: _json_safe(value) for key, value in raw_row.items()},
        ensure_ascii=False,
        sort_keys=True,
    )
    return standard


def _raw_row(headers: list[str], values: list[object]) -> dict[str, object]:
    row: dict[str, object] = {}
    for index, header in enumerate(headers):
        if not header:
            continue
        value = values[index] if index < len(values) else None
        row[header] = _json_safe(value)
    return row


def _direction(row: dict[str, object]) -> str:
    if row.get("account_category_std") == "发行类债券" or row.get("asset_class_std") == "发行类债券":
        return "LIABILITY"
    return "ASSET"


def _position_key(row: dict[str, object]) -> str:
    canonical = "|".join(str(row.get(field) or NULL_TEXT) for field in KEY_FIELDS)
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()


def _metadata(
    *,
    batch_id: int | None,
    source_version: str | None,
    no_data: bool,
) -> dict[str, object]:
    return {
        "source_version": source_version,
        "rule_version": RULE_VERSION,
        "batch_id": batch_id,
        "stale": False,
        "fallback": False,
        "no_data": no_data,
    }


def _trace(
    *,
    batch_id: int | None,
    source_file_hash: str | None,
    duplicate_of_batch_id: int | None = None,
) -> dict[str, object]:
    return {
        "request_id": f"req_ledger_{uuid4().hex[:12]}",
        "source_file_hash": source_file_hash,
        "batch_id": batch_id,
        "duplicate_of_batch_id": duplicate_of_batch_id,
    }


def _batch_summary_data(summary: dict[str, object]) -> dict[str, object]:
    return LedgerImportBatchSummary(
        batch_id=int(summary["batch_id"]),
        file_name=str(summary["file_name"]),
        file_hash=str(summary["file_hash"]),
        as_of_date=str(summary["as_of_date"]),
        status=str(summary["status"]),
        row_count=int(summary["row_count"]),
        error_count=int(summary["error_count"]),
        source_version=str(summary["source_version"]),
        rule_version=str(summary["rule_version"]),
    ).model_dump(mode="json")


def _header_text(value: object) -> str:
    return _cell_text(value)


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(value.quantize(Decimal("1")))
        return format(value, "f")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_text(value: object) -> str:
    text = unicodedata.normalize("NFKC", _cell_text(value)).strip().upper()
    return text or NULL_TEXT


def _normalize_mixed(value: object) -> str:
    text = unicodedata.normalize("NFKC", _cell_text(value)).strip().upper()
    return text or NULL_TEXT


def _normalize_issuance_alias(value: object) -> object:
    if value in {"发行类债劵", "发行类债券"}:
        return "发行类债券"
    return value


def _normalize_decimal(value: object) -> Decimal | None:
    text = _cell_text(value).replace(",", "")
    if not text:
        return None
    try:
        if text.endswith("%"):
            return Decimal(text[:-1].strip()) / Decimal("100")
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid decimal value: {value!r}") from exc


def _normalize_integer(value: object) -> int | None:
    decimal_value = _normalize_decimal(value)
    if decimal_value is None:
        return None
    return int(decimal_value)


def _normalize_date(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _cell_text(value)
    if not text:
        return None
    normalized = text.replace("/", "-")
    if len(normalized) == 8 and normalized.isdigit():
        return f"{normalized[:4]}-{normalized[4:6]}-{normalized[6:]}"
    try:
        return date.fromisoformat(normalized[:10]).isoformat()
    except ValueError:
        return normalized


def _xls_cell_value(cell, *, datemode: int) -> object:
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xldate_as_datetime(cell.value, datemode)
    return cell.value


def _json_safe(value: object) -> object:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def _decode_csv_content(content: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _is_blank_row(row: dict[str, object]) -> bool:
    return all(_cell_text(value) == "" for value in row.values())
