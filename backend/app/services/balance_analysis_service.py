from __future__ import annotations

"""
Balance-analysis read surfaces: DuckDB access only via `BalanceAnalysisRepository` and other read repositories.

Formal fact writes (`replace_formal_balance_rows`, snapshot tables) are restricted to `backend/app/tasks/` workers.
"""

import csv
import importlib
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.core_finance.balance_analysis import FormalTywBalanceFactRow, FormalZqtzBalanceFactRow
from backend.app.governance.formal_compute_lineage import (
    resolve_completed_formal_build_lineage,
    resolve_formal_manifest_lineage,
)
from backend.app.governance.locks import LockDefinition, acquire_lock
from backend.app.governance.settings import Settings
from backend.app.repositories.balance_analysis_decision_repo import BalanceAnalysisDecisionRepository
from backend.app.repositories.balance_analysis_repo import BalanceAnalysisRepository
from backend.app.repositories.governance_repo import (
    CACHE_BUILD_RUN_STREAM,
    GovernanceRepository,
)
from backend.app.schemas.balance_analysis import (
    BalanceAnalysisBasisBreakdownPayload,
    BalanceAnalysisBasisBreakdownRow,
    BalanceAnalysisDecisionItemRow,
    BalanceAnalysisDecisionItemStatusRow,
    BalanceAnalysisDecisionItemsSection,
    BalanceAnalysisDecisionItemsPayload,
    BalanceAnalysisDecisionStatusRecord,
    BalanceAnalysisDecisionStatusUpdateRequest,
    BalanceAnalysisDatesPayload,
    BalanceAnalysisDetailRow,
    BalanceAnalysisEventCalendarRow,
    BalanceAnalysisEventCalendarSection,
    BalanceAnalysisPayload,
    BalanceAnalysisRiskAlertRow,
    BalanceAnalysisRiskAlertsSection,
    BalanceAnalysisSummaryRow,
    BalanceAnalysisSummaryTablePayload,
    BalanceAnalysisTableRow,
    BalanceAnalysisWorkbookCard,
    BalanceAnalysisWorkbookColumn,
    BalanceAnalysisWorkbookPayload,
    BalanceAnalysisWorkbookTable,
)
from backend.app.schemas.materialize import CacheBuildRunRecord
from backend.app.services.formal_result_runtime import (
    build_formal_result_envelope,
    build_formal_result_meta,
)
from backend.app.tasks.balance_analysis_materialize import (
    BALANCE_ANALYSIS_LOCK,
    CACHE_KEY,
    CACHE_VERSION,
    RULE_VERSION,
    materialize_balance_analysis_facts,
)

BALANCE_ANALYSIS_JOB_NAME = "balance_analysis_materialize"
PENDING_SOURCE_VERSION = "sv_balance_analysis_pending"
ALLOWED_BALANCE_POSITION_SCOPES = frozenset({"asset", "liability", "all"})
ALLOWED_BALANCE_CURRENCY_BASES = frozenset({"native", "CNY"})
IN_FLIGHT_STATUSES = {"queued", "running"}
STALE_IN_FLIGHT_AFTER = timedelta(hours=1)
EXPORT_WORKBOOK_TABLES = (
    ("债券持仓", ("zqtz_balance", "bond_business_types")),
    ("同业持仓", ("tyw_balance", "counterparty_types")),
    ("期限分布", ("maturity_distribution", "maturity_gap")),
    ("利率分布", ("rate_distribution",)),
)
EXCEL_HEADER_FONT = Font(bold=True)
EXCEL_RIGHT_ALIGNMENT = Alignment(horizontal="right")
EXCEL_AMOUNT_FORMAT = "#,##0.00000000"
EXCEL_NUMBER_FORMAT = "0.00000000"
EXCEL_INTEGER_FORMAT = "0"
DEFAULT_BALANCE_DECISION_UPDATED_BY = "balance-analysis-ui"


class BalanceAnalysisRefreshServiceError(RuntimeError):
    pass


class BalanceAnalysisRefreshConflictError(RuntimeError):
    pass


def refresh_balance_analysis(settings: Settings, *, report_date: str) -> dict[str, object]:
    try:
        with acquire_lock(
            _refresh_trigger_lock(report_date=report_date),
            base_dir=settings.governance_path,
            timeout_seconds=0.1,
        ):
            existing = _latest_inflight_refresh(settings, report_date=report_date)
            if existing is not None:
                raise BalanceAnalysisRefreshConflictError(
                    f"Balance-analysis refresh already in progress for report_date={report_date}."
                )

            run_id = _build_run_id()
            queued_at = datetime.now(timezone.utc).isoformat()
            GovernanceRepository(base_dir=settings.governance_path).append(
                CACHE_BUILD_RUN_STREAM,
                {
                    **CacheBuildRunRecord(
                        run_id=run_id,
                        job_name=BALANCE_ANALYSIS_JOB_NAME,
                        status="queued",
                        cache_key=CACHE_KEY,
                        cache_version=CACHE_VERSION,
                        lock=BALANCE_ANALYSIS_LOCK.key,
                        source_version=PENDING_SOURCE_VERSION,
                        vendor_version="vv_none",
                    ).model_dump(),
                    "report_date": report_date,
                    "queued_at": queued_at,
                },
            )
            try:
                materialize_balance_analysis_facts.send(
                    report_date=report_date,
                    duckdb_path=str(settings.duckdb_path),
                    governance_dir=str(settings.governance_path),
                    run_id=run_id,
                )
            except Exception as exc:
                _record_dispatch_failure(
                    settings=settings,
                    run_id=run_id,
                    report_date=report_date,
                    error_message="Balance-analysis refresh queue dispatch failed.",
                )
                raise BalanceAnalysisRefreshServiceError(
                    "Balance-analysis refresh queue dispatch failed."
                ) from exc

            return {
                "status": "queued",
                "run_id": run_id,
                "job_name": BALANCE_ANALYSIS_JOB_NAME,
                "trigger_mode": "async",
                "cache_key": CACHE_KEY,
                "report_date": report_date,
            }
    except TimeoutError as exc:
        raise BalanceAnalysisRefreshConflictError(
            f"Balance-analysis refresh already in progress for report_date={report_date}."
        ) from exc


def balance_analysis_refresh_status(settings: Settings, *, run_id: str) -> dict[str, object]:
    records = [
        record
        for record in GovernanceRepository(base_dir=settings.governance_path).read_all(CACHE_BUILD_RUN_STREAM)
        if str(record.get("cache_key")) == CACHE_KEY
        and str(record.get("job_name")) == BALANCE_ANALYSIS_JOB_NAME
        and str(record.get("run_id")) == run_id
    ]
    if not records:
        raise ValueError(f"Unknown balance-analysis refresh run_id={run_id}")
    latest = records[-1]
    status = str(latest.get("status", "unknown"))
    return {
        **latest,
        "trigger_mode": "async" if status in {"queued", "running"} else "terminal",
    }


def balance_analysis_dates_envelope(*, duckdb_path: str, governance_dir: str) -> dict[str, object]:
    repo = BalanceAnalysisRepository(duckdb_path)
    payload = BalanceAnalysisDatesPayload(report_dates=repo.list_report_dates())
    lineage = resolve_formal_manifest_lineage(
        governance_dir=governance_dir,
        cache_key=CACHE_KEY,
    )
    meta = build_formal_result_meta(
        trace_id="tr_balance_analysis_dates",
        result_kind="balance-analysis.dates",
        cache_version=_resolve_balance_cache_version(lineage),
        source_version=str(lineage["source_version"]),
        rule_version=str(lineage["rule_version"]),
    )
    return build_formal_result_envelope(
        result_meta=meta,
        result_payload=payload.model_dump(mode="json"),
    )


def balance_analysis_overview_envelope(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
) -> dict[str, object]:
    _validate_balance_overview_filters(
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    repo = BalanceAnalysisRepository(duckdb_path)
    if report_date not in repo.list_report_dates():
        raise ValueError(f"No balance-analysis data found for report_date={report_date}.")

    overview = repo.fetch_formal_overview(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    build_lineage = resolve_completed_formal_build_lineage(
        governance_dir=governance_dir,
        cache_key=CACHE_KEY,
        job_name=BALANCE_ANALYSIS_JOB_NAME,
        report_date=report_date,
    )
    meta = build_formal_result_meta(
        trace_id=f"tr_balance_analysis_overview_{report_date}_{position_scope}_{currency_basis}",
        result_kind="balance-analysis.overview",
        cache_version=_resolve_balance_cache_version(build_lineage),
        source_version=_require_balance_lineage_value(
            build_lineage["source_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="source_version",
        ),
        rule_version=_require_balance_lineage_value(
            (build_lineage.get("rule_version") if build_lineage is not None else None)
            or overview.get("rule_version"),
            report_date=report_date,
            field_name="rule_version",
        ),
    )
    return build_formal_result_envelope(
        result_meta=meta,
        result_payload={
            "report_date": str(overview["report_date"]),
            "position_scope": str(overview["position_scope"]),
            "currency_basis": str(overview["currency_basis"]),
            "detail_row_count": int(overview["detail_row_count"]),
            "summary_row_count": int(overview["summary_row_count"]),
            "total_market_value_amount": _as_decimal(overview["total_market_value_amount"]),
            "total_amortized_cost_amount": _as_decimal(overview["total_amortized_cost_amount"]),
            "total_accrued_interest_amount": _as_decimal(overview["total_accrued_interest_amount"]),
        },
    )


def balance_analysis_summary_envelope(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
    limit: int = 50,
    offset: int = 0,
) -> dict[str, object]:
    _validate_balance_overview_filters(
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    repo = BalanceAnalysisRepository(duckdb_path)
    if report_date not in repo.list_report_dates():
        raise ValueError(f"No balance-analysis data found for report_date={report_date}.")

    table = repo.fetch_formal_summary_table(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        limit=limit,
        offset=offset,
    )
    build_lineage = resolve_completed_formal_build_lineage(
        governance_dir=governance_dir,
        cache_key=CACHE_KEY,
        job_name=BALANCE_ANALYSIS_JOB_NAME,
        report_date=report_date,
    )
    meta = build_formal_result_meta(
        trace_id=f"tr_balance_analysis_summary_{report_date}_{position_scope}_{currency_basis}_{offset}_{limit}",
        result_kind="balance-analysis.summary",
        cache_version=_resolve_balance_cache_version(build_lineage),
        source_version=_require_balance_lineage_value(
            build_lineage["source_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="source_version",
        ),
        rule_version=_require_balance_lineage_value(
            build_lineage["rule_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="rule_version",
        ),
    )
    payload = BalanceAnalysisSummaryTablePayload(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        limit=limit,
        offset=offset,
        total_rows=int(table["total_rows"]),
        rows=[_to_summary_table_row(row) for row in table["rows"]],
    )
    return build_formal_result_envelope(
        result_meta=meta,
        result_payload=payload.model_dump(mode="json"),
    )


def balance_analysis_basis_breakdown_envelope(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
) -> dict[str, object]:
    _validate_balance_overview_filters(
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    repo = BalanceAnalysisRepository(duckdb_path)
    if report_date not in repo.list_report_dates():
        raise ValueError(f"No balance-analysis data found for report_date={report_date}.")

    breakdown_rows = repo.fetch_formal_basis_breakdown(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    build_lineage = resolve_completed_formal_build_lineage(
        governance_dir=governance_dir,
        cache_key=CACHE_KEY,
        job_name=BALANCE_ANALYSIS_JOB_NAME,
        report_date=report_date,
    )
    meta = build_formal_result_meta(
        trace_id=f"tr_balance_analysis_basis_breakdown_{report_date}_{position_scope}_{currency_basis}",
        result_kind="balance-analysis.basis-breakdown",
        cache_version=_resolve_balance_cache_version(build_lineage),
        source_version=_require_balance_lineage_value(
            build_lineage["source_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="source_version",
        ),
        rule_version=_require_balance_lineage_value(
            build_lineage["rule_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="rule_version",
        ),
    )
    payload = BalanceAnalysisBasisBreakdownPayload(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        rows=[_to_basis_breakdown_row(row) for row in breakdown_rows],
    )
    return build_formal_result_envelope(
        result_meta=meta,
        result_payload=payload.model_dump(mode="json"),
    )


def export_balance_analysis_summary_csv(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
) -> tuple[str, str]:
    _validate_balance_overview_filters(
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    repo = BalanceAnalysisRepository(duckdb_path)
    if report_date not in repo.list_report_dates():
        raise ValueError(f"No balance-analysis data found for report_date={report_date}.")

    table = repo.fetch_formal_summary_table(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        limit=None,
        offset=0,
    )
    build_lineage = resolve_completed_formal_build_lineage(
        governance_dir=governance_dir,
        cache_key=CACHE_KEY,
        job_name=BALANCE_ANALYSIS_JOB_NAME,
        report_date=report_date,
    )
    source_version = _require_balance_lineage_value(
        build_lineage["source_version"] if build_lineage is not None else None,
        report_date=report_date,
        field_name="source_version",
    )
    rule_version = _require_balance_lineage_value(
        build_lineage["rule_version"] if build_lineage is not None else None,
        report_date=report_date,
        field_name="rule_version",
    )
    rows = [_to_summary_table_row(row) for row in table["rows"]]
    filename = f"balance-analysis-summary-{report_date}-{position_scope}-{currency_basis}.csv"
    return filename, _build_balance_summary_csv(
        rows,
        report_date=report_date,
        source_version=source_version,
        rule_version=rule_version,
    )


def export_balance_analysis_workbook_xlsx(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
) -> tuple[str, bytes]:
    workbook_payload = balance_analysis_workbook_envelope(
        duckdb_path=duckdb_path,
        governance_dir=governance_dir,
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )["result"]
    filename = f"资产负债分析_{report_date}.xlsx"
    return filename, _build_balance_analysis_workbook_xlsx_bytes(workbook_payload)


def balance_analysis_detail_envelope(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
) -> dict[str, object]:
    repo = BalanceAnalysisRepository(duckdb_path)
    if report_date not in repo.list_report_dates():
        raise ValueError(f"No balance-analysis data found for report_date={report_date}.")

    zqtz_rows = repo.fetch_formal_zqtz_rows(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    tyw_rows = repo.fetch_formal_tyw_rows(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )

    details = [
        *[_to_zqtz_detail_row(row) for row in zqtz_rows],
        *[_to_tyw_detail_row(row) for row in tyw_rows],
    ]
    summary = _build_summary_rows(details)
    build_lineage = resolve_completed_formal_build_lineage(
        governance_dir=governance_dir,
        cache_key=CACHE_KEY,
        job_name=BALANCE_ANALYSIS_JOB_NAME,
        report_date=report_date,
    )

    payload = BalanceAnalysisPayload(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        details=details,
        summary=summary,
    )
    meta = build_formal_result_meta(
        trace_id=f"tr_balance_analysis_detail_{report_date}_{position_scope}_{currency_basis}",
        result_kind="balance-analysis.detail",
        cache_version=_resolve_balance_cache_version(build_lineage),
        source_version=(
            str(build_lineage["source_version"])
            if build_lineage is not None
            else _combine_lineage_values([*zqtz_rows, *tyw_rows], "source_version")
        ),
        rule_version=(
            str(build_lineage["rule_version"])
            if build_lineage is not None and str(build_lineage.get("rule_version") or "").strip()
            else _combine_lineage_values([*zqtz_rows, *tyw_rows], "rule_version") or RULE_VERSION
        ),
    )
    return build_formal_result_envelope(
        result_meta=meta,
        result_payload=payload.model_dump(mode="json"),
    )


def balance_analysis_workbook_envelope(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
) -> dict[str, object]:
    _validate_balance_overview_filters(
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    workbook, build_lineage = _build_balance_workbook_payload(
        duckdb_path=duckdb_path,
        governance_dir=governance_dir,
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    payload = BalanceAnalysisWorkbookPayload(
        report_date=workbook["report_date"],
        position_scope=workbook["position_scope"],
        currency_basis=workbook["currency_basis"],
        cards=[
            BalanceAnalysisWorkbookCard(**card)
            for card in workbook["cards"]
        ],
        tables=[
            BalanceAnalysisWorkbookTable(
                key=table["key"],
                title=table["title"],
                section_kind=table["section_kind"],
                columns=[BalanceAnalysisWorkbookColumn(**column) for column in table["columns"]],
                rows=table["rows"],
            )
            for table in workbook["tables"]
            if str(table.get("section_kind")) == "table"
        ],
        operational_sections=[
            *[
                BalanceAnalysisDecisionItemsSection(
                    key="decision_items",
                    title=str(section["title"]),
                    section_kind="decision_items",
                    columns=[BalanceAnalysisWorkbookColumn(**column) for column in section["columns"]],
                    rows=[BalanceAnalysisDecisionItemRow(**row) for row in section["rows"]],
                )
                for section in workbook["tables"]
                if str(section.get("section_kind")) == "decision_items"
            ],
            *[
                BalanceAnalysisEventCalendarSection(
                    key="event_calendar",
                    title=str(section["title"]),
                    section_kind="event_calendar",
                    columns=[BalanceAnalysisWorkbookColumn(**column) for column in section["columns"]],
                    rows=[BalanceAnalysisEventCalendarRow(**row) for row in section["rows"]],
                )
                for section in workbook["tables"]
                if str(section.get("section_kind")) == "event_calendar"
            ],
            *[
                BalanceAnalysisRiskAlertsSection(
                    key="risk_alerts",
                    title=str(section["title"]),
                    section_kind="risk_alerts",
                    columns=[BalanceAnalysisWorkbookColumn(**column) for column in section["columns"]],
                    rows=[BalanceAnalysisRiskAlertRow(**row) for row in section["rows"]],
                )
                for section in workbook["tables"]
                if str(section.get("section_kind")) == "risk_alerts"
            ],
        ],
    )
    meta = build_formal_result_meta(
        trace_id=f"tr_balance_analysis_workbook_{report_date}_{position_scope}_{currency_basis}",
        result_kind="balance-analysis.workbook",
        cache_version=_resolve_balance_cache_version(build_lineage),
        source_version=_require_balance_lineage_value(
            build_lineage["source_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="source_version",
        ),
        rule_version=_require_balance_lineage_value(
            build_lineage["rule_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="rule_version",
        ),
    )
    return build_formal_result_envelope(
        result_meta=meta,
        result_payload=payload.model_dump(mode="json"),
    )


def balance_analysis_decision_items_envelope(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"] = "all",
    currency_basis: Literal["native", "CNY"] = "CNY",
) -> dict[str, object]:
    _validate_balance_overview_filters(
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    workbook, build_lineage = _build_balance_workbook_payload(
        duckdb_path=duckdb_path,
        governance_dir=governance_dir,
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    decision_section = _extract_generated_decision_section(workbook)
    latest_statuses = BalanceAnalysisDecisionRepository(governance_dir).list_latest_statuses(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
    )
    payload = BalanceAnalysisDecisionItemsPayload(
        report_date=report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        columns=[
            BalanceAnalysisWorkbookColumn(**column)
            for column in decision_section.get("columns", [])
        ],
        rows=[
            _to_decision_item_status_row(row, latest_statuses)
            for row in decision_section.get("rows", [])
        ],
    )
    meta = build_formal_result_meta(
        trace_id=f"tr_balance_analysis_decision_items_{report_date}_{position_scope}_{currency_basis}",
        result_kind="balance-analysis.decision-items",
        cache_version=_resolve_balance_cache_version(build_lineage),
        source_version=_require_balance_lineage_value(
            build_lineage["source_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="source_version",
        ),
        rule_version=_require_balance_lineage_value(
            build_lineage["rule_version"] if build_lineage is not None else None,
            report_date=report_date,
            field_name="rule_version",
        ),
    )
    return build_formal_result_envelope(
        result_meta=meta,
        result_payload=payload.model_dump(mode="json"),
    )


def update_balance_analysis_decision_status(
    *,
    duckdb_path: str,
    governance_dir: str,
    update: BalanceAnalysisDecisionStatusUpdateRequest,
    updated_by: str,
) -> BalanceAnalysisDecisionStatusRecord:
    workbook, _build_lineage = _build_balance_workbook_payload(
        duckdb_path=duckdb_path,
        governance_dir=governance_dir,
        report_date=update.report_date,
        position_scope=update.position_scope,
        currency_basis=update.currency_basis,
    )
    valid_decision_keys = {
        _build_decision_key(row)
        for row in _extract_generated_decision_section(workbook).get("rows", [])
        if isinstance(row, dict)
    }
    if update.decision_key not in valid_decision_keys:
        raise ValueError(
            "Unknown balance-analysis decision_key for the requested report_date and filters."
        )

    record = BalanceAnalysisDecisionStatusRecord(
        decision_key=update.decision_key,
        status=update.status,
        updated_at=datetime.now(timezone.utc).isoformat(),
        updated_by=(updated_by or DEFAULT_BALANCE_DECISION_UPDATED_BY).strip()
        or DEFAULT_BALANCE_DECISION_UPDATED_BY,
        comment=update.comment,
    )
    BalanceAnalysisDecisionRepository(governance_dir).append_status(
        {
            "report_date": update.report_date,
            "position_scope": update.position_scope,
            "currency_basis": update.currency_basis,
            **record.model_dump(mode="json"),
        }
    )
    return record


def _to_zqtz_detail_row(row: dict[str, object]) -> BalanceAnalysisDetailRow:
    instrument_code = str(row["instrument_code"])
    return BalanceAnalysisDetailRow(
        source_family="zqtz",
        report_date=str(row["report_date"]),
        row_key=f"zqtz:{instrument_code}:{row['currency_basis']}:{row['position_scope']}",
        display_name=instrument_code,
        position_scope=str(row["position_scope"]),
        currency_basis=str(row["currency_basis"]),
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        market_value_amount=_as_decimal(row["market_value_amount"]),
        amortized_cost_amount=_as_decimal(row["amortized_cost_amount"]),
        accrued_interest_amount=_as_decimal(row["accrued_interest_amount"]),
        is_issuance_like=bool(row["is_issuance_like"]),
    )


def _to_tyw_detail_row(row: dict[str, object]) -> BalanceAnalysisDetailRow:
    position_id = str(row["position_id"])
    principal = _as_decimal(row["principal_amount"])
    accrued = _as_decimal(row["accrued_interest_amount"])
    return BalanceAnalysisDetailRow(
        source_family="tyw",
        report_date=str(row["report_date"]),
        row_key=f"tyw:{position_id}:{row['currency_basis']}:{row['position_scope']}",
        display_name=position_id,
        position_scope=str(row["position_scope"]),
        currency_basis=str(row["currency_basis"]),
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        market_value_amount=principal,
        amortized_cost_amount=principal,
        accrued_interest_amount=accrued,
        is_issuance_like=None,
    )


def _to_basis_breakdown_row(row: dict[str, object]) -> BalanceAnalysisBasisBreakdownRow:
    return BalanceAnalysisBasisBreakdownRow(
        source_family=str(row["source_family"]),  # type: ignore[arg-type]
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        position_scope=str(row["position_scope"]),  # type: ignore[arg-type]
        currency_basis=str(row["currency_basis"]),  # type: ignore[arg-type]
        detail_row_count=int(row["detail_row_count"]),
        market_value_amount=_as_decimal(row["market_value_amount"]),
        amortized_cost_amount=_as_decimal(row["amortized_cost_amount"]),
        accrued_interest_amount=_as_decimal(row["accrued_interest_amount"]),
    )


def _to_summary_table_row(row: dict[str, object]) -> BalanceAnalysisTableRow:
    return BalanceAnalysisTableRow(
        row_key=str(row["row_key"]),
        source_family=str(row["source_family"]),
        display_name=str(row["display_name"]),
        owner_name=str(row["owner_name"]),
        category_name=str(row["category_name"]),
        position_scope=str(row["position_scope"]),
        currency_basis=str(row["currency_basis"]),
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        detail_row_count=int(row["detail_row_count"]),
        market_value_amount=_as_decimal(row["market_value_amount"]),
        amortized_cost_amount=_as_decimal(row["amortized_cost_amount"]),
        accrued_interest_amount=_as_decimal(row["accrued_interest_amount"]),
    )


def _build_balance_workbook_payload(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"],
    currency_basis: Literal["native", "CNY"],
) -> tuple[dict[str, Any], dict[str, object] | None]:
    repo = BalanceAnalysisRepository(duckdb_path)
    if report_date not in repo.list_report_dates():
        raise ValueError(f"No balance-analysis data found for report_date={report_date}.")

    zqtz_native_rows = [
        _to_formal_zqtz_fact_row(row)
        for row in repo.fetch_formal_zqtz_rows(
            report_date=report_date,
            position_scope=position_scope,
            currency_basis="native",
        )
    ]
    tyw_native_rows = [
        _to_formal_tyw_fact_row(row)
        for row in repo.fetch_formal_tyw_rows(
            report_date=report_date,
            position_scope=position_scope,
            currency_basis="native",
        )
    ]
    zqtz_currency_rows = [
        _to_formal_zqtz_fact_row(row)
        for row in repo.fetch_formal_zqtz_rows(
            report_date=report_date,
            position_scope=position_scope,
            currency_basis="CNY",
        )
    ]
    workbook_mod = importlib.import_module("backend.app.core_finance.balance_analysis_workbook")
    workbook_mod = importlib.reload(workbook_mod)
    workbook = workbook_mod.build_balance_analysis_workbook_payload(
        report_date=zqtz_native_rows[0].report_date if zqtz_native_rows else tyw_native_rows[0].report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        zqtz_rows=zqtz_native_rows,
        tyw_rows=tyw_native_rows,
        zqtz_currency_rows=zqtz_currency_rows,
    )
    return workbook, resolve_completed_formal_build_lineage(
        governance_dir=governance_dir,
        cache_key=CACHE_KEY,
        job_name=BALANCE_ANALYSIS_JOB_NAME,
        report_date=report_date,
    )


def _extract_generated_decision_section(workbook: dict[str, Any]) -> dict[str, Any]:
    for section in workbook.get("tables", []):
        if str(section.get("section_kind")) == "decision_items":
            return dict(section)
    return {"columns": [], "rows": []}


def _build_decision_key(row: dict[str, object]) -> str:
    return "::".join(
        [
            str(row.get("rule_id") or "").strip(),
            str(row.get("source_section") or "").strip(),
            str(row.get("title") or "").strip(),
        ]
    )


def _default_pending_decision_status(decision_key: str) -> BalanceAnalysisDecisionStatusRecord:
    return BalanceAnalysisDecisionStatusRecord(
        decision_key=decision_key,
        status="pending",
        updated_at=None,
        updated_by=None,
        comment=None,
    )


def _to_decision_item_status_row(
    row: dict[str, object],
    latest_statuses: dict[str, dict[str, object]],
) -> BalanceAnalysisDecisionItemStatusRow:
    decision_key = _build_decision_key(row)
    latest_status = latest_statuses.get(decision_key)
    return BalanceAnalysisDecisionItemStatusRow(
        decision_key=decision_key,
        title=str(row["title"]),
        action_label=str(row["action_label"]),
        severity=str(row["severity"]),
        reason=str(row["reason"]),
        source_section=str(row["source_section"]),
        rule_id=str(row["rule_id"]),
        rule_version=str(row["rule_version"]),
        latest_status=(
            BalanceAnalysisDecisionStatusRecord(**latest_status)
            if latest_status is not None
            else _default_pending_decision_status(decision_key)
        ),
    )


def _to_formal_zqtz_fact_row(row: dict[str, object]) -> FormalZqtzBalanceFactRow:
    maturity_date = str(row.get("maturity_date") or "").strip()
    return FormalZqtzBalanceFactRow(
        report_date=_parse_date(str(row["report_date"])),
        instrument_code=str(row["instrument_code"]),
        instrument_name=str(row.get("instrument_name") or ""),
        portfolio_name=str(row.get("portfolio_name") or ""),
        cost_center=str(row.get("cost_center") or ""),
        account_category=str(row.get("account_category") or ""),
        asset_class=str(row.get("asset_class") or ""),
        bond_type=str(row.get("bond_type") or ""),
        issuer_name=str(row.get("issuer_name") or ""),
        industry_name=str(row.get("industry_name") or ""),
        rating=str(row.get("rating") or ""),
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        position_scope=str(row["position_scope"]),
        currency_basis=str(row["currency_basis"]),
        currency_code=str(row.get("currency_code") or ""),
        face_value_amount=_as_decimal(row["face_value_amount"]),
        market_value_amount=_as_decimal(row["market_value_amount"]),
        amortized_cost_amount=_as_decimal(row["amortized_cost_amount"]),
        accrued_interest_amount=_as_decimal(row["accrued_interest_amount"]),
        coupon_rate=_optional_decimal(row.get("coupon_rate")),
        ytm_value=_optional_decimal(row.get("ytm_value")),
        maturity_date=_parse_date(maturity_date) if maturity_date else None,
        interest_mode=str(row.get("interest_mode") or ""),
        is_issuance_like=bool(row["is_issuance_like"]),
        overdue_principal_days=_optional_nonnegative_int(row.get("overdue_principal_days")),
        overdue_interest_days=_optional_nonnegative_int(row.get("overdue_interest_days")),
        value_date=_parse_date(value_date_raw)
        if (value_date_raw := str(row.get("value_date") or "").strip())
        else None,
        customer_attribute=str(row.get("customer_attribute") or ""),
        source_version=str(row.get("source_version") or ""),
        rule_version=str(row.get("rule_version") or ""),
        ingest_batch_id=str(row.get("ingest_batch_id") or ""),
        trace_id=str(row.get("trace_id") or ""),
    )


def _to_formal_tyw_fact_row(row: dict[str, object]) -> FormalTywBalanceFactRow:
    maturity_date = str(row.get("maturity_date") or "").strip()
    return FormalTywBalanceFactRow(
        report_date=_parse_date(str(row["report_date"])),
        position_id=str(row["position_id"]),
        product_type=str(row.get("product_type") or ""),
        position_side=str(row.get("position_side") or ""),
        counterparty_name=str(row.get("counterparty_name") or ""),
        account_type=str(row.get("account_type") or ""),
        special_account_type=str(row.get("special_account_type") or ""),
        core_customer_type=str(row.get("core_customer_type") or ""),
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        position_scope=str(row["position_scope"]),
        currency_basis=str(row["currency_basis"]),
        currency_code=str(row.get("currency_code") or ""),
        principal_amount=_as_decimal(row["principal_amount"]),
        accrued_interest_amount=_as_decimal(row["accrued_interest_amount"]),
        funding_cost_rate=_optional_decimal(row.get("funding_cost_rate")),
        maturity_date=_parse_date(maturity_date) if maturity_date else None,
        source_version=str(row.get("source_version") or ""),
        rule_version=str(row.get("rule_version") or ""),
        ingest_batch_id=str(row.get("ingest_batch_id") or ""),
        trace_id=str(row.get("trace_id") or ""),
    )


def _build_summary_rows(details: list[BalanceAnalysisDetailRow]) -> list[BalanceAnalysisSummaryRow]:
    grouped: dict[tuple[str, str, str], dict[str, Decimal | int]] = {}
    for row in details:
        key = (row.source_family, row.position_scope, row.currency_basis)
        bucket = grouped.setdefault(
            key,
            {
                "row_count": 0,
                "market_value_amount": Decimal("0"),
                "amortized_cost_amount": Decimal("0"),
                "accrued_interest_amount": Decimal("0"),
            },
        )
        bucket["row_count"] = int(bucket["row_count"]) + 1
        bucket["market_value_amount"] = _as_decimal(bucket["market_value_amount"]) + row.market_value_amount
        bucket["amortized_cost_amount"] = _as_decimal(bucket["amortized_cost_amount"]) + row.amortized_cost_amount
        bucket["accrued_interest_amount"] = _as_decimal(bucket["accrued_interest_amount"]) + row.accrued_interest_amount

    return [
        BalanceAnalysisSummaryRow(
            source_family=source_family,
            position_scope=position_scope,
            currency_basis=currency_basis,
            row_count=int(values["row_count"]),
            market_value_amount=_as_decimal(values["market_value_amount"]),
            amortized_cost_amount=_as_decimal(values["amortized_cost_amount"]),
            accrued_interest_amount=_as_decimal(values["accrued_interest_amount"]),
        )
        for (source_family, position_scope, currency_basis), values in sorted(grouped.items())
    ]


def _combine_lineage_values(rows: list[dict[str, object]], field_name: str) -> str:
    values = sorted(
        {
            str(row.get(field_name) or "").strip()
            for row in rows
            if str(row.get(field_name) or "").strip()
        }
    )
    return "__".join(values) or "sv_balance_analysis_empty"


def _build_balance_summary_csv(
    rows: list[BalanceAnalysisTableRow],
    *,
    report_date: str,
    source_version: str,
    rule_version: str,
) -> str:
    output = StringIO()
    fieldnames = [
        "row_key",
        "source_family",
        "display_name",
        "owner_name",
        "category_name",
        "position_scope",
        "currency_basis",
        "invest_type_std",
        "accounting_basis",
        "detail_row_count",
        "market_value_amount",
        "amortized_cost_amount",
        "accrued_interest_amount",
        "report_date",
        "source_version",
        "rule_version",
    ]
    writer = csv.DictWriter(output, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                **row.model_dump(mode="json"),
                "report_date": report_date,
                "source_version": source_version,
                "rule_version": rule_version,
            }
        )
    return output.getvalue()


def _build_balance_analysis_workbook_xlsx_bytes(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "概览"
    _write_workbook_cards_sheet(overview_sheet, payload.get("cards") or [])

    for sheet_name, candidate_keys in EXPORT_WORKBOOK_TABLES:
        table = _pick_workbook_export_table(payload, *candidate_keys)
        sheet = workbook.create_sheet(title=sheet_name)
        _write_workbook_table_sheet(sheet, table)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _pick_workbook_export_table(payload: dict[str, Any], *candidate_keys: str) -> dict[str, Any]:
    table_map = {
        str(table.get("key")): table
        for table in list(payload.get("tables") or [])
        if isinstance(table, dict)
    }
    for key in candidate_keys:
        if key in table_map:
            return table_map[key]
    raise RuntimeError(
        "Balance-analysis workbook export table unavailable. "
        f"Expected one of {candidate_keys!r}."
    )


def _write_workbook_cards_sheet(sheet: Worksheet, cards: list[dict[str, Any]]) -> None:
    sheet.append(["label", "value"])
    _style_header_row(sheet, column_count=2)
    for card in cards:
        value = _coerce_excel_value(card.get("value"))
        sheet.append([card.get("label"), value])
        _style_numeric_cell(sheet.cell(row=sheet.max_row, column=2), value, column_key="value")
    _autosize_sheet_columns(sheet)


def _write_workbook_table_sheet(sheet: Worksheet, table: dict[str, Any]) -> None:
    columns = list(table.get("columns") or [])
    rows = list(table.get("rows") or [])
    headers = [str(column.get("label") or "") for column in columns]
    column_keys = [str(column.get("key") or "") for column in columns]

    sheet.append(headers)
    _style_header_row(sheet, column_count=len(headers))

    for row in rows:
        values = []
        coerced_values: list[object] = []
        for column_key in column_keys:
            coerced = _coerce_excel_value(row.get(column_key))
            values.append(coerced)
            coerced_values.append(coerced)
        sheet.append(values)
        for column_index, (column_key, value) in enumerate(zip(column_keys, coerced_values), start=1):
            _style_numeric_cell(sheet.cell(row=sheet.max_row, column=column_index), value, column_key=column_key)

    _autosize_sheet_columns(sheet)


def _style_header_row(sheet: Worksheet, *, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        sheet.cell(row=1, column=column_index).font = EXCEL_HEADER_FONT


def _style_numeric_cell(cell, value: object, *, column_key: str) -> None:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return
    cell.alignment = EXCEL_RIGHT_ALIGNMENT
    if _is_amount_column(column_key):
        cell.number_format = EXCEL_AMOUNT_FORMAT
        return
    if isinstance(value, int):
        cell.number_format = EXCEL_INTEGER_FORMAT
        return
    if isinstance(value, Decimal) and value == value.to_integral_value():
        cell.number_format = EXCEL_INTEGER_FORMAT
        return
    cell.number_format = EXCEL_NUMBER_FORMAT


def _is_amount_column(column_key: str) -> bool:
    normalized = column_key.lower()
    return normalized == "value" or normalized.endswith("_amount") or normalized.endswith("_value")


def _coerce_excel_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return ""
        try:
            return Decimal(stripped)
        except InvalidOperation:
            return value
    return value


def _autosize_sheet_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max(len(value) for value in values) if values else 0
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 40)


def _resolve_balance_cache_version(lineage: dict[str, object] | None) -> str:
    if lineage is not None:
        resolved = str(lineage.get("cache_version") or "").strip()
        if resolved:
            return resolved
    return CACHE_VERSION


def _validate_balance_overview_filters(*, position_scope: str, currency_basis: str) -> None:
    if position_scope not in ALLOWED_BALANCE_POSITION_SCOPES:
        raise ValueError(
            f"Unsupported balance-analysis position_scope={position_scope}. "
            f"Expected one of {sorted(ALLOWED_BALANCE_POSITION_SCOPES)}."
        )
    if currency_basis not in ALLOWED_BALANCE_CURRENCY_BASES:
        raise ValueError(
            f"Unsupported balance-analysis currency_basis={currency_basis}. "
            f"Expected one of {sorted(ALLOWED_BALANCE_CURRENCY_BASES)}."
        )


def _require_balance_lineage_value(value: object, *, report_date: str, field_name: str) -> str:
    resolved = str(value or "").strip()
    if not resolved:
        raise RuntimeError(
            f"Canonical balance-analysis {field_name} unavailable for report_date={report_date}."
        )
    return resolved


def _refresh_trigger_lock(*, report_date: str) -> LockDefinition:
    return LockDefinition(
        key=f"{BALANCE_ANALYSIS_LOCK.key}:{report_date}:trigger",
        ttl_seconds=30,
    )


def _load_refresh_run_records(settings: Settings) -> list[dict[str, object]]:
    return [
        record
        for record in GovernanceRepository(base_dir=settings.governance_path).read_all(CACHE_BUILD_RUN_STREAM)
        if str(record.get("cache_key")) == CACHE_KEY
        and str(record.get("job_name")) == BALANCE_ANALYSIS_JOB_NAME
    ]


def _latest_inflight_refresh(
    settings: Settings,
    *,
    report_date: str,
) -> dict[str, object] | None:
    by_run_id: dict[str, dict[str, object]] = {}
    for record in _load_refresh_run_records(settings):
        if str(record.get("report_date")) != report_date:
            continue
        by_run_id[str(record.get("run_id"))] = record
    stale_records: list[dict[str, object]] = []
    for record in reversed(list(by_run_id.values())):
        if str(record.get("status")) in IN_FLIGHT_STATUSES:
            if _is_stale_inflight_record(record):
                stale_records.append(record)
                continue
            return record
    for record in stale_records:
        _mark_stale_inflight_run(
            settings=settings,
            run_id=str(record.get("run_id")),
            report_date=report_date,
            error_message="Marked stale balance-analysis refresh run as failed.",
        )
    return None


def _is_stale_inflight_record(record: dict[str, object]) -> bool:
    for field_name in ("started_at", "queued_at", "created_at"):
        raw_value = str(record.get(field_name) or "").strip()
        if not raw_value:
            continue
        timestamp = _parse_timestamp(raw_value)
        return datetime.now(timezone.utc) - timestamp > STALE_IN_FLIGHT_AFTER
    return True


def _parse_timestamp(raw_value: str) -> datetime:
    normalized = raw_value.replace("Z", "+00:00") if raw_value.endswith("Z") else raw_value
    parsed = datetime.fromisoformat(normalized)
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def _record_dispatch_failure(
    *,
    settings: Settings,
    run_id: str,
    report_date: str,
    error_message: str,
) -> None:
    GovernanceRepository(base_dir=settings.governance_path).append(
        CACHE_BUILD_RUN_STREAM,
        {
            "run_id": run_id,
            "job_name": BALANCE_ANALYSIS_JOB_NAME,
            "status": "failed",
            "cache_key": CACHE_KEY,
            "lock": BALANCE_ANALYSIS_LOCK.key,
            "source_version": "sv_balance_analysis_failed",
            "vendor_version": "vv_none",
            "report_date": report_date,
            "error_message": error_message,
        },
    )


def _mark_stale_inflight_run(
    *,
    settings: Settings,
    run_id: str,
    report_date: str,
    error_message: str,
) -> None:
    GovernanceRepository(base_dir=settings.governance_path).append(
        CACHE_BUILD_RUN_STREAM,
        {
            "run_id": run_id,
            "job_name": BALANCE_ANALYSIS_JOB_NAME,
            "status": "failed",
            "cache_key": CACHE_KEY,
            "lock": BALANCE_ANALYSIS_LOCK.key,
            "source_version": "sv_balance_analysis_stale",
            "vendor_version": "vv_none",
            "report_date": report_date,
            "error_message": error_message,
            "finished_at": datetime.now(timezone.utc).isoformat(),
        },
    )


def _build_run_id() -> str:
    return f"{BALANCE_ANALYSIS_JOB_NAME}:{datetime.now(timezone.utc).isoformat()}"


def _as_decimal(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _optional_decimal(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    return _as_decimal(value)


def _optional_nonnegative_int(value: object) -> int:
    if value in (None, ""):
        return 0
    try:
        return max(0, int(Decimal(str(value))))
    except (InvalidOperation, ValueError, TypeError):
        return 0


def _parse_date(raw_value: str):
    return datetime.strptime(raw_value, "%Y-%m-%d").date()
