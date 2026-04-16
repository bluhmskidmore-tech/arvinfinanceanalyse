from __future__ import annotations

import hashlib
import os
from calendar import monthrange
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from backend.app.repositories.governance_repo import SOURCE_MANIFEST_STREAM, GovernanceRepository
from backend.app.services.source_rules import describe_source_file


SUPPORTED_PNL_SOURCE_FAMILIES = ("pnl", "pnl_514", "pnl_516", "pnl_517")
MANIFEST_ELIGIBLE_STATUSES = {"completed", "rerun"}
PNL_SOURCE_RULE_VERSION = "rv_pnl_source_parse_v1"


@dataclass(slots=True, frozen=True)
class PnlSourceSnapshot:
    source_family: str
    report_date: str
    path: Path
    source_version: str
    ingest_batch_id: str
    created_at: str


@dataclass(slots=True, frozen=True)
class PnlRefreshInput:
    report_date: str
    is_month_end: bool
    fi_rows: list[dict[str, object]]
    nonstd_rows_by_type: dict[str, list[dict[str, object]]]


def resolve_pnl_data_input_root() -> Path:
    configured_root = os.getenv("MOSS_DATA_INPUT_ROOT")
    if configured_root:
        return Path(configured_root).expanduser()
    return Path(__file__).resolve().parents[3] / "data_input"


def load_latest_pnl_refresh_input(
    *,
    governance_dir: str | Path,
    data_root: str | Path | None = None,
    report_date: str | None = None,
) -> PnlRefreshInput:
    resolved_data_root = Path(data_root) if data_root is not None else resolve_pnl_data_input_root()
    manifest_candidates = _manifest_candidates(governance_dir)
    direct_candidates = _direct_candidates(resolved_data_root)

    if report_date is None:
        candidates = manifest_candidates if manifest_candidates else direct_candidates
        if not candidates:
            raise ValueError("No eligible pnl source files found for formal refresh.")
        target_report_date = max(candidate.report_date for candidate in candidates)
    else:
        target_report_date = report_date
        candidates = _merge_candidates_for_report_date(
            manifest_candidates=manifest_candidates,
            direct_candidates=direct_candidates,
            report_date=report_date,
        )
        if not candidates:
            raise ValueError(f"No eligible pnl source files found for report_date={report_date}.")

    selected_by_family = {
        family: _latest_candidate_for_family(candidates, family=family, report_date=target_report_date)
        for family in SUPPORTED_PNL_SOURCE_FAMILIES
    }

    fi_rows: list[dict[str, object]] = []
    nonstd_rows_by_type: dict[str, list[dict[str, object]]] = {}
    for family, snapshot in selected_by_family.items():
        if snapshot is None:
            continue
        if family == "pnl":
            fi_rows = _parse_fi_rows(snapshot)
            continue
        bucket = family.removeprefix("pnl_")
        nonstd_rows_by_type[bucket] = _parse_nonstd_rows(snapshot, bucket=bucket)

    if not fi_rows and not any(nonstd_rows_by_type.values()):
        raise ValueError(f"Latest pnl source bundle for report_date={target_report_date} contains no usable rows.")

    report_day = date.fromisoformat(target_report_date)
    return PnlRefreshInput(
        report_date=target_report_date,
        is_month_end=report_day.day == monthrange(report_day.year, report_day.month)[1],
        fi_rows=fi_rows,
        nonstd_rows_by_type={
            bucket: filtered_rows
            for bucket, rows in nonstd_rows_by_type.items()
            if (filtered_rows := _filter_nonstd_rows_for_report_month(rows, report_day))
        },
    )


def _merge_candidates_for_report_date(
    *,
    manifest_candidates: list[PnlSourceSnapshot],
    direct_candidates: list[PnlSourceSnapshot],
    report_date: str,
) -> list[PnlSourceSnapshot]:
    merged: list[PnlSourceSnapshot] = []
    for family in SUPPORTED_PNL_SOURCE_FAMILIES:
        family_manifest = [
            candidate
            for candidate in manifest_candidates
            if candidate.report_date == report_date and candidate.source_family == family
        ]
        family_direct = [
            candidate
            for candidate in direct_candidates
            if candidate.report_date == report_date and candidate.source_family == family
        ]
        if family_manifest:
            merged.extend(family_manifest)
        elif family_direct:
            merged.extend(family_direct)
    return merged


def _manifest_candidates(governance_dir: str | Path) -> list[PnlSourceSnapshot]:
    rows = GovernanceRepository(base_dir=governance_dir).read_all(SOURCE_MANIFEST_STREAM)
    snapshots: list[PnlSourceSnapshot] = []
    for row in rows:
        source_family = str(row.get("source_family", ""))
        archived_path = row.get("archived_path")
        if source_family not in SUPPORTED_PNL_SOURCE_FAMILIES:
            continue
        if str(row.get("status", "")) not in MANIFEST_ELIGIBLE_STATUSES:
            continue
        if archived_path in (None, ""):
            continue

        path = Path(str(archived_path))
        if not path.exists() or _is_processed_path(path):
            continue

        report_date = str(row.get("report_date") or "")
        if not report_date:
            metadata = describe_source_file(str(row.get("source_file") or path.name))
            report_date = str(metadata.report_date or "")
        if not report_date:
            continue

        snapshots.append(
            PnlSourceSnapshot(
                source_family=source_family,
                report_date=report_date,
                path=path,
                source_version=str(row.get("source_version") or "sv_pnl_source_missing"),
                ingest_batch_id=str(row.get("ingest_batch_id") or "ib_pnl_manifest"),
                created_at=str(row.get("created_at") or ""),
            )
        )
    return snapshots


def _direct_candidates(data_root: Path) -> list[PnlSourceSnapshot]:
    candidates: list[PnlSourceSnapshot] = []
    directory_specs = {
        "pnl": ("pnl", "*.xls"),
        "pnl_514": ("pnl_514", "*.xlsx"),
        "pnl_516": ("pnl_516", "*.xlsx"),
        "pnl_517": ("pnl_517", "*.xlsx"),
    }

    for family, (directory_name, pattern) in directory_specs.items():
        source_dir = data_root / directory_name
        if not source_dir.exists():
            continue
        for path in sorted(source_dir.glob(pattern)):
            metadata = describe_source_file(path.name)
            if metadata.source_family != family or metadata.report_date is None:
                continue
            stat = path.stat()
            candidates.append(
                PnlSourceSnapshot(
                    source_family=family,
                    report_date=metadata.report_date,
                    path=path,
                    source_version=_build_source_version(path),
                    ingest_batch_id="ib_pnl_direct",
                    created_at=str(stat.st_mtime_ns),
                )
            )
    return candidates


def _latest_candidate_for_family(
    candidates: list[PnlSourceSnapshot],
    *,
    family: str,
    report_date: str,
) -> PnlSourceSnapshot | None:
    matches = [
        candidate
        for candidate in candidates
        if candidate.source_family == family and candidate.report_date == report_date
    ]
    if not matches:
        return None
    return max(
        matches,
        key=lambda candidate: (
            candidate.created_at,
            candidate.ingest_batch_id,
            str(candidate.path),
        ),
    )


def _parse_fi_rows(snapshot: PnlSourceSnapshot) -> list[dict[str, object]]:
    metadata = describe_source_file(snapshot.path.name)
    report_date = snapshot.report_date or metadata.report_date
    workbook = xlrd.open_workbook(str(snapshot.path))
    sheet = workbook.sheet_by_index(0)
    headers = [str(sheet.cell_value(0, column)).strip() for column in range(sheet.ncols)]
    rows: list[dict[str, object]] = []

    for row_index in range(1, sheet.nrows):
        raw_row = {
            headers[column]: sheet.cell_value(row_index, column)
            for column in range(sheet.ncols)
            if headers[column]
        }
        instrument_code = _cell_text(raw_row.get("债券代码"))
        if not instrument_code:
            continue

        rows.append(
            {
                "report_date": report_date,
                "instrument_code": instrument_code,
                "portfolio_name": _cell_text(raw_row.get("投资组合")),
                "cost_center": _cell_text(raw_row.get("成本中心")),
                "invest_type_raw": _cell_text(raw_row.get("投资类型")),
                "interest_income_514": _to_decimal(raw_row.get("利息514")),
                # The FI source column is named T损益516; the formal thin slice uses the governed sign convention.
                "fair_value_change_516": _to_decimal(raw_row.get("T损益516")) * Decimal("-1"),
                "capital_gain_517": _to_decimal(raw_row.get("投资收益517")),
                "manual_adjustment": Decimal("0"),
                # The source file's 币种 describes instrument currency. The current formal PnL slice stores basis,
                # so only explicit CNX markers stay CNX and all other rows land in the CNY fact partition.
                "currency_basis": _resolve_currency_basis(_cell_text(raw_row.get("币种"))),
                "source_version": snapshot.source_version,
                "rule_version": PNL_SOURCE_RULE_VERSION,
                "ingest_batch_id": snapshot.ingest_batch_id,
                "trace_id": f"{snapshot.path.name}:fi:{len(rows) + 1}",
            }
        )
    return rows


def _parse_nonstd_rows(snapshot: PnlSourceSnapshot, *, bucket: str) -> list[dict[str, object]]:
    workbook = load_workbook(snapshot.path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    headers = [
        "" if value is None else str(value).strip()
        for value in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
    ]
    rows: list[dict[str, object]] = []

    try:
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            raw_row = {
                headers[index]: values[index]
                for index in range(min(len(headers), len(values)))
                if headers[index]
            }
            account_code = _cell_text(raw_row.get("科目号") or raw_row.get("科目代码") or raw_row.get("会计科目"))
            asset_code = _cell_text(raw_row.get("资产代码"))
            if not account_code and not asset_code:
                continue

            rows.append(
                {
                    "voucher_date": _cell_text(raw_row.get("账务日期")),
                    "account_code": account_code,
                    "asset_code": asset_code,
                    "portfolio_name": _cell_text(raw_row.get("投资组合")),
                    "cost_center": _cell_text(raw_row.get("成本中心")),
                    "dc_flag": _cell_text(raw_row.get("借贷标识") or raw_row.get("方向")),
                    "event_type": _cell_text(raw_row.get("会计事件")),
                    "raw_amount": _to_decimal(raw_row.get("金额") if raw_row.get("金额") not in (None, "") else raw_row.get("AMOUNT")),
                    "source_file": snapshot.path.name,
                    "source_version": snapshot.source_version,
                    "rule_version": PNL_SOURCE_RULE_VERSION,
                    "ingest_batch_id": snapshot.ingest_batch_id,
                    "trace_id": f"{snapshot.path.name}:{bucket}:{len(rows) + 1}",
                }
            )
    finally:
        workbook.close()

    return rows


def _resolve_currency_basis(raw_currency: str) -> str:
    normalized = raw_currency.strip().upper()
    if normalized in {"CNX", "综本"}:
        return "CNX"
    return "CNY"


def _filter_nonstd_rows_for_report_month(
    rows: list[dict[str, object]],
    report_day: date,
) -> list[dict[str, object]]:
    filtered: list[dict[str, object]] = []
    for row in rows:
        voucher_date = date.fromisoformat(str(row["voucher_date"]))
        if voucher_date.year != report_day.year or voucher_date.month != report_day.month:
            continue
        filtered.append(row)
    return filtered


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_decimal(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    return Decimal(str(value))


def _is_processed_path(path: Path) -> bool:
    return any(part.lower() == "processed" for part in path.parts)


def _build_source_version(path: Path) -> str:
    stat = path.stat()
    seed = f"{path.name}:{stat.st_size}:{stat.st_mtime_ns}"
    digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12]
    return f"sv_pnl_{digest}"
