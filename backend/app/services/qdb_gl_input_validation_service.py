from __future__ import annotations

import hashlib
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Final

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from backend.app.schemas.qdb_gl_contract import (
    QdbGlBaselineBinding,
    QdbGlBaselineValidationEvidence,
    QdbGlContractCheck,
    QdbGlContractFinding,
    QdbGlLineage,
)


RULE_VERSION = "rv_qdb_gl_input_contract_v1"

LEDGER_FILE_PATTERN: Final = re.compile(r"^总账对账(?P<report_month>\d{6})\.xlsx$", re.IGNORECASE)
AVERAGE_FILE_PATTERN: Final = re.compile(r"^日均(?P<report_month>\d{6})\.xlsx$", re.IGNORECASE)

LEDGER_HEADERS: Final[list[str]] = [
    "组合科目代码",
    "组合科目名称",
    "币种",
    "期初余额",
    "本期借方",
    "本期贷方",
    "期末余额",
]
LEDGER_CANONICAL_SHEETS: Final[tuple[tuple[str, str], ...]] = (
    ("综本", "CNX"),
    ("人民币", "CNY"),
)

AVERAGE_BLOCK_HEADERS: Final[list[str]] = ["币种", "科目", "科目日均余额"]
AVERAGE_CANONICAL_SHEETS: Final[tuple[str, ...]] = ("年", "月")
ALLOWED_CURRENCY_GROUPS: Final[set[str]] = {"CNX", "CNY"}
RECONCILIATION_TOLERANCE: Final = Decimal("0.01")
CHECK_ORDER: Final[tuple[str, ...]] = (
    "source_binding",
    "header_row",
    "row_shape",
    "required_raw_fields",
    "account_code_text_preserved",
    "currency_grouping",
    "reconciliation_contract",
)


def discover_qdb_gl_baseline_bindings(source_dir: str | Path) -> list[QdbGlBaselineBinding]:
    source_dir = Path(source_dir)
    if not source_dir.exists():
        return []

    bindings: list[QdbGlBaselineBinding] = []
    for path in sorted(source_dir.glob("*.xlsx")):
        binding = bind_qdb_gl_baseline_source(path)
        if binding is not None:
            bindings.append(binding)
    return sorted(bindings, key=lambda item: (item.source_kind, item.report_month, item.path.name))


def bind_qdb_gl_baseline_source(path: str | Path) -> QdbGlBaselineBinding | None:
    path = Path(path)
    match = LEDGER_FILE_PATTERN.match(path.name)
    if match is not None:
        return QdbGlBaselineBinding(
            source_kind="ledger_reconciliation",
            report_month=match.group("report_month"),
            path=path,
            source_version=_build_source_version(path),
        )

    match = AVERAGE_FILE_PATTERN.match(path.name)
    if match is not None:
        return QdbGlBaselineBinding(
            source_kind="average_balance",
            report_month=match.group("report_month"),
            path=path,
            source_version=_build_source_version(path),
        )

    return None


def validate_qdb_gl_baseline_source(path: str | Path) -> QdbGlBaselineValidationEvidence:
    path = Path(path)
    binding = bind_qdb_gl_baseline_source(path)
    source_kind = binding.source_kind if binding is not None else "unknown"
    report_month = binding.report_month if binding is not None else None
    source_version = binding.source_version if binding is not None else _build_source_version(path)
    trace_id = _build_trace_id(source_kind, source_version)
    checks = _build_default_checks(reconciliation_not_applicable=(source_kind == "average_balance"))

    if binding is None:
        _record_failure(
            checks,
            "source_binding",
            message=f"File name {path.name} does not match the QDB GL baseline source-binding contract.",
            cell_ref=None,
        )
        return _build_validation_evidence(
            binding_status="rejected",
            source_kind=source_kind,
            report_month=report_month,
            source_version=source_version,
            trace_id=trace_id,
            bound_currency_groups=[],
            lineage=QdbGlLineage(
                source_file=path.name,
                source_kind=source_kind,
                report_month=report_month,
                source_version=source_version,
                rule_version=RULE_VERSION,
                trace_id=trace_id,
                sheet_names=[],
            ),
            checks=checks,
        )

    workbook = None
    sheet_names: list[str] = []
    bound_currency_groups: set[str] = set()

    try:
        workbook = load_workbook(binding.path, read_only=True, data_only=True)
        checks["source_binding"].status_label = "pass"
        if binding.source_kind == "ledger_reconciliation":
            sheet_names, bound_currency_groups = _validate_ledger_workbook(workbook, checks)
        else:
            sheet_names, bound_currency_groups = _validate_average_workbook(workbook, checks)
    except Exception as exc:
        _record_failure(
            checks,
            "source_binding",
            message=f"Failed to open workbook: {exc}",
        )
    finally:
        if workbook is not None:
            workbook.close()

    return _build_validation_evidence(
        binding_status="bound",
        source_kind=binding.source_kind,
        report_month=binding.report_month,
        source_version=binding.source_version,
        trace_id=trace_id,
        bound_currency_groups=sorted(bound_currency_groups),
        lineage=QdbGlLineage(
            source_file=binding.path.name,
            source_kind=binding.source_kind,
            report_month=binding.report_month,
            source_version=binding.source_version,
            rule_version=RULE_VERSION,
            trace_id=trace_id,
            sheet_names=sheet_names,
        ),
        checks=checks,
    )


def _validate_ledger_workbook(workbook, checks: dict[str, QdbGlContractCheck]) -> tuple[list[str], set[str]]:
    sheet_names: list[str] = []
    bound_currency_groups: set[str] = set()

    for sheet_name, expected_currency in LEDGER_CANONICAL_SHEETS:
        if sheet_name not in workbook.sheetnames:
            _record_failure(
                checks,
                "source_binding",
                message=f"Missing canonical sheet {sheet_name} required by the ledger baseline binding contract.",
                sheet_name=sheet_name,
            )
            continue

        sheet_names.append(sheet_name)
        worksheet = workbook[sheet_name]
        _validate_ledger_header(worksheet, checks)
        _validate_ledger_rows(worksheet, expected_currency=expected_currency, checks=checks, bound_currency_groups=bound_currency_groups)

    if bound_currency_groups != ALLOWED_CURRENCY_GROUPS:
        _record_failure(
            checks,
            "currency_grouping",
            message="Ledger workbook must expose both CNX and CNY canonical currency groups.",
        )

    return sheet_names, bound_currency_groups


def _validate_ledger_header(worksheet, checks: dict[str, QdbGlContractCheck]) -> None:
    header_row = next(worksheet.iter_rows(min_row=6, max_row=6, values_only=True), tuple())
    for column_index, expected_header in enumerate(LEDGER_HEADERS, start=1):
        actual = _normalize_text(header_row[column_index - 1] if len(header_row) >= column_index else None)
        if actual != expected_header:
            _record_failure(
                checks,
                "header_row",
                message=f"Expected header {expected_header}, got {actual or '<blank>'}.",
                sheet_name=worksheet.title,
                row_locator=6,
                cell_ref=f"{get_column_letter(column_index)}6",
            )


def _validate_ledger_rows(
    worksheet,
    *,
    expected_currency: str,
    checks: dict[str, QdbGlContractCheck],
    bound_currency_groups: set[str],
) -> None:
    for row_index, row in enumerate(worksheet.iter_rows(min_row=7, values_only=True), start=7):
        row_values = list(row)
        if len(row_values) < 7:
            row_values.extend([None] * (7 - len(row_values)))
        if all(_is_blank(value) for value in row_values[:7]):
            continue

        account_code_value, account_name, currency_value, beginning_value, debit_value, credit_value, ending_value = row_values[:7]
        required_cells = {
            "A": account_code_value,
            "B": account_name,
            "C": currency_value,
            "D": beginning_value,
            "E": debit_value,
            "F": credit_value,
            "G": ending_value,
        }
        for column_letter, value in required_cells.items():
            if _is_blank(value):
                _record_failure(
                    checks,
                    "row_shape",
                    message="Ledger row must expose the 7-column core envelope before field-level validation.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{column_letter}{row_index}",
                )
                _record_failure(
                    checks,
                    "required_raw_fields",
                    message=f"Ledger row is missing required raw field {column_letter}.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{column_letter}{row_index}",
                )

        account_code = _normalize_account_code(account_code_value)
        if account_code is None:
            _record_failure(
                checks,
                "account_code_text_preserved",
                message="Account code must remain a digit-only text value without scientific notation or fractional loss.",
                sheet_name=worksheet.title,
                row_locator=row_index,
                cell_ref=f"A{row_index}",
            )

        currency_code = _normalize_currency(currency_value)
        if currency_code is None:
            _record_failure(
                checks,
                "currency_grouping",
                message="Currency code must be CNX or CNY.",
                sheet_name=worksheet.title,
                row_locator=row_index,
                cell_ref=f"C{row_index}",
            )
        else:
            bound_currency_groups.add(currency_code)
            if currency_code != expected_currency:
                _record_failure(
                    checks,
                    "currency_grouping",
                    message=f"Ledger row currency {currency_code} does not match canonical sheet currency {expected_currency}.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"C{row_index}",
                )

        beginning_amount = _to_decimal(beginning_value)
        debit_amount = _to_decimal(debit_value)
        credit_amount = _to_decimal(credit_value)
        ending_amount = _to_decimal(ending_value)
        if None not in {beginning_amount, debit_amount, credit_amount, ending_amount}:
            delta = beginning_amount + debit_amount - credit_amount - ending_amount
            if abs(delta) > RECONCILIATION_TOLERANCE:
                _record_failure(
                    checks,
                    "reconciliation_contract",
                    message="Ledger row failed the baseline reconciliation contract: 期初余额 + 本期借方 - 本期贷方 must equal 期末余额.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"G{row_index}",
                )


def _validate_average_workbook(workbook, checks: dict[str, QdbGlContractCheck]) -> tuple[list[str], set[str]]:
    sheet_names: list[str] = []
    bound_currency_groups: set[str] = set()

    for sheet_name in AVERAGE_CANONICAL_SHEETS:
        if sheet_name not in workbook.sheetnames:
            _record_failure(
                checks,
                "source_binding",
                message=f"Missing canonical sheet {sheet_name} required by the average baseline binding contract.",
                sheet_name=sheet_name,
            )
            continue

        sheet_names.append(sheet_name)
        worksheet = workbook[sheet_name]
        _validate_average_header(worksheet, checks)
        _validate_average_rows(worksheet, checks=checks, bound_currency_groups=bound_currency_groups)

    if bound_currency_groups != ALLOWED_CURRENCY_GROUPS:
        _record_failure(
            checks,
            "currency_grouping",
            message="Average workbook must expose both CNX and CNY currency groups.",
        )

    return sheet_names, bound_currency_groups


def _validate_average_header(worksheet, checks: dict[str, QdbGlContractCheck]) -> None:
    header_row = list(next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True), tuple()))
    if not header_row:
        _record_failure(
            checks,
            "header_row",
            message="Average workbook header row is missing.",
            sheet_name=worksheet.title,
            row_locator=3,
        )
        return

    _parse_average_block_specs(header_row, checks, worksheet.title)


def _validate_average_rows(
    worksheet,
    *,
    checks: dict[str, QdbGlContractCheck],
    bound_currency_groups: set[str],
) -> None:
    for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
        row_values = list(row)
        if all(_is_blank(value) for value in row_values):
            continue

        recognized_tuple = False
        column_index = 0
        while column_index < len(row_values):
            currency_value = row_values[column_index]
            currency_code = _normalize_currency(currency_value)
            if currency_code is None:
                column_index += 1
                continue

            account_code_value = row_values[column_index + 1] if len(row_values) > column_index + 1 else None
            balance_value = row_values[column_index + 2] if len(row_values) > column_index + 2 else None
            recognized_tuple = True

            if _is_blank(account_code_value):
                _record_failure(
                    checks,
                    "row_shape",
                    message="Average canonical tuple is missing account_code_raw after a valid currency marker.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{get_column_letter(column_index + 2)}{row_index}",
                )
                _record_failure(
                    checks,
                    "required_raw_fields",
                    message="Average canonical tuple is missing required raw field account_code_raw.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{get_column_letter(column_index + 2)}{row_index}",
                )
                column_index += 1
                continue

            account_code = _normalize_account_code(account_code_value)
            if account_code is None:
                _record_failure(
                    checks,
                    "account_code_text_preserved",
                    message="Average workbook account code must remain a digit-only text value without scientific notation or fractional loss.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{get_column_letter(column_index + 2)}{row_index}",
                )

            if _is_blank(balance_value):
                _record_failure(
                    checks,
                    "row_shape",
                    message="Average canonical tuple is missing avg_balance_raw after currency/account_code.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{get_column_letter(column_index + 3)}{row_index}",
                )
                _record_failure(
                    checks,
                    "required_raw_fields",
                    message="Average canonical tuple is missing required raw field avg_balance_raw.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{get_column_letter(column_index + 3)}{row_index}",
                )
                column_index += 2
                continue

            if _to_decimal(balance_value) is None:
                _record_failure(
                    checks,
                    "row_shape",
                    message="Average workbook balance field must be numeric.",
                    sheet_name=worksheet.title,
                    row_locator=row_index,
                    cell_ref=f"{get_column_letter(column_index + 3)}{row_index}",
                )
                column_index += 3
                continue

            bound_currency_groups.add(currency_code)
            column_index += 3



def _build_default_checks(*, reconciliation_not_applicable: bool) -> dict[str, QdbGlContractCheck]:
    checks = {
        check_id: QdbGlContractCheck(check_id=check_id, status_label="pass")
        for check_id in CHECK_ORDER
    }
    if reconciliation_not_applicable:
        checks["reconciliation_contract"].status_label = "not_applicable"
    return checks


def _parse_average_block_specs(
    header_row: list[object],
    checks: dict[str, QdbGlContractCheck],
    sheet_name: str,
) -> list[tuple[int, int]]:
    block_specs: list[tuple[int, int]] = []
    column_index = 0
    row_length = len(header_row)

    while column_index < row_length:
        if all(_is_blank(value) for value in header_row[column_index:]):
            break

        start_column = column_index + 1
        header_values = [
            _normalize_text(header_row[column_index + offset] if row_length > column_index + offset else None)
            for offset in range(3)
        ]
        for offset, expected_header in enumerate(AVERAGE_BLOCK_HEADERS):
            actual = header_values[offset]
            if actual != expected_header:
                _record_failure(
                    checks,
                    "header_row",
                    message=f"Average block header mismatch at offset {offset + 1}: expected {expected_header}, got {actual or '<blank>'}.",
                    sheet_name=sheet_name,
                    row_locator=3,
                    cell_ref=f"{get_column_letter(start_column + offset)}3",
                )

        if row_length <= column_index + 3:
            block_specs.append((start_column, 3))
            column_index += 3
            continue

        spacer_value = _normalize_text(header_row[column_index + 3])
        if spacer_value == "":
            block_specs.append((start_column, 4))
            column_index += 4
            continue

        _record_failure(
            checks,
            "row_shape",
            message="Average block spacer column must stay blank when present.",
            sheet_name=sheet_name,
            row_locator=3,
            cell_ref=f"{get_column_letter(start_column + 3)}3",
        )
        block_specs.append((start_column, 4))
        column_index += 4

    return block_specs


def _build_validation_evidence(
    *,
    binding_status: str,
    source_kind: str,
    report_month: str | None,
    source_version: str,
    trace_id: str,
    bound_currency_groups: list[str],
    lineage: QdbGlLineage,
    checks: dict[str, QdbGlContractCheck],
) -> QdbGlBaselineValidationEvidence:
    ordered_checks = [checks[check_id] for check_id in CHECK_ORDER]
    admissible = all(check.status_label in {"pass", "not_applicable"} for check in ordered_checks)
    if binding_status != "bound":
        admissible = False

    return QdbGlBaselineValidationEvidence(
        binding_status=binding_status,
        source_kind=source_kind,
        report_month=report_month,
        admissible=admissible,
        status_label="pass" if admissible else "fail",
        source_version=source_version,
        rule_version=RULE_VERSION,
        trace_id=trace_id,
        bound_currency_groups=bound_currency_groups,
        lineage=lineage,
        checks=ordered_checks,
    )


def _record_failure(
    checks: dict[str, QdbGlContractCheck],
    check_id: str,
    *,
    message: str,
    sheet_name: str | None = None,
    row_locator: int | None = None,
    cell_ref: str | None = None,
) -> None:
    check = checks[check_id]
    if check.status_label != "not_applicable":
        check.status_label = "fail"
    check.findings.append(
        QdbGlContractFinding(
            message=message,
            sheet_name=sheet_name,
            row_locator=row_locator,
            cell_ref=cell_ref,
        )
    )


def _build_source_version(path: Path) -> str:
    if not path.exists():
        digest = hashlib.sha256(str(path).encode("utf-8")).hexdigest()[:12]
        return f"sv_qdb_gl_{digest}"

    stat = path.stat()
    seed = f"{path.name}:{stat.st_size}:{stat.st_mtime_ns}"
    digest = hashlib.sha256(seed.encode("utf-8")).hexdigest()[:12]
    return f"sv_qdb_gl_{digest}"


def _build_trace_id(source_kind: str, source_version: str) -> str:
    return f"tr_qdb_gl_{source_kind}_{source_version.removeprefix('sv_qdb_gl_')}"


def _normalize_account_code(value: object) -> str | None:
    if value is None:
        return None

    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        if not value.is_integer():
            return None
        text = str(int(value))
    else:
        text = str(value).strip()

    if not text or "e" in text.lower() or not text.isdigit():
        return None
    if len(text) > 11:
        return None
    return text


def _normalize_currency(value: object) -> str | None:
    text = _normalize_text(value).upper()
    if text in ALLOWED_CURRENCY_GROUPS:
        return text
    return None


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).replace("\xa0", " ").strip()


def _to_decimal(value: object) -> Decimal | None:
    if _is_blank(value):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _is_blank(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False
