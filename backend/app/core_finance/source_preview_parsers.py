from __future__ import annotations

import hashlib
from pathlib import Path

import xlrd
from backend.app.schemas.source_preview import (
    NonstdPnlPreviewRow,
    PnlPreviewRow,
    TywPreviewRow,
    ZqtzPreviewRow,
)
from backend.app.services.source_rules import (
    classify_nonstd_pnl_preview,
    classify_pnl_preview,
    classify_tyw_preview,
    classify_zqtz_preview,
    describe_source_file,
)
from openpyxl import load_workbook

RULE_VERSION = "rv_phase1_source_preview_v1"

ZQTZ_BOND_CODE = "\u503a\u5238\u4ee3\u53f7"
ZQTZ_BOND_NAME = "\u503a\u5238\u540d\u79f0"
ZQTZ_DATE = "\u65e5\u671f"
ZQTZ_BUSINESS_TYPE1 = "\u4e1a\u52a1\u79cd\u7c7b1"
ZQTZ_ACCOUNT_CATEGORY = "\u8d26\u6237\u7c7b\u522b"

TYW_PRODUCT_TYPE = "\u4ea7\u54c1\u7c7b\u578b"
TYW_COUNTERPARTY_NAME = "\u5bf9\u624b\u65b9\u540d\u79f0"
TYW_INVESTMENT_PORTFOLIO = "\u6295\u8d44\u7ec4\u5408"
TYW_CBIRC_TYPE = "\u4f1a\u8ba1\u7c7b\u578b_\u94f6\u4fdd\u76d1\u4f1a"
TYW_PBOC_TYPE = "\u4f1a\u8ba1\u7c7b\u578b_\u4eba\u884c"
TYW_CORE_CUSTOMER_TYPE = "\u6838\u5fc3\u5ba2\u6237\u7c7b\u578b"
TYW_ACCOUNT_TYPE = "\u8d26\u6237\u7c7b\u578b"
TYW_SPECIAL_ACCOUNT_TYPE = "\u7279\u6b8a\u8d26\u6237\u7c7b\u578b"
TYW_CUSTODY_ACCOUNT_NAME = "\u6258\u7ba1\u8d26\u6237\u540d\u79f0"
TYW_TRACE_FIELDS = {
    TYW_PRODUCT_TYPE,
    TYW_INVESTMENT_PORTFOLIO,
    TYW_CBIRC_TYPE,
    TYW_PBOC_TYPE,
    TYW_CORE_CUSTOMER_TYPE,
    TYW_ACCOUNT_TYPE,
    TYW_SPECIAL_ACCOUNT_TYPE,
    TYW_CUSTODY_ACCOUNT_NAME,
}


def build_source_version(path: Path) -> str:
    stat = path.stat()
    seed = f"{path.name}:{stat.st_size}:{stat.st_mtime_ns}"
    return f"sv_{hashlib.sha256(seed.encode('utf-8')).hexdigest()[:12]}"


def parse_source_file(
    path: Path,
    ingest_batch_id: str,
    source_version: str,
    source_file_name: str | None = None,
) -> tuple[str, str | None, list[dict[str, object]], list[dict[str, object]]]:
    metadata = describe_source_file(source_file_name or path.name)
    source_family = metadata.source_family
    if source_family == "pnl":
        return _parse_pnl_source_file(
            path=path,
            ingest_batch_id=ingest_batch_id,
            source_version=source_version,
            metadata=metadata,
        )

    if source_family in {"pnl_514", "pnl_516", "pnl_517"}:
        return _parse_nonstd_pnl_source_file(
            path=path,
            ingest_batch_id=ingest_batch_id,
            source_version=source_version,
            metadata=metadata,
        )

    if source_family not in {"zqtz", "tyw"}:
        return source_family, metadata.report_date, [], []

    sheet = xlrd.open_workbook(str(path)).sheet_by_index(0)
    headers = [str(sheet.cell_value(1, column)).strip() for column in range(sheet.ncols)]
    rows: list[dict[str, object]] = []
    traces: list[dict[str, object]] = []
    row_locator = 0

    for row_index in range(2, sheet.nrows):
        raw_row = {
            headers[column]: sheet.cell_value(row_index, column)
            for column in range(sheet.ncols)
            if headers[column]
        }
        row_locator += 1
        if source_family == "zqtz":
            preview = classify_zqtz_preview(raw_row)
            row_record = ZqtzPreviewRow(
                ingest_batch_id=ingest_batch_id,
                row_locator=row_locator,
                report_date=metadata.report_date or _text(raw_row, ZQTZ_DATE) or None,
                business_type_primary=str(preview["business_type_primary"]),
                business_type_final=str(preview["business_type_final"]),
                asset_group=str(preview["asset_group"]),
                instrument_code=_text(raw_row, ZQTZ_BOND_CODE),
                instrument_name=_text(raw_row, ZQTZ_BOND_NAME),
                account_category=_text(raw_row, ZQTZ_ACCOUNT_CATEGORY),
                manual_review_needed=bool(preview["manual_review_needed"]),
            ).model_dump(mode="json")
            trace_rows = _zqtz_trace_rows(raw_row, row_record)
        else:
            preview = classify_tyw_preview(raw_row)
            row_record = TywPreviewRow(
                ingest_batch_id=ingest_batch_id,
                row_locator=row_locator,
                report_date=metadata.report_date,
                business_type_primary=str(preview["business_type_primary"]),
                product_group=str(preview["product_group"]),
                institution_category=str(preview["institution_category"]),
                special_nature=str(preview["special_nature"]),
                counterparty_name=_text(raw_row, TYW_COUNTERPARTY_NAME),
                investment_portfolio=str(preview["investment_portfolio"]),
                manual_review_needed=bool(preview["manual_review_needed"]),
            ).model_dump(mode="json")
            trace_rows = _tyw_trace_rows(raw_row, row_record)

        row_record["source_version"] = source_version
        row_record["rule_version"] = RULE_VERSION
        rows.append(row_record)
        traces.extend(trace_rows)

    return source_family, metadata.report_date, rows, traces


def _parse_pnl_source_file(
    path: Path,
    ingest_batch_id: str,
    source_version: str,
    metadata,
) -> tuple[str, str | None, list[dict[str, object]], list[dict[str, object]]]:
    sheet = xlrd.open_workbook(str(path)).sheet_by_index(0)
    headers = [str(sheet.cell_value(0, column)).strip() for column in range(sheet.ncols)]
    rows: list[dict[str, object]] = []
    traces: list[dict[str, object]] = []
    row_locator = 0

    for row_index in range(1, sheet.nrows):
        raw_row = {
            headers[column]: sheet.cell_value(row_index, column)
            for column in range(sheet.ncols)
            if headers[column]
        }
        if not _text(raw_row, "\u503a\u5238\u4ee3\u7801"):
            continue

        row_locator += 1
        preview = classify_pnl_preview(raw_row)
        row_record = PnlPreviewRow(
            ingest_batch_id=ingest_batch_id,
            row_locator=row_locator,
            report_date=metadata.report_date,
            instrument_code=str(preview["instrument_code"]),
            invest_type_raw=str(preview["invest_type_raw"]),
            portfolio_name=str(preview["portfolio_name"]),
            cost_center=str(preview["cost_center"]),
            currency=str(preview["currency"]),
            manual_review_needed=bool(preview["manual_review_needed"]),
        ).model_dump(mode="json")
        row_record["source_family"] = "pnl"
        row_record["source_version"] = source_version
        row_record["rule_version"] = RULE_VERSION
        rows.append(row_record)
        traces.extend(_pnl_trace_rows(raw_row, row_record))

    return "pnl", metadata.report_date, rows, traces


def _parse_nonstd_pnl_source_file(
    path: Path,
    ingest_batch_id: str,
    source_version: str,
    metadata,
) -> tuple[str, str | None, list[dict[str, object]], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    headers = [
        "" if value is None else str(value).strip()
        for value in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
    ]
    rows: list[dict[str, object]] = []
    traces: list[dict[str, object]] = []
    row_locator = 0
    bucket = metadata.source_family.removeprefix("pnl_")

    for values in worksheet.iter_rows(min_row=3, values_only=True):
        raw_row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        if not _text(raw_row, "\u8d44\u4ea7\u4ee3\u7801"):
            continue

        row_locator += 1
        preview = classify_nonstd_pnl_preview(raw_row, bucket=bucket)
        row_record = NonstdPnlPreviewRow(
            ingest_batch_id=ingest_batch_id,
            row_locator=row_locator,
            report_date=metadata.report_date,
            journal_type=str(preview["journal_type"]),
            product_type=str(preview["product_type"]),
            asset_code=str(preview["asset_code"]),
            account_code=str(preview["account_code"]),
            dc_flag_raw=str(preview["dc_flag_raw"]),
            raw_amount=str(preview["raw_amount"]),
            manual_review_needed=bool(preview["manual_review_needed"]),
        ).model_dump(mode="json")
        row_record["source_family"] = metadata.source_family
        row_record["source_version"] = source_version
        row_record["rule_version"] = RULE_VERSION
        rows.append(row_record)
        traces.extend(_nonstd_pnl_trace_rows(raw_row, row_record))

    return metadata.source_family, metadata.report_date, rows, traces


def _zqtz_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    rows = [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": _row_locator_value(row_record),
            "trace_step": 1,
            "field_name": ZQTZ_BUSINESS_TYPE1,
            "field_value": _text(raw_row, ZQTZ_BUSINESS_TYPE1),
            "derived_label": str(row_record["business_type_primary"]),
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "zqtz",
        }
    ]
    if str(row_record["business_type_primary"]) == "\u5176\u4ed6\u503a\u5238":
        rows.append(
            {
                "ingest_batch_id": str(row_record["ingest_batch_id"]),
                "row_locator": _row_locator_value(row_record),
                "trace_step": 2,
                "field_name": ZQTZ_BOND_CODE,
                "field_value": _text(raw_row, ZQTZ_BOND_CODE),
                "derived_label": str(row_record["business_type_final"]),
                "manual_review_needed": bool(row_record["manual_review_needed"]),
                "source_family": "zqtz",
            }
        )
    rows.append(
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": _row_locator_value(row_record),
            "trace_step": 3,
            "field_name": "asset_group_map",
            "field_value": str(row_record["business_type_final"]),
            "derived_label": str(row_record["asset_group"]),
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "zqtz",
        }
    )
    return rows


def _tyw_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    trace_spec = [
        (1, TYW_PRODUCT_TYPE, str(row_record["business_type_primary"])),
        (2, TYW_INVESTMENT_PORTFOLIO, str(row_record["product_group"])),
        (3, TYW_CBIRC_TYPE, str(row_record["institution_category"])),
        (4, TYW_PBOC_TYPE, str(row_record["institution_category"])),
        (5, TYW_CORE_CUSTOMER_TYPE, str(row_record["institution_category"])),
        (6, TYW_ACCOUNT_TYPE, str(row_record["special_nature"])),
        (7, TYW_SPECIAL_ACCOUNT_TYPE, str(row_record["special_nature"])),
        (8, TYW_CUSTODY_ACCOUNT_NAME, str(row_record["special_nature"])),
    ]
    return [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": _row_locator_value(row_record),
            "trace_step": trace_step,
            "field_name": field_name,
            "field_value": _text(raw_row, field_name),
            "derived_label": derived_label,
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "tyw",
        }
        for trace_step, field_name, derived_label in trace_spec
    ]


def _pnl_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    trace_spec = [
        (1, "\u6295\u8d44\u7c7b\u578b", str(row_record["invest_type_raw"])),
        (2, "\u503a\u5238\u4ee3\u7801", str(row_record["instrument_code"])),
        (3, "\u6210\u672c\u4e2d\u5fc3", str(row_record["cost_center"])),
    ]
    return [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": _row_locator_value(row_record),
            "trace_step": trace_step,
            "field_name": field_name,
            "field_value": _text(raw_row, field_name),
            "derived_label": derived_label,
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "pnl",
        }
        for trace_step, field_name, derived_label in trace_spec
    ]


def _nonstd_pnl_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    trace_spec = [
        (1, "科目号", str(row_record["account_code"])),
        (2, "资产代码", str(row_record["asset_code"])),
        (3, "借贷标识", str(row_record["dc_flag_raw"])),
    ]
    return [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": _row_locator_value(row_record),
            "trace_step": trace_step,
            "field_name": field_name,
            "field_value": _text(raw_row, field_name),
            "derived_label": derived_label,
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": str(row_record["source_family"]),
        }
        for trace_step, field_name, derived_label in trace_spec
    ]


def _text(row: dict[str, object], key: str) -> str:
    value = row.get(key, "")
    if value is None:
        return ""
    return str(value).strip()


def _row_locator_value(row_record: dict[str, object]) -> int:
    return int(str(row_record["row_locator"]))
