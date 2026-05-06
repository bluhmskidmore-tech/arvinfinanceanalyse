from __future__ import annotations

from calendar import monthrange
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Any

from backend.app.core_finance.reconciliation_checks import position_vs_ledger_diff
from openpyxl import Workbook, load_workbook

ZERO = Decimal("0")
ONE_HUNDRED_MILLION = Decimal("100000000")

CONFIG = {
    "DEVIATION_WARN": Decimal("5"),
    "DEVIATION_ALERT": Decimal("10"),
    "DEVIATION_CRITICAL": Decimal("20"),
    "MIN_AMOUNT_11D": Decimal("1"),
}

LOAN_BALANCE_CODES_3D = {"122", "123", "129", "130", "132", "136"}
DEPOSIT_BALANCE_CODES_3D = {
    "201",
    "202",
    "203",
    "204",
    "205",
    "211",
    "215",
    "216",
    "217",
    "225",
    "243",
    "244",
    "251",
}
TERM_DEPOSIT_CODES_3D = {"205", "215"}
DEMAND_DEPOSIT_CODES_3D = {"201", "211"}
INVESTMENT_BALANCE_CODES_3D = {"141", "142", "143", "144", "145"}
LIQUID_ASSET_CODES_3D = {"101", "110", "114", "116"}


def _effective_analysis_config(overrides: dict[str, Any] | None) -> dict[str, Decimal]:
    merged = {key: CONFIG[key] for key in CONFIG}
    if not overrides:
        return merged
    for key, value in overrides.items():
        if key in merged:
            merged[key] = Decimal(str(value))
    return merged

CATEGORY_NAMES_3D = {
    "101": "现金",
    "110": "存放央行",
    "114": "存放同业",
    "116": "存放境外同业",
    "120": "同业借出/贴现",
    "122": "个人贷款",
    "123": "公司贷款",
    "130": "逾期贷款",
    "131": "贷款减值准备",
    "132": "福费廷/押汇",
    "133": "应收利息",
    "140": "买入返售",
    "141": "FVTPL金融资产",
    "142": "AC债券投资",
    "143": "AC其他投资",
    "144": "FVOCI金融资产",
    "145": "长期股权投资",
    "201": "单位活期存款",
    "205": "单位定期存款",
    "211": "活期储蓄存款",
    "215": "定期储蓄存款",
    "234": "同业存放",
    "235": "同业定期存款",
    "241": "拆入资金",
    "251": "保证金存款",
    "255": "卖出回购",
    "272": "应付债券/存单",
}

INDUSTRY_NAMES = {
    "01": "农林牧渔",
    "02": "采矿业",
    "03": "制造业",
    "04": "电力燃气水",
    "05": "建筑业",
    "06": "交通运输",
    "07": "信息技术",
    "08": "批发零售",
    "09": "住宿餐饮",
    "10": "金融业",
    "11": "房地产",
    "12": "租赁商务",
    "13": "科技服务",
    "14": "水利环境",
    "15": "居民服务",
    "16": "教育",
    "17": "卫生社保",
    "18": "文体娱乐",
    "19": "公共管理",
    "99": "大额存单",
}

DAILY_AVG_COLUMN_MAP = [
    (1, 2, "CNX", "3d"),
    (5, 6, "CNX", "5d"),
    (9, 10, "CNX", "7d"),
    (13, 14, "CNX", "11d"),
    (17, 18, "CNY", "3d"),
    (21, 22, "CNY", "5d"),
    (25, 26, "CNY", "7d"),
    (29, 30, "CNY", "11d"),
]

SEGMENT_BASE_SCALE_SOURCE = "分部基础数据（2026）"
SEGMENT_BASE_SCALE_MICRO_LOAN_MISSING_SOURCE = "source_missing: 标准日均源不含80297微贷金融支行专段"
SEGMENT_SCALE_COMPARE_SOURCE = "月度分析-分部情况：总账对账+日均同源历史月重建"
COMPANY_SCALE_SOURCE = "公司规模：总账对账+日均同源科目重建"
COMPANY_SCALE_COMPARE_SOURCE = "月度分析-公司板块：总账对账+日均同源历史月重建"
RETAIL_SCALE_SOURCE = "零售规模：总账对账+日均同源科目重建"
RETAIL_SCALE_BRANCH_LOAN_MISSING_SOURCE = "source_missing: 零售分支行个贷依赖80297微贷金融支行专段"
RETAIL_SCALE_COMPARE_SOURCE = "月度分析-零售板块：总账对账+日均同源历史月重建"
FINANCIAL_MARKET_SCALE_SOURCE = "金融市场规模：总账对账+日均同源科目重建"
FINANCIAL_MARKET_SCALE_COMPARE_SOURCE = "月度分析-金融市场：总账对账+日均同源历史月重建"
INCOME_RATE_ANALYSIS_SOURCE = "收益率分析：总账收益科目+日均规模重建"
INCOME_RATE_ATTRIBUTION_SOURCE = "收益量价归因：总账收益科目+日均规模按年累计同比拆解"
INCOME_RATE_MISSING_SOURCE = "source_missing: 财务指标表该项依赖外部营收分项/FTP/收益率来源，当前总账+日均闭环未确认"
INCOME_RATE_PERSONAL_LOAN_SCALE_MISSING_SOURCE = "source_missing: 个人贷款收益率分母依赖信用卡生息规模/80297微贷拆分，当前总账+日均闭环未确认"
DEPOSIT_INTEREST_SPLIT_SOURCE = "存款利息拆分：总账521利息支出+日均存款规模重建"
PARENT_COMPANY_REVENUE_SOURCE = "母公司营收分项：总账损益科目可复算部分"
PARENT_COMPANY_REVENUE_MISSING_SOURCE = "source_missing: 母公司营收分项该行依赖外部营收分项/FTP/非息明细来源，当前总账+日均闭环未确认"


def parse_daily_avg(filepath: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(filename=str(filepath), read_only=True, data_only=True)
    try:
        result: dict[str, list[dict[str, Any]]] = {}
        for sheet_name, period_label in (("月", "月日均"), ("年", "年日均")):
            worksheet = workbook[sheet_name]
            for code_col, value_col, currency, level in DAILY_AVG_COLUMN_MAP:
                key = f"{period_label}_{currency}_{level}"
                rows: list[dict[str, Any]] = []
                for row in worksheet.iter_rows(min_row=4, values_only=True):
                    code = _normalize_account_code(row[code_col] if len(row) > code_col else None)
                    if not code:
                        continue
                    value = _to_decimal(row[value_col] if len(row) > value_col else None)
                    if value is None:
                        continue
                    rows.append({"科目代码": code, "日均余额": value, "币种": currency, "级别": level})
                result[key] = rows
        return result
    finally:
        workbook.close()


def parse_general_ledger(filepath: str | Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(filename=str(filepath), read_only=True, data_only=True)
    try:
        result: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            rows: list[dict[str, Any]] = []
            for row in worksheet.iter_rows(min_row=7, values_only=True):
                code = _normalize_account_code(row[0] if len(row) > 0 else None)
                if not code:
                    continue
                currency = str(row[2] or "").strip()
                if not currency:
                    continue
                opening = _to_decimal(row[3] if len(row) > 3 else None) or ZERO
                debit = _to_decimal(row[4] if len(row) > 4 else None) or ZERO
                credit = _to_decimal(row[5] if len(row) > 5 else None) or ZERO
                closing = _to_decimal(row[6] if len(row) > 6 else None) or ZERO
                rows.append(
                    {
                        "科目代码": code,
                        "科目名称": str(row[1] or "").strip(),
                        "币种": currency,
                        "期初余额": opening,
                        "本期借方": debit,
                        "本期贷方": credit,
                        "期末余额": closing,
                        "变动额": closing - opening,
                        "本期净额": debit - credit,
                        "大类1": code[:1],
                        "大类3": code[:3],
                        "大类5": code[:5],
                    }
                )
            result[sheet_name] = rows
        return result
    finally:
        workbook.close()


def merge_all(gl_data: dict[str, list[dict[str, Any]]], rj_data: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    cnx_rows = [row for row in gl_data.get("综本", []) if row["币种"] == "CNX"]
    merged: dict[str, Any] = {}

    rj_m_11 = _index_amounts(rj_data.get("月日均_CNX_11d", []))
    rj_y_11 = _index_amounts(rj_data.get("年日均_CNX_11d", []))
    merged["11位"] = [{**row, "月日均": rj_m_11.get(row["科目代码"]), "年日均": rj_y_11.get(row["科目代码"])} for row in cnx_rows]

    grouped_3 = _group_sum(cnx_rows, key="大类3")
    rj_m_3 = _index_amounts(rj_data.get("月日均_CNX_3d", []))
    rj_y_3 = _index_amounts(rj_data.get("年日均_CNX_3d", []))
    rj_m_5 = _index_amounts(rj_data.get("月日均_CNX_5d", []))
    rj_y_5 = _index_amounts(rj_data.get("年日均_CNX_5d", []))
    merged["3位"] = [
        {
            **row,
            "科目代码": row["group_key"],
            "名称": CATEGORY_NAMES_3D.get(row["group_key"], ""),
            "月日均": rj_m_3.get(row["group_key"]),
            "年日均": rj_y_3.get(row["group_key"]),
        }
        for row in grouped_3
    ]
    merged["日均_3位"] = _daily_average_level_rows(month_avg=rj_m_3, year_avg=rj_y_3)
    merged["日均_5位"] = _daily_average_level_rows(month_avg=rj_m_5, year_avg=rj_y_5)

    for prefix, label in (("123", "公司贷款"), ("201", "活期存款"), ("205", "定期存款"), ("251", "保证金")):
        subset = [row for row in cnx_rows if row["大类3"] == prefix]
        grouped_5 = _group_sum(subset, key="大类5", include_name=True)
        merged[f"5位_{label}"] = [
            {
                **row,
                "科目代码": row["group_key"],
                "行业代码": row["group_key"][3:5],
                "行业名称": INDUSTRY_NAMES.get(row["group_key"][3:5], ""),
                "月日均": rj_m_5.get(row["group_key"]),
                "年日均": rj_y_5.get(row["group_key"]),
            }
            for row in grouped_5
        ]

    cny_index = {row["科目代码"]: row for row in gl_data.get("人民币", [])}
    merged["外币分析"] = []
    for row in cnx_rows:
        cny_row = cny_index.get(row["科目代码"])
        if cny_row is None:
            continue
        cnx_value = row["期末余额"]
        cny_value = cny_row["期末余额"]
        foreign_value = cnx_value - cny_value
        foreign_share = None if cnx_value == ZERO else foreign_value / abs(cnx_value) * Decimal("100")
        merged["外币分析"].append(
            {
                "科目代码": row["科目代码"],
                "科目名称": row["科目名称"],
                "期末余额_综本": cnx_value,
                "期末余额_人民币": cny_value,
                "外币部分": foreign_value,
                "外币占比%": foreign_share,
            }
        )
    return merged


def build_qdb_gl_monthly_analysis_workbook(
    *,
    report_month: str,
    merged_data: dict[str, Any],
    threshold_overrides: dict[str, int | float] | None = None,
    comparison_data: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    analysis_cfg = _effective_analysis_config(threshold_overrides)
    m3 = compute_deviation(merged_data.get("3位", []))
    m11 = compute_deviation(merged_data.get("11位", []))
    metrics = compute_asset_liability_structure(m3)
    gap_rows = compute_industry_gap(merged_data)
    alert_rows = [
        *generate_alerts(merged_data, analysis_config=analysis_cfg),
        *_position_ledger_reconciliation_alerts(
            _qdb_gl_position_vs_ledger_check(
                position_totals=metrics,
                ledger_rows_3d=m3,
            )
        ),
    ]
    foreign_rows = build_foreign_currency_rows(merged_data.get("外币分析", []))

    sheets = [
        _sheet("overview", "经营概览", ["指标", "值"], [
            {"指标": "总资产(亿)", "值": _display_number(_to_yi(metrics["总资产"]))},
            {"指标": "总负债(亿)", "值": _display_number(_to_yi(metrics["总负债"]))},
            {"指标": "贷款总额(亿)", "值": _display_number(_to_yi(metrics["贷款总额"]))},
            {"指标": "存款总额(亿)", "值": _display_number(_to_yi(metrics["存款总额"]))},
            {"指标": "投资总额(亿)", "值": _display_number(_to_yi(metrics["投资总额"]))},
            {"指标": "存贷比%", "值": _display_number(metrics["存贷比%"])},
            {"指标": "贷款减值准备率%", "值": _display_number(metrics["贷款减值准备率%"])},
            {"指标": "定期化率%", "值": _display_number(metrics["定期化率%"])},
            {"指标": "活期率%", "值": _display_number(metrics["活期率%"])},
            {"指标": "高流动性占比%", "值": _display_number(metrics["高流动性占比%"])},
        ]),
        _sheet("summary_3d", "3位科目总览", ["科目代码", "名称", "期初余额", "期末余额", "变动额", "月日均", "年日均", "偏离额", "偏离%", "趋势额", "趋势%"], [_normalize_amount_row(row, include_trend=True) for row in _sort_rows(m3, "期末余额")]),
        _sheet("asset_structure", "资产结构", ["科目代码", "名称", "期末余额", "月日均", "偏离%", "趋势%"], [
            {"科目代码": row["科目代码"], "名称": row["名称"], "期末余额": _display_number(_to_yi(row["期末余额"])), "月日均": _display_number(_to_yi(row.get("月日均"))), "偏离%": _display_number(row.get("偏离%")), "趋势%": _display_number(row.get("趋势%"))}
            for row in _sort_rows([row for row in m3 if str(row["科目代码"]).startswith("1")], "期末余额")
        ]),
        _sheet("liability_structure", "负债结构", ["科目代码", "名称", "期末余额", "月日均", "偏离%", "趋势%"], [
            {"科目代码": row["科目代码"], "名称": row["名称"], "期末余额": _display_number(_to_yi(row["期末余额"])), "月日均": _display_number(_to_yi(row.get("月日均"))), "偏离%": _display_number(row.get("偏离%")), "趋势%": _display_number(row.get("趋势%"))}
            for row in _sort_rows([row for row in m3 if str(row["科目代码"]).startswith("2")], "期末余额")
        ]),
        _industry_sheet("loan_industry", "贷款行业", compute_deviation(merged_data.get("5位_公司贷款", []))),
        _industry_sheet("deposit_demand_industry", "存款行业_活期", compute_deviation(merged_data.get("5位_活期存款", []))),
        _industry_sheet("deposit_term_industry", "存款行业_定期", compute_deviation(merged_data.get("5位_定期存款", []))),
        _sheet("industry_gap", "行业存贷差", ["行业", "贷款期末", "存款期末", "存贷差_时点", "贷款月日均", "存款月日均", "存贷差_日均"], gap_rows),
        _sheet("top_11d", "11位偏离TOP", ["科目代码", "科目名称", "期末余额", "月日均", "年日均", "偏离额", "偏离%", "趋势额", "趋势%"], build_11d_top_rows(m11, analysis_config=analysis_cfg)),
        _sheet("alerts", "异动预警", ["科目代码", "科目名称", "预警级别", "期末余额(亿)", "月日均(亿)", "偏离额(亿)", "偏离%", "异动类型"], alert_rows),
        _sheet("foreign_currency", "外币分析", ["科目代码", "科目名称", "期末余额_综本", "期末余额_人民币", "外币部分", "外币占比%"], foreign_rows),
    ]
    segment_base_scale_sheet = _build_segment_base_scale_sheet(report_month=report_month, merged_data=merged_data)
    if segment_base_scale_sheet is not None:
        sheets.append(segment_base_scale_sheet)
    segment_scale_compare_sheet = _build_segment_scale_compare_sheet(
        merged_data=merged_data,
        comparison_data=comparison_data,
    )
    if segment_scale_compare_sheet is not None:
        sheets.append(segment_scale_compare_sheet)
    company_scale_sheet = _build_company_scale_sheet(
        report_month=report_month,
        merged_data=merged_data,
    )
    if company_scale_sheet is not None:
        sheets.append(company_scale_sheet)
    company_scale_compare_sheet = _build_company_scale_compare_sheet(
        merged_data=merged_data,
        comparison_data=comparison_data,
    )
    if company_scale_compare_sheet is not None:
        sheets.append(company_scale_compare_sheet)
    retail_scale_sheet = _build_retail_scale_sheet(
        report_month=report_month,
        merged_data=merged_data,
    )
    if retail_scale_sheet is not None:
        sheets.append(retail_scale_sheet)
    retail_scale_compare_sheet = _build_retail_scale_compare_sheet(
        merged_data=merged_data,
        comparison_data=comparison_data,
    )
    if retail_scale_compare_sheet is not None:
        sheets.append(retail_scale_compare_sheet)
    financial_market_scale_sheet = _build_financial_market_scale_sheet(
        report_month=report_month,
        merged_data=merged_data,
    )
    if financial_market_scale_sheet is not None:
        sheets.append(financial_market_scale_sheet)
    financial_market_scale_compare_sheet = _build_financial_market_scale_compare_sheet(
        merged_data=merged_data,
        comparison_data=comparison_data,
    )
    if financial_market_scale_compare_sheet is not None:
        sheets.append(financial_market_scale_compare_sheet)
    income_rate_analysis_sheet = _build_income_rate_analysis_sheet(
        report_month=report_month,
        merged_data=merged_data,
    )
    if income_rate_analysis_sheet is not None:
        sheets.append(income_rate_analysis_sheet)
    income_rate_attribution_sheet = _build_income_rate_attribution_sheet(
        report_month=report_month,
        merged_data=merged_data,
        comparison_data=comparison_data,
    )
    if income_rate_attribution_sheet is not None:
        sheets.append(income_rate_attribution_sheet)
    deposit_interest_split_sheet = _build_deposit_interest_split_sheet(
        report_month=report_month,
        merged_data=merged_data,
        comparison_data=comparison_data,
    )
    if deposit_interest_split_sheet is not None:
        sheets.append(deposit_interest_split_sheet)
    parent_company_revenue_sheet = _build_parent_company_revenue_components_sheet(
        report_month=report_month,
        merged_data=merged_data,
        comparison_data=comparison_data,
    )
    if parent_company_revenue_sheet is not None:
        sheets.append(parent_company_revenue_sheet)
    return {"report_month": report_month, "sheets": sheets}


def export_qdb_gl_monthly_analysis_workbook_xlsx_bytes(workbook_payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    active = workbook.active
    for index, sheet_payload in enumerate(workbook_payload.get("sheets", [])):
        worksheet = active if index == 0 else workbook.create_sheet()
        worksheet.title = str(sheet_payload["title"])
        columns = list(sheet_payload.get("columns", []))
        for column_index, header in enumerate(columns, start=1):
            worksheet.cell(row=1, column=column_index, value=header)
        for row_index, row in enumerate(sheet_payload.get("rows", []), start=2):
            for column_index, header in enumerate(columns, start=1):
                worksheet.cell(row=row_index, column=column_index, value=_excel_value(row.get(header)))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def compute_deviation(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    computed: list[dict[str, Any]] = []
    for row in rows:
        current = dict(row)
        month_avg = _as_decimal(row.get("月日均"))
        year_avg = _as_decimal(row.get("年日均"))
        ending = _as_decimal(row.get("期末余额"))
        current["偏离额"] = None if month_avg is None else ending - month_avg
        current["偏离%"] = _safe_pct(current["偏离额"], month_avg)
        current["趋势额"] = None if month_avg is None or year_avg is None else month_avg - year_avg
        current["趋势%"] = _safe_pct(current["趋势额"], year_avg)
        computed.append(current)
    return computed


def compute_asset_liability_structure(rows_3d: list[dict[str, Any]]) -> dict[str, Decimal]:
    rows_by_code = {row["科目代码"]: row for row in rows_3d}

    def value(code: str, field: str = "期末余额") -> Decimal:
        row = rows_by_code.get(code)
        return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)

    def sum_codes(codes: set[str]) -> Decimal:
        return sum((value(code) for code in codes), ZERO)

    # 财务指标表的总账口径按资产/负债科目净额汇总；不能把所有正余额相加，
    # 否则损益类、表外类正余额会被误算进总资产。
    total_assets = sum(
        (_as_decimal(row.get("期末余额")) or ZERO
         for row in rows_3d
         if str(row["科目代码"]).startswith("1")),
        ZERO,
    )
    total_liabilities = abs(
        sum(
            (_as_decimal(row.get("期末余额")) or ZERO
             for row in rows_3d
             if str(row["科目代码"]).startswith("2")),
            ZERO,
        )
    )
    loan_total = sum_codes(LOAN_BALANCE_CODES_3D)
    deposit_total = abs(sum_codes(DEPOSIT_BALANCE_CODES_3D))
    provision = abs(value("131"))
    investment_total = sum_codes(INVESTMENT_BALANCE_CODES_3D)
    liquid_total = sum_codes(LIQUID_ASSET_CODES_3D)
    term_deposit = abs(sum_codes(TERM_DEPOSIT_CODES_3D))
    demand_deposit = abs(sum_codes(DEMAND_DEPOSIT_CODES_3D))
    return {
        "贷款总额": loan_total,
        "存款总额": deposit_total,
        "总资产": total_assets,
        "总负债": total_liabilities,
        "投资总额": investment_total,
        "存贷比%": _pct(loan_total, deposit_total),
        "贷款减值准备率%": _pct(provision, loan_total),
        "定期化率%": _pct(term_deposit, deposit_total),
        "活期率%": _pct(demand_deposit, deposit_total),
        "高流动性占比%": _pct(liquid_total, total_assets),
    }


def _qdb_gl_position_vs_ledger_check(
    *,
    position_totals: dict[str, Any],
    ledger_rows_3d: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    ledger_totals = _qdb_gl_ledger_totals(ledger_rows_3d)
    total_assets = _as_decimal(position_totals.get("总资产")) or ZERO
    total_liabilities = _as_decimal(position_totals.get("总负债")) or ZERO
    return position_vs_ledger_diff(
        {
            "total_assets": float(total_assets),
            "total_liabilities": float(total_liabilities),
            "net_assets": float(total_assets - total_liabilities),
        },
        ledger_totals,
        threshold_yuan=0.01,
    )


def _qdb_gl_ledger_totals(rows_3d: list[dict[str, Any]]) -> dict[str, float]:
    total_assets = sum(
        (_as_decimal(row.get("期末余额")) or ZERO)
        for row in rows_3d
        if str(row.get("科目代码") or "").startswith("1")
    )
    total_liabilities = abs(
        sum(
            (_as_decimal(row.get("期末余额")) or ZERO)
            for row in rows_3d
            if str(row.get("科目代码") or "").startswith("2")
        )
    )
    return {
        "total_assets": float(total_assets),
        "total_liabilities": float(total_liabilities),
        "net_assets": float(total_assets - total_liabilities),
    }


def _position_ledger_reconciliation_alerts(checks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    alerts: list[dict[str, Any]] = []
    for check in checks:
        if not check["breached"]:
            continue
        alerts.append(
            {
                "科目代码": check["dimension"],
                "科目名称": "Position vs Ledger",
                "预警级别": "严重",
                "期末余额(亿)": _display_number(_to_yi(Decimal(str(check["position_value"])))),
                "月日均(亿)": _display_number(_to_yi(Decimal(str(check["ledger_value"])))),
                "偏离额(亿)": _display_number(_to_yi(Decimal(str(check["diff"])))),
                "偏离%": None,
                "异动类型": "position_vs_ledger_reconciliation",
            }
        )
    return alerts


def compute_industry_gap(merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    loans = merged_data.get("5位_公司贷款", [])
    demand = {row["行业代码"]: row for row in merged_data.get("5位_活期存款", [])}
    term = {row["行业代码"]: row for row in merged_data.get("5位_定期存款", [])}
    rows: list[dict[str, Any]] = []
    for row in loans:
        industry_code = row["行业代码"]
        demand_row = demand.get(industry_code)
        term_row = term.get(industry_code)
        deposit_end = abs(_as_decimal(demand_row.get("期末余额")) or ZERO) + abs(_as_decimal(term_row.get("期末余额")) or ZERO)
        deposit_avg = abs(_as_decimal(demand_row.get("月日均")) or ZERO) + abs(_as_decimal(term_row.get("月日均")) or ZERO)
        loan_end = _as_decimal(row.get("期末余额")) or ZERO
        loan_avg = _as_decimal(row.get("月日均")) or ZERO
        rows.append({"行业": row.get("行业名称", ""), "贷款期末": _display_number(_to_yi(loan_end)), "存款期末": _display_number(_to_yi(deposit_end)), "存贷差_时点": _display_number(_to_yi(loan_end - deposit_end)), "贷款月日均": _display_number(_to_yi(loan_avg)), "存款月日均": _display_number(_to_yi(deposit_avg)), "存贷差_日均": _display_number(_to_yi(loan_avg - deposit_avg))})
    return rows


def generate_alerts(
    merged_data: dict[str, Any],
    *,
    analysis_config: dict[str, Decimal] | None = None,
) -> list[dict[str, Any]]:
    cfg = analysis_config if analysis_config is not None else CONFIG
    rows_11 = compute_deviation(merged_data.get("11位", []))
    alerts: list[dict[str, Any]] = []
    for row in rows_11:
        deviation_pct = _as_decimal(row.get("偏离%"))
        deviation_amount = _as_decimal(row.get("偏离额"))
        if deviation_pct is None or deviation_amount is None:
            continue
        abs_pct = abs(deviation_pct)
        abs_yi = abs(_to_yi(deviation_amount))
        if abs_pct > cfg["DEVIATION_CRITICAL"] and abs_yi > Decimal("5"):
            level = "严重"
        elif abs_pct > cfg["DEVIATION_ALERT"] and abs_yi > Decimal("2"):
            level = "中度"
        elif abs_pct > cfg["DEVIATION_WARN"] and abs_yi > Decimal("1"):
            level = "轻度"
        else:
            continue
        code3 = str(row["科目代码"])[:3]
        first_digit = str(row["科目代码"])[:1]
        if code3 == "130":
            anomaly = "逾期贷款异动"
        elif code3 in {"140", "255", "234", "235", "114", "120"} and abs_pct > Decimal("30"):
            anomaly = "同业业务月末操纵嫌疑"
        elif first_digit == "2" and deviation_pct < Decimal("-10"):
            anomaly = "负债月末冲量"
        elif first_digit == "1" and deviation_pct < Decimal("-20"):
            anomaly = "资产月末压降"
        else:
            anomaly = "时点vs日均显著偏离"
        alerts.append({"科目代码": row["科目代码"], "科目名称": row.get("科目名称", ""), "预警级别": level, "期末余额(亿)": _display_number(_to_yi(_as_decimal(row.get("期末余额")) or ZERO)), "月日均(亿)": _display_number(_to_yi(_as_decimal(row.get("月日均")) or ZERO)), "偏离额(亿)": _display_number(_to_yi(deviation_amount)), "偏离%": _display_number(deviation_pct), "异动类型": anomaly})
    return alerts


def build_foreign_currency_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    filtered = [row for row in rows if abs(_as_decimal(row.get("外币部分")) or ZERO) > Decimal("10000000")]
    ordered = sorted(filtered, key=lambda row: abs(_as_decimal(row.get("外币部分")) or ZERO), reverse=True)
    return [{"科目代码": row["科目代码"], "科目名称": row["科目名称"], "期末余额_综本": _display_number(_to_yi(_as_decimal(row.get("期末余额_综本")) or ZERO)), "期末余额_人民币": _display_number(_to_yi(_as_decimal(row.get("期末余额_人民币")) or ZERO)), "外币部分": _display_number(_to_yi(_as_decimal(row.get("外币部分")) or ZERO)), "外币占比%": _display_number(_as_decimal(row.get("外币占比%")))} for row in ordered[:30]]


def build_11d_top_rows(
    rows: list[dict[str, Any]],
    *,
    analysis_config: dict[str, Decimal] | None = None,
) -> list[dict[str, Any]]:
    cfg = analysis_config if analysis_config is not None else CONFIG
    filtered = [row for row in rows if abs(_as_decimal(row.get("月日均")) or ZERO) > cfg["MIN_AMOUNT_11D"] * ONE_HUNDRED_MILLION]
    ordered = sorted(filtered, key=lambda row: abs(_as_decimal(row.get("偏离%")) or ZERO), reverse=True)[:50]
    return [{"科目代码": row["科目代码"], "科目名称": row.get("科目名称", ""), "期末余额": _display_number(_to_yi(_as_decimal(row.get("期末余额")) or ZERO)), "月日均": _display_number(_to_yi(_as_decimal(row.get("月日均")) or ZERO)), "年日均": _display_number(_to_yi(_as_decimal(row.get("年日均")) or ZERO)), "偏离额": _display_number(_to_yi(_as_decimal(row.get("偏离额")) or ZERO)), "偏离%": _display_number(_as_decimal(row.get("偏离%"))), "趋势额": _display_number(_to_yi(_as_decimal(row.get("趋势额")) or ZERO)), "趋势%": _display_number(_as_decimal(row.get("趋势%")))} for row in ordered]


def _build_segment_base_scale_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
) -> dict[str, Any] | None:
    if not report_month.startswith("2026"):
        return None

    rows = _segment_base_scale_raw_rows(merged_data)
    if not rows:
        return None
    return _sheet(
        "segment_base_scale",
        "分部基础规模",
        ["指标", "时点余额", "年日均", "月日均", "口径来源"],
        [
            {
                "指标": row["指标"],
                "时点余额": _display_yi(row["时点余额"]),
                "年日均": _display_yi(row["年日均"]),
                "月日均": _display_yi(row["月日均"]),
                "口径来源": row["口径来源"],
            }
            for row in rows
        ],
    )


def _segment_base_scale_raw_rows(merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    rows_3 = {row["科目代码"]: row for row in merged_data.get("3位", [])}
    rows_11 = list(merged_data.get("11位", []))
    if not rows_3 or not rows_11:
        return []

    def value_3d(code: str, field: str) -> Decimal | None:
        row = rows_3.get(code)
        return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)

    def value_11d(code: str, field: str) -> Decimal | None:
        for row in rows_11:
            if row["科目代码"] == code:
                return _as_decimal(row.get(field)) or ZERO
        return ZERO

    def sum_11d(prefix: str, field: str) -> Decimal | None:
        return sum(
            (_as_decimal(row.get(field)) or ZERO for row in rows_11 if row["科目代码"].startswith(prefix)),
            ZERO,
        )

    def sum_or_none(*values: Decimal | None) -> Decimal | None:
        if any(value is None for value in values):
            return None
        return sum((value for value in values if value is not None), ZERO)

    def amount_row(name: str, spot: Decimal | None, year_avg: Decimal | None, month_avg: Decimal | None, source: str) -> dict[str, Any]:
        return {
            "指标": name,
            "时点余额": spot,
            "年日均": year_avg,
            "月日均": month_avg,
            "口径来源": source,
        }

    def company_deposit(field: str) -> Decimal | None:
        company_demand_203 = value_3d("203", field) if field == "期末余额" else value_11d("20301000001", field)
        company_demand_204 = value_3d("204", field) if field == "期末余额" else value_11d("20401000001", field)
        return sum_or_none(
            value_3d("201", field),
            company_demand_203,
            company_demand_204,
            value_3d("202", field),
            None if sum_11d("20250", field) is None else -sum_11d("20250", field),
            value_3d("205", field),
            value_3d("225", field),
            value_3d("243", field),
            value_3d("244", field),
            value_3d("251", field),
            value_11d("21601020001", field),
        )

    def savings_deposit(field: str) -> Decimal | None:
        total = sum_or_none(
            sum_11d("20250", field),
            value_3d("211", field),
            value_3d("215", field),
            sum_11d("21702", field),
            value_11d("21602020001", field),
        )
        return None if total is None else -total

    def corporate_loan(field: str) -> Decimal | None:
        return sum_or_none(
            value_3d("123", field),
            value_11d("13001000002", field),
            value_11d("13003000002", field),
            value_3d("132", field),
            value_3d("136", field),
            None if sum_11d("13604", field) is None else -sum_11d("13604", field),
            value_3d("129", field),
        )

    def personal_loan(field: str) -> Decimal | None:
        return sum_or_none(
            value_3d("122", field),
            value_11d("13001000001", field),
            value_11d("13003000001", field),
            sum_11d("13604", field),
        )

    def credit_card(field: str) -> Decimal | None:
        return sum_11d("13604", field)

    company_spot = company_deposit("期末余额")
    company_year = company_deposit("年日均")
    company_month = company_deposit("月日均")
    savings_spot = savings_deposit("期末余额")
    savings_year = savings_deposit("年日均")
    savings_month = savings_deposit("月日均")
    corporate_loan_spot = corporate_loan("期末余额")
    corporate_loan_year = corporate_loan("年日均")
    corporate_loan_month = corporate_loan("月日均")
    personal_loan_spot = personal_loan("期末余额")
    personal_loan_year = personal_loan("年日均")
    personal_loan_month = personal_loan("月日均")
    credit_card_spot = credit_card("期末余额")
    credit_card_year = credit_card("年日均")
    credit_card_month = credit_card("月日均")

    return [
        amount_row("公司存款合计", None if company_spot is None else -company_spot, None if company_year is None else -company_year, None if company_month is None else -company_month, SEGMENT_BASE_SCALE_SOURCE),
        amount_row("储蓄存款合计", savings_spot, savings_year, savings_month, SEGMENT_BASE_SCALE_SOURCE),
        amount_row("公司贷款合计", corporate_loan_spot, corporate_loan_year, corporate_loan_month, SEGMENT_BASE_SCALE_SOURCE),
        amount_row("个人贷款合计", personal_loan_spot, personal_loan_year, personal_loan_month, SEGMENT_BASE_SCALE_SOURCE),
        # 标准总账对账-日均源缺少 80297 微贷金融支行专段，不能从 merged_data 正确拆出微贷中心。
        amount_row("微贷中心", None, None, None, SEGMENT_BASE_SCALE_MICRO_LOAN_MISSING_SOURCE),
        amount_row("信用卡", credit_card_spot, credit_card_year, credit_card_month, SEGMENT_BASE_SCALE_SOURCE),
        amount_row(
            "存款合计",
            None if company_spot is None or savings_spot is None else -company_spot + savings_spot,
            None if company_year is None or savings_year is None else -company_year + savings_year,
            None if company_month is None or savings_month is None else -company_month + savings_month,
            SEGMENT_BASE_SCALE_SOURCE,
        ),
    ]


def _build_segment_scale_compare_sheet(
    *,
    merged_data: dict[str, Any],
    comparison_data: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not comparison_data:
        return None

    current_rows = {row["指标"]: row for row in _segment_base_scale_raw_rows(merged_data)}
    if not current_rows:
        return None

    comparison_rows_by_kind = {
        kind: {row["指标"]: row for row in _segment_base_scale_raw_rows(data)}
        for kind, data in comparison_data.items()
        if isinstance(data, dict)
    }
    compare_specs = [
        ("时点同比", "时点余额", "prior_year", "上年同期"),
        ("时点环比", "时点余额", "prior_month", "上月"),
        ("年日均同比", "年日均", "prior_year", "上年同期"),
        ("月日均环比", "月日均", "prior_month", "上月"),
    ]
    rows: list[dict[str, Any]] = []
    for metric_name, current_row in current_rows.items():
        for compare_label, field, comparison_key, comparison_caption in compare_specs:
            comparison_rows = comparison_rows_by_kind.get(comparison_key)
            if not comparison_rows:
                continue
            comparison_row = comparison_rows.get(metric_name)
            current_value = _as_decimal(current_row.get(field))
            comparison_value = _as_decimal(comparison_row.get(field)) if comparison_row else None
            delta = None if current_value is None or comparison_value is None else current_value - comparison_value
            current_source = str(current_row.get("口径来源") or "")
            comparison_source = str((comparison_row or {}).get("口径来源") or "")
            source = SEGMENT_SCALE_COMPARE_SOURCE
            if current_source.startswith("source_missing:"):
                source = current_source
            elif comparison_source.startswith("source_missing:"):
                source = comparison_source
            elif comparison_row is None:
                source = f"source_missing: {comparison_caption}缺少同名分部规模行"
            rows.append(
                {
                    "指标": metric_name,
                    "口径": compare_label,
                    "本期": _display_yi(current_value),
                    "对比期": _display_yi(comparison_value),
                    "增减额": _display_yi(delta),
                    "增减幅%": _display_number(_safe_pct(delta, comparison_value)),
                    "口径来源": source,
                }
            )
    if not rows:
        return None
    return _sheet(
        "segment_scale_compare",
        "分部规模同比环比",
        ["指标", "口径", "本期", "对比期", "增减额", "增减幅%", "口径来源"],
        rows,
    )


def _build_company_scale_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
) -> dict[str, Any] | None:
    if not report_month.startswith("2026"):
        return None

    rows = _company_scale_raw_rows(merged_data)
    if not rows:
        return None
    return _sheet(
        "company_scale",
        "公司规模",
        ["指标", "时点余额", "年日均", "月日均", "口径来源"],
        [
            {
                "指标": row["指标"],
                "时点余额": _display_yi(row["时点余额"]),
                "年日均": _display_yi(row["年日均"]),
                "月日均": _display_yi(row["月日均"]),
                "口径来源": row["口径来源"],
            }
            for row in rows
        ],
    )


def _company_scale_raw_rows(merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    rows_3 = {row["科目代码"]: row for row in merged_data.get("3位", [])}
    avg_rows_3 = {row["科目代码"]: row for row in merged_data.get("日均_3位", [])}
    avg_rows_5 = {row["科目代码"]: row for row in merged_data.get("日均_5位", [])}
    rows_11 = list(merged_data.get("11位", []))
    if not rows_3 or not rows_11:
        return []

    def value_3d(code: str, field: str) -> Decimal:
        if field in {"年日均", "月日均"}:
            avg_row = avg_rows_3.get(code)
            if avg_row is not None:
                return _as_decimal(avg_row.get(field)) or ZERO
        row = rows_3.get(code)
        return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)

    def value_11d(code: str, field: str) -> Decimal:
        for row in rows_11:
            if row["科目代码"] == code:
                return _as_decimal(row.get(field)) or ZERO
        return ZERO

    def value_5d(code: str, field: str) -> Decimal:
        if field in {"年日均", "月日均"}:
            row = avg_rows_5.get(code)
            return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)
        return sum(
            (_as_decimal(row.get(field)) or ZERO for row in rows_11 if row["科目代码"].startswith(code)),
            ZERO,
        )

    def company_demand_deposit(field: str) -> Decimal:
        company_203 = value_3d("203", field) if field == "期末余额" else value_11d("20301000001", field)
        company_204 = value_3d("204", field) if field == "期末余额" else value_11d("20401000001", field)
        return -(value_3d("201", field) + company_203 + company_204 + value_3d("243", field))

    def company_term_deposit(field: str) -> Decimal:
        return -(
            value_3d("202", field)
            - value_5d("20250", field)
            + value_3d("205", field)
            + value_3d("225", field)
            + value_3d("244", field)
            + value_3d("251", field)
        )

    def company_structured_deposit(field: str) -> Decimal:
        return -value_5d("21601", field)

    def company_deposit_total(field: str) -> Decimal:
        company_203 = value_3d("203", field) if field == "期末余额" else value_11d("20301000001", field)
        company_204 = value_3d("204", field) if field == "期末余额" else value_11d("20401000001", field)
        return -(
            value_3d("201", field)
            + value_3d("202", field)
            + company_203
            + company_204
            - value_5d("20250", field)
            + value_3d("205", field)
            + value_3d("225", field)
            + value_3d("243", field)
            + value_3d("244", field)
            + value_3d("251", field)
            + value_11d("21601020001", field)
        )

    def company_general_loan(field: str) -> Decimal:
        return (
            value_3d("123", field)
            + value_11d("13001000002", field)
            + value_11d("13003000002", field)
            + value_3d("132", field)
            + value_3d("136", field)
            - value_5d("13604", field)
        )

    def company_bill(field: str) -> Decimal:
        return value_3d("129", field)

    def company_loan_total(field: str) -> Decimal:
        return company_general_loan(field) + company_bill(field)

    def amount_row(name: str, spot: Decimal, year_avg: Decimal, month_avg: Decimal) -> dict[str, Any]:
        return {
            "指标": name,
            "时点余额": spot,
            "年日均": year_avg,
            "月日均": month_avg,
            "口径来源": COMPANY_SCALE_SOURCE,
        }

    return [
        amount_row("公司存款-活期", company_demand_deposit("期末余额"), company_demand_deposit("年日均"), company_demand_deposit("月日均")),
        amount_row("公司存款-定期", company_term_deposit("期末余额"), company_term_deposit("年日均"), company_term_deposit("月日均")),
        amount_row("公司存款-结构性", company_structured_deposit("期末余额"), company_structured_deposit("年日均"), company_structured_deposit("月日均")),
        amount_row("公司存款合计", company_deposit_total("期末余额"), company_deposit_total("年日均"), company_deposit_total("月日均")),
        amount_row("公司贷款-一般贷款", company_general_loan("期末余额"), company_general_loan("年日均"), company_general_loan("月日均")),
        amount_row("公司贷款-票据", company_bill("期末余额"), company_bill("年日均"), company_bill("月日均")),
        amount_row("公司贷款合计", company_loan_total("期末余额"), company_loan_total("年日均"), company_loan_total("月日均")),
    ]


def _build_company_scale_compare_sheet(
    *,
    merged_data: dict[str, Any],
    comparison_data: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not comparison_data:
        return None

    current_rows = {row["指标"]: row for row in _company_scale_raw_rows(merged_data)}
    if not current_rows:
        return None

    comparison_rows_by_kind = {
        kind: {row["指标"]: row for row in _company_scale_raw_rows(data)}
        for kind, data in comparison_data.items()
        if isinstance(data, dict)
    }
    compare_specs = [
        ("时点同比", "时点余额", "prior_year"),
        ("时点环比", "时点余额", "prior_month"),
        ("年日均同比", "年日均", "prior_year"),
        ("月日均环比", "月日均", "prior_month"),
    ]
    rows: list[dict[str, Any]] = []
    for metric_name, current_row in current_rows.items():
        for compare_label, field, comparison_key in compare_specs:
            comparison_rows = comparison_rows_by_kind.get(comparison_key)
            if not comparison_rows:
                continue
            comparison_row = comparison_rows.get(metric_name)
            current_value = _as_decimal(current_row.get(field))
            comparison_value = _as_decimal(comparison_row.get(field)) if comparison_row else None
            delta = None if current_value is None or comparison_value is None else current_value - comparison_value
            source = COMPANY_SCALE_COMPARE_SOURCE
            if comparison_row is None:
                source = "source_missing: 对比期缺少同名公司规模行"
            rows.append(
                {
                    "指标": metric_name,
                    "口径": compare_label,
                    "本期": _display_yi(current_value),
                    "对比期": _display_yi(comparison_value),
                    "增减额": _display_yi(delta),
                    "增减幅%": _display_number(_safe_pct(delta, comparison_value)),
                    "口径来源": source,
                }
            )
    if not rows:
        return None
    return _sheet(
        "company_scale_compare",
        "公司规模同比环比",
        ["指标", "口径", "本期", "对比期", "增减额", "增减幅%", "口径来源"],
        rows,
    )


def _build_retail_scale_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
) -> dict[str, Any] | None:
    if not report_month.startswith("2026"):
        return None

    rows = _retail_scale_raw_rows(merged_data)
    if not rows:
        return None
    return _sheet(
        "retail_scale",
        "零售规模",
        ["指标", "时点余额", "年日均", "月日均", "口径来源"],
        [
            {
                "指标": row["指标"],
                "时点余额": _display_yi(row["时点余额"]),
                "年日均": _display_yi(row["年日均"]),
                "月日均": _display_yi(row["月日均"]),
                "口径来源": row["口径来源"],
            }
            for row in rows
        ],
    )


def _retail_scale_raw_rows(merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    rows_3 = {row["科目代码"]: row for row in merged_data.get("3位", [])}
    avg_rows_3 = {row["科目代码"]: row for row in merged_data.get("日均_3位", [])}
    avg_rows_5 = {row["科目代码"]: row for row in merged_data.get("日均_5位", [])}
    rows_11 = list(merged_data.get("11位", []))
    if not rows_3 or not rows_11:
        return []

    def value_3d(code: str, field: str) -> Decimal:
        if field in {"年日均", "月日均"}:
            avg_row = avg_rows_3.get(code)
            if avg_row is not None:
                return _as_decimal(avg_row.get(field)) or ZERO
        row = rows_3.get(code)
        return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)

    def value_11d(code: str, field: str) -> Decimal:
        for row in rows_11:
            if row["科目代码"] == code:
                return _as_decimal(row.get(field)) or ZERO
        return ZERO

    def value_5d(code: str, field: str) -> Decimal:
        if field in {"年日均", "月日均"}:
            row = avg_rows_5.get(code)
            return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)
        return sum(
            (_as_decimal(row.get(field)) or ZERO for row in rows_11 if row["科目代码"].startswith(code)),
            ZERO,
        )

    def savings_demand_deposit(field: str) -> Decimal:
        return -(value_3d("211", field) + value_5d("21702", field))

    def savings_term_deposit(field: str) -> Decimal:
        return -(value_3d("215", field) + value_5d("20250", field))

    def savings_structured_deposit(field: str) -> Decimal:
        return -value_5d("21602", field)

    def savings_deposit_total(field: str) -> Decimal:
        return -(
            value_5d("20250", field)
            + value_3d("211", field)
            + value_3d("215", field)
            + value_5d("21702", field)
            + value_11d("21602020001", field)
        )

    def personal_loan_total(field: str) -> Decimal:
        return value_3d("122", field) + value_11d("13001000001", field) + value_11d("13003000001", field) + value_5d("13604", field)

    def credit_card(field: str) -> Decimal:
        return value_5d("13604", field)

    def amount_row(name: str, spot: Decimal | None, year_avg: Decimal | None, month_avg: Decimal | None, source: str) -> dict[str, Any]:
        return {
            "指标": name,
            "时点余额": spot,
            "年日均": year_avg,
            "月日均": month_avg,
            "口径来源": source,
        }

    return [
        amount_row("零售存款-活期", savings_demand_deposit("期末余额"), savings_demand_deposit("年日均"), savings_demand_deposit("月日均"), RETAIL_SCALE_SOURCE),
        amount_row("零售存款-定期", savings_term_deposit("期末余额"), savings_term_deposit("年日均"), savings_term_deposit("月日均"), RETAIL_SCALE_SOURCE),
        amount_row("零售存款-结构性", savings_structured_deposit("期末余额"), savings_structured_deposit("年日均"), savings_structured_deposit("月日均"), RETAIL_SCALE_SOURCE),
        amount_row("零售存款合计", savings_deposit_total("期末余额"), savings_deposit_total("年日均"), savings_deposit_total("月日均"), RETAIL_SCALE_SOURCE),
        amount_row("零售贷款-分支行个贷", None, None, None, RETAIL_SCALE_BRANCH_LOAN_MISSING_SOURCE),
        amount_row("参考：微贷中心", None, None, None, SEGMENT_BASE_SCALE_MICRO_LOAN_MISSING_SOURCE),
        amount_row("参考：信用卡", credit_card("期末余额"), credit_card("年日均"), credit_card("月日均"), RETAIL_SCALE_SOURCE),
        amount_row("参考：个人贷款合计", personal_loan_total("期末余额"), personal_loan_total("年日均"), personal_loan_total("月日均"), RETAIL_SCALE_SOURCE),
    ]


def _build_retail_scale_compare_sheet(
    *,
    merged_data: dict[str, Any],
    comparison_data: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not comparison_data:
        return None

    current_rows = {row["指标"]: row for row in _retail_scale_raw_rows(merged_data)}
    if not current_rows:
        return None

    comparison_rows_by_kind = {
        kind: {row["指标"]: row for row in _retail_scale_raw_rows(data)}
        for kind, data in comparison_data.items()
        if isinstance(data, dict)
    }
    compare_specs = [
        ("时点同比", "时点余额", "prior_year"),
        ("时点环比", "时点余额", "prior_month"),
        ("年日均同比", "年日均", "prior_year"),
        ("月日均环比", "月日均", "prior_month"),
    ]
    rows: list[dict[str, Any]] = []
    for metric_name, current_row in current_rows.items():
        for compare_label, field, comparison_key in compare_specs:
            comparison_rows = comparison_rows_by_kind.get(comparison_key)
            if not comparison_rows:
                continue
            comparison_row = comparison_rows.get(metric_name)
            current_value = _as_decimal(current_row.get(field))
            comparison_value = _as_decimal(comparison_row.get(field)) if comparison_row else None
            delta = None if current_value is None or comparison_value is None else current_value - comparison_value
            current_source = str(current_row.get("口径来源") or "")
            comparison_source = str((comparison_row or {}).get("口径来源") or "")
            source = RETAIL_SCALE_COMPARE_SOURCE
            if current_source.startswith("source_missing:"):
                source = current_source
            elif comparison_source.startswith("source_missing:"):
                source = comparison_source
            elif comparison_row is None:
                source = "source_missing: 对比期缺少同名零售规模行"
            rows.append(
                {
                    "指标": metric_name,
                    "口径": compare_label,
                    "本期": _display_yi(current_value),
                    "对比期": _display_yi(comparison_value),
                    "增减额": _display_yi(delta),
                    "增减幅%": _display_number(_safe_pct(delta, comparison_value)),
                    "口径来源": source,
                }
            )
    if not rows:
        return None
    return _sheet(
        "retail_scale_compare",
        "零售规模同比环比",
        ["指标", "口径", "本期", "对比期", "增减额", "增减幅%", "口径来源"],
        rows,
    )


def _build_financial_market_scale_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
) -> dict[str, Any] | None:
    if not report_month.startswith("2026"):
        return None

    rows = _financial_market_scale_raw_rows(merged_data)
    if not rows:
        return None
    return _sheet(
        "financial_market_scale",
        "金融市场规模",
        ["指标", "时点余额", "年日均", "月日均", "口径来源"],
        [
            {
                "指标": row["指标"],
                "时点余额": _display_yi(row["时点余额"]),
                "年日均": _display_yi(row["年日均"]),
                "月日均": _display_yi(row["月日均"]),
                "口径来源": row["口径来源"],
            }
            for row in rows
        ],
    )


def _financial_market_scale_raw_rows(merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    rows_3 = {row["科目代码"]: row for row in merged_data.get("3位", [])}
    avg_rows_3 = {row["科目代码"]: row for row in merged_data.get("日均_3位", [])}
    avg_rows_5 = {row["科目代码"]: row for row in merged_data.get("日均_5位", [])}
    rows_11 = list(merged_data.get("11位", []))
    if not rows_3 or not rows_11:
        return []

    def value_3d(code: str, field: str) -> Decimal:
        if field in {"年日均", "月日均"}:
            avg_row = avg_rows_3.get(code)
            if avg_row is not None:
                return _as_decimal(avg_row.get(field)) or ZERO
        row = rows_3.get(code)
        return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)

    def sum_3d(codes: tuple[str, ...], field: str) -> Decimal:
        return sum((value_3d(code, field) for code in codes), ZERO)

    def sum_11d(prefix: str, field: str) -> Decimal:
        return sum(
            (_as_decimal(row.get(field)) or ZERO for row in rows_11 if row["科目代码"].startswith(prefix)),
            ZERO,
        )

    def value_5d(code: str, field: str) -> Decimal:
        if field in {"年日均", "月日均"}:
            row = avg_rows_5.get(code)
            return ZERO if row is None else (_as_decimal(row.get(field)) or ZERO)
        return ZERO

    def interest_earning_bonds(field: str) -> Decimal:
        return (
            sum_3d(("142", "143", "144"), field)
            - sum_11d("14301010001", field)
            - sum_11d("14301010002", field)
        )

    def fvtpl(field: str) -> Decimal:
        return value_3d("141", field)

    def interbank_assets(field: str) -> Decimal:
        return (
            sum_3d(("120", "121", "140"), field)
            - value_5d("14004", field)
            - value_5d("14005", field)
        )

    def interbank_liabilities(field: str) -> Decimal:
        return -(
            sum_3d(("234", "235", "241", "242", "255"), field)
            + sum_11d("27205000001", field)
            + sum_11d("27206000001", field)
        )

    def amount_row(name: str, spot: Decimal, year_avg: Decimal, month_avg: Decimal) -> dict[str, Any]:
        return {
            "指标": name,
            "时点余额": spot,
            "年日均": year_avg,
            "月日均": month_avg,
            "口径来源": FINANCIAL_MARKET_SCALE_SOURCE,
        }

    return [
        amount_row("生息债券投资", interest_earning_bonds("期末余额"), interest_earning_bonds("年日均"), interest_earning_bonds("月日均")),
        amount_row("FVTPL", fvtpl("期末余额"), fvtpl("年日均"), fvtpl("月日均")),
        amount_row("同业资产", interbank_assets("期末余额"), interbank_assets("年日均"), interbank_assets("月日均")),
        amount_row("同业负债", interbank_liabilities("期末余额"), interbank_liabilities("年日均"), interbank_liabilities("月日均")),
    ]


def _build_financial_market_scale_compare_sheet(
    *,
    merged_data: dict[str, Any],
    comparison_data: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not comparison_data:
        return None

    current_rows = {row["指标"]: row for row in _financial_market_scale_raw_rows(merged_data)}
    if not current_rows:
        return None

    comparison_rows_by_kind = {
        kind: {row["指标"]: row for row in _financial_market_scale_raw_rows(data)}
        for kind, data in comparison_data.items()
        if isinstance(data, dict)
    }
    compare_specs = [
        ("时点同比", "时点余额", "prior_year"),
        ("时点环比", "时点余额", "prior_month"),
        ("年日均同比", "年日均", "prior_year"),
        ("月日均环比", "月日均", "prior_month"),
    ]
    rows: list[dict[str, Any]] = []
    for metric_name, current_row in current_rows.items():
        for compare_label, field, comparison_key in compare_specs:
            comparison_rows = comparison_rows_by_kind.get(comparison_key)
            if not comparison_rows:
                continue
            comparison_row = comparison_rows.get(metric_name)
            current_value = _as_decimal(current_row.get(field))
            comparison_value = _as_decimal(comparison_row.get(field)) if comparison_row else None
            delta = None if current_value is None or comparison_value is None else current_value - comparison_value
            source = FINANCIAL_MARKET_SCALE_COMPARE_SOURCE
            if comparison_row is None:
                source = "source_missing: 对比期缺少同名金融市场规模行"
            rows.append(
                {
                    "指标": metric_name,
                    "口径": compare_label,
                    "本期": _display_yi(current_value),
                    "对比期": _display_yi(comparison_value),
                    "增减额": _display_yi(delta),
                    "增减幅%": _display_number(_safe_pct(delta, comparison_value)),
                    "口径来源": source,
                }
            )
    if not rows:
        return None
    return _sheet(
        "financial_market_scale_compare",
        "金融市场规模同比环比",
        ["指标", "口径", "本期", "对比期", "增减额", "增减幅%", "口径来源"],
        rows,
    )


def _build_income_rate_analysis_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
) -> dict[str, Any] | None:
    if not report_month.startswith("2026"):
        return None

    rows = _income_rate_raw_rows(report_month=report_month, merged_data=merged_data)
    if not rows:
        return None
    return _sheet(
        "income_rate_analysis",
        "收益率分析（总账可复算）",
        ["指标", "板块", "收益类别", "年日均规模", "总账收益/支出", "年化收益率/付息率%", "口径来源"],
        [
            {
                "指标": row["指标"],
                "板块": row["板块"],
                "收益类别": row["收益类别"],
                "年日均规模": _display_yi(row["年日均规模"]),
                "总账收益/支出": _display_yi(row["总账收益/支出"]),
                "年化收益率/付息率%": _display_number(row["年化收益率/付息率%"]),
                "口径来源": row["口径来源"],
            }
            for row in rows
        ],
    )


def _income_rate_raw_rows(*, report_month: str, merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    rows_3 = {row["科目代码"]: row for row in merged_data.get("3位", [])}
    rows_11 = list(merged_data.get("11位", []))
    if not rows_3 or not rows_11:
        return []

    days = _ytd_days(report_month)
    company_rows = {row["指标"]: row for row in _company_scale_raw_rows(merged_data)}
    retail_rows = {row["指标"]: row for row in _retail_scale_raw_rows(merged_data)}

    def value_3d(code: str) -> Decimal:
        row = rows_3.get(code)
        return ZERO if row is None else (_as_decimal(row.get("期末余额")) or ZERO)

    def sum_11d(prefix: str) -> Decimal:
        return sum(
            (_as_decimal(row.get("期末余额")) or ZERO for row in rows_11 if row["科目代码"].startswith(prefix)),
            ZERO,
        )

    def value_11d(code: str) -> Decimal:
        for row in rows_11:
            if row["科目代码"] == code:
                return _as_decimal(row.get("期末余额")) or ZERO
        return ZERO

    company_loan_income = -(
        value_3d("501")
        - sum_11d("50109")
        - sum_11d("50110")
        + value_11d("50110000002")
        + sum_11d("50206")
        - sum_11d("50115")
    )
    personal_loan_income = -(sum_11d("50109") + sum_11d("50110") - value_11d("50110000002"))
    discount_callback = -sum_11d("50115")
    company_deposit_expense = (
        sum_11d("52101")
        + sum_11d("52102")
        + sum_11d("52110")
        - value_11d("52110000001")
        + value_11d("52105000001")
        + value_11d("52106000002")
    )
    savings_deposit_expense = (
        sum_11d("52103")
        + sum_11d("52104")
        + value_11d("52105000002")
        + value_11d("52106000001")
        + value_11d("52110000001")
    )

    def scale(metric_name: str, source_rows: dict[str, dict[str, Any]]) -> Decimal | None:
        row = source_rows.get(metric_name)
        return None if row is None else _as_decimal(row.get("年日均"))

    def amount_row(
        *,
        name: str,
        segment: str,
        income_type: str,
        year_avg_scale: Decimal | None,
        amount: Decimal | None,
        source: str,
    ) -> dict[str, Any]:
        return {
            "指标": name,
            "板块": segment,
            "收益类别": income_type,
            "年日均规模": year_avg_scale,
            "总账收益/支出": amount,
            "年化收益率/付息率%": _annualized_rate_pct(amount, year_avg_scale, days),
            "口径来源": source,
        }

    return [
        amount_row(
            name="公司贷款利息收入",
            segment="公司板块",
            income_type="贷款利息收入",
            year_avg_scale=scale("公司贷款合计", company_rows),
            amount=company_loan_income,
            source=INCOME_RATE_ANALYSIS_SOURCE,
        ),
        amount_row(
            name="个人贷款利息收入",
            segment="参考：个人贷款总量",
            income_type="贷款利息收入",
            year_avg_scale=None,
            amount=personal_loan_income,
            source=INCOME_RATE_PERSONAL_LOAN_SCALE_MISSING_SOURCE,
        ),
        amount_row(
            name="公司存款利息支出",
            segment="公司板块",
            income_type="存款利息支出",
            year_avg_scale=scale("公司存款合计", company_rows),
            amount=company_deposit_expense,
            source=INCOME_RATE_ANALYSIS_SOURCE,
        ),
        amount_row(
            name="储蓄存款利息支出",
            segment="零售存款",
            income_type="存款利息支出",
            year_avg_scale=scale("零售存款合计", retail_rows),
            amount=savings_deposit_expense,
            source=INCOME_RATE_ANALYSIS_SOURCE,
        ),
        amount_row(
            name="折现回拨",
            segment="贷款折现",
            income_type="贷款利息收入调节项",
            year_avg_scale=None,
            amount=discount_callback,
            source=INCOME_RATE_ANALYSIS_SOURCE,
        ),
        amount_row(
            name="金融投资利息收入",
            segment="金融市场",
            income_type="金融投资利息收入",
            year_avg_scale=None,
            amount=None,
            source=INCOME_RATE_MISSING_SOURCE,
        ),
        amount_row(
            name="同业资产负债利息净收入",
            segment="金融市场",
            income_type="同业资产负债利息净收入",
            year_avg_scale=None,
            amount=None,
            source=INCOME_RATE_MISSING_SOURCE,
        ),
    ]


def _build_income_rate_attribution_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
    comparison_data: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not comparison_data:
        return None
    prior_year_data = comparison_data.get("prior_year")
    if not isinstance(prior_year_data, dict):
        return None

    current_rows = {
        row["指标"]: row for row in _income_rate_raw_rows(report_month=report_month, merged_data=merged_data)
    }
    prior_year_month = f"{int(report_month[:4]) - 1}{report_month[4:]}"
    prior_rows = {
        row["指标"]: row
        for row in _income_rate_raw_rows(report_month=prior_year_month, merged_data=prior_year_data)
    }
    current_days = _ytd_days(report_month)
    prior_days = _ytd_days(prior_year_month)
    rows: list[dict[str, Any]] = []
    for metric_name, current_row in current_rows.items():
        prior_row = prior_rows.get(metric_name)
        if prior_row is None:
            continue
        current_amount = _as_decimal(current_row.get("总账收益/支出"))
        prior_amount = _as_decimal(prior_row.get("总账收益/支出"))
        current_scale = _as_decimal(current_row.get("年日均规模"))
        prior_scale = _as_decimal(prior_row.get("年日均规模"))
        current_rate = _annualized_rate(current_amount, current_scale, current_days)
        prior_rate = _annualized_rate(prior_amount, prior_scale, prior_days)
        source = INCOME_RATE_ATTRIBUTION_SOURCE
        volume_effect = None
        rate_effect = None
        delta = None if current_amount is None or prior_amount is None else current_amount - prior_amount
        if None in {current_scale, prior_scale, current_rate, prior_rate}:
            if str(current_row.get("口径来源") or "").startswith("source_missing:"):
                source = str(current_row["口径来源"])
            elif str(prior_row.get("口径来源") or "").startswith("source_missing:"):
                source = str(prior_row["口径来源"])
            else:
                source = "source_missing: 收益量价归因缺少年日均规模或收益率"
        else:
            volume_effect = (current_scale - prior_scale) * current_rate * Decimal(current_days) / Decimal(365)
            rate_effect = prior_scale * (
                current_rate * Decimal(current_days) / Decimal(365)
                - prior_rate * Decimal(prior_days) / Decimal(365)
            )
        check_gap = None if delta is None or volume_effect is None or rate_effect is None else delta - volume_effect - rate_effect
        rows.append(
            {
                "指标": metric_name,
                "板块": current_row["板块"],
                "本期收益/支出": _display_yi(current_amount),
                "对比期收益/支出": _display_yi(prior_amount),
                "增减额": _display_yi(delta),
                "规模贡献": _display_yi(volume_effect),
                "利率贡献": _display_yi(rate_effect),
                "校验差异": _display_yi(check_gap),
                "口径来源": source,
            }
        )
    if not rows:
        return None
    return _sheet(
        "income_rate_attribution",
        "收益量价归因（年累计同比）",
        ["指标", "板块", "本期收益/支出", "对比期收益/支出", "增减额", "规模贡献", "利率贡献", "校验差异", "口径来源"],
        rows,
    )


def _build_deposit_interest_split_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
    comparison_data: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not report_month.startswith("2026"):
        return None

    current_rows = _deposit_interest_split_raw_rows(merged_data)
    if not current_rows:
        return None

    prior_year_rows = _raw_rows_by_metric(
        _deposit_interest_split_raw_rows((comparison_data or {}).get("prior_year", {}))
    )
    prior_month_rows = _raw_rows_by_metric(
        _deposit_interest_split_raw_rows((comparison_data or {}).get("prior_month", {}))
    )
    two_months_ago_rows = _raw_rows_by_metric(
        _deposit_interest_split_raw_rows((comparison_data or {}).get("two_months_ago", {}))
    )

    ytd_days = _ytd_days(report_month)
    current_month_days = monthrange(int(report_month[:4]), int(report_month[4:6]))[1]
    prior_month_key = _shift_report_month_key(report_month, -1)

    rows: list[dict[str, Any]] = []
    for row in current_rows:
        metric_name = row["指标"]
        prior_year_row = prior_year_rows.get(metric_name)
        prior_month_row = prior_month_rows.get(metric_name)
        two_months_ago_row = two_months_ago_rows.get(metric_name)
        current_amount = _as_decimal(row.get("年累计利息支出"))
        prior_year_amount = _as_decimal(prior_year_row.get("年累计利息支出")) if prior_year_row else None
        prior_month_ytd_amount = _as_decimal(prior_month_row.get("年累计利息支出")) if prior_month_row else None
        two_months_ago_ytd_amount = (
            _as_decimal(two_months_ago_row.get("年累计利息支出")) if two_months_ago_row else None
        )
        current_month_amount = _month_amount(report_month, current_amount, prior_month_ytd_amount)
        prior_month_amount = (
            _month_amount(prior_month_key, prior_month_ytd_amount, two_months_ago_ytd_amount)
            if prior_month_key is not None
            else None
        )
        yoy_delta = None if current_amount is None or prior_year_amount is None else current_amount - prior_year_amount
        mom_delta = (
            None
            if current_month_amount is None or prior_month_amount is None
            else current_month_amount - prior_month_amount
        )
        rows.append(
            {
                "指标": metric_name,
                "板块": row["板块"],
                "本期年日均": _display_yi(_as_decimal(row.get("年日均规模"))),
                "年累计利息支出": _display_yi(current_amount),
                "年化付息率%": _display_number(
                    _annualized_rate_pct(current_amount, _as_decimal(row.get("年日均规模")), ytd_days)
                ),
                "同比增减额": _display_yi(yoy_delta),
                "本月月日均": _display_yi(_as_decimal(row.get("月日均规模"))),
                "本月利息支出": _display_yi(current_month_amount),
                "本月付息率%": _display_number(
                    _annualized_rate_pct(
                        current_month_amount,
                        _as_decimal(row.get("月日均规模")),
                        current_month_days,
                    )
                ),
                "环比增减额": _display_yi(mom_delta),
                "口径来源": row["口径来源"],
            }
        )

    return _sheet(
        "deposit_interest_split",
        "存款利息拆分",
        [
            "指标",
            "板块",
            "本期年日均",
            "年累计利息支出",
            "年化付息率%",
            "同比增减额",
            "本月月日均",
            "本月利息支出",
            "本月付息率%",
            "环比增减额",
            "口径来源",
        ],
        rows,
    )


def _deposit_interest_split_raw_rows(merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    if not isinstance(merged_data, dict):
        return []
    rows_11 = list(merged_data.get("11位", []))
    if not rows_11:
        return []

    company_rows = {row["指标"]: row for row in _company_scale_raw_rows(merged_data)}
    retail_rows = {row["指标"]: row for row in _retail_scale_raw_rows(merged_data)}

    def sum_11d(prefix: str) -> Decimal:
        return sum(
            (_as_decimal(row.get("期末余额")) or ZERO for row in rows_11 if row["科目代码"].startswith(prefix)),
            ZERO,
        )

    def value_11d(code: str) -> Decimal:
        for row in rows_11:
            if row["科目代码"] == code:
                return _as_decimal(row.get("期末余额")) or ZERO
        return ZERO

    def scale(metric_name: str, source_rows: dict[str, dict[str, Any]], field: str) -> Decimal | None:
        row = source_rows.get(metric_name)
        return None if row is None else _as_decimal(row.get(field))

    def sum_scale(
        left_metric: str,
        left_rows: dict[str, dict[str, Any]],
        right_metric: str,
        right_rows: dict[str, dict[str, Any]],
        field: str,
    ) -> Decimal | None:
        left = scale(left_metric, left_rows, field)
        right = scale(right_metric, right_rows, field)
        if left is None or right is None:
            return None
        return left + right

    def amount_row(
        name: str,
        segment: str,
        scale_metric: str,
        scale_rows: dict[str, dict[str, Any]],
        amount: Decimal,
    ) -> dict[str, Any]:
        return {
            "指标": name,
            "板块": segment,
            "年日均规模": scale(scale_metric, scale_rows, "年日均"),
            "月日均规模": scale(scale_metric, scale_rows, "月日均"),
            "年累计利息支出": amount,
            "口径来源": DEPOSIT_INTEREST_SPLIT_SOURCE,
        }

    company_demand = sum_11d("52101")
    company_term = sum_11d("52102") + value_11d("52105000001") + sum_11d("52110") - value_11d("52110000001")
    company_structured = value_11d("52106000002")
    savings_demand = sum_11d("52103") + value_11d("52110000001")
    savings_term = sum_11d("52104") + value_11d("52105000002")
    savings_structured = value_11d("52106000001")
    total_amount = company_demand + company_term + company_structured + savings_demand + savings_term + savings_structured

    return [
        amount_row("公司存款", "公司存款", "公司存款合计", company_rows, company_demand + company_term + company_structured),
        amount_row("公司存款-活期", "公司存款", "公司存款-活期", company_rows, company_demand),
        amount_row("公司存款-定期", "公司存款", "公司存款-定期", company_rows, company_term),
        amount_row("公司存款-结构性", "公司存款", "公司存款-结构性", company_rows, company_structured),
        amount_row("储蓄存款", "零售存款", "零售存款合计", retail_rows, savings_demand + savings_term + savings_structured),
        amount_row("储蓄存款-活期", "零售存款", "零售存款-活期", retail_rows, savings_demand),
        amount_row("储蓄存款-定期", "零售存款", "零售存款-定期", retail_rows, savings_term),
        amount_row("储蓄存款-结构性", "零售存款", "零售存款-结构性", retail_rows, savings_structured),
        {
            "指标": "存款利息支出合计",
            "板块": "母公司",
            "年日均规模": sum_scale("公司存款合计", company_rows, "零售存款合计", retail_rows, "年日均"),
            "月日均规模": sum_scale("公司存款合计", company_rows, "零售存款合计", retail_rows, "月日均"),
            "年累计利息支出": total_amount,
            "口径来源": DEPOSIT_INTEREST_SPLIT_SOURCE,
        },
    ]


def _build_parent_company_revenue_components_sheet(
    *,
    report_month: str,
    merged_data: dict[str, Any],
    comparison_data: dict[str, dict[str, Any]] | None,
) -> dict[str, Any] | None:
    if not report_month.startswith("2026"):
        return None

    current_rows = _parent_company_revenue_raw_rows(merged_data)
    if not current_rows:
        return None

    prior_year_rows = _raw_rows_by_metric(
        _parent_company_revenue_raw_rows((comparison_data or {}).get("prior_year", {}))
    )
    prior_month_rows = _raw_rows_by_metric(
        _parent_company_revenue_raw_rows((comparison_data or {}).get("prior_month", {}))
    )
    two_months_ago_rows = _raw_rows_by_metric(
        _parent_company_revenue_raw_rows((comparison_data or {}).get("two_months_ago", {}))
    )
    prior_month_key = _shift_report_month_key(report_month, -1)

    rows: list[dict[str, Any]] = []
    for row in current_rows:
        metric_name = row["指标"]
        prior_year_row = prior_year_rows.get(metric_name)
        prior_month_row = prior_month_rows.get(metric_name)
        two_months_ago_row = two_months_ago_rows.get(metric_name)
        current_ytd = _as_decimal(row.get("年累计金额"))
        prior_year_ytd = _as_decimal(prior_year_row.get("年累计金额")) if prior_year_row else None
        prior_month_ytd = _as_decimal(prior_month_row.get("年累计金额")) if prior_month_row else None
        two_months_ago_ytd = _as_decimal(two_months_ago_row.get("年累计金额")) if two_months_ago_row else None
        current_month = _month_amount(report_month, current_ytd, prior_month_ytd)
        prior_month = (
            _month_amount(prior_month_key, prior_month_ytd, two_months_ago_ytd)
            if prior_month_key is not None
            else None
        )
        yoy_delta = None if current_ytd is None or prior_year_ytd is None else current_ytd - prior_year_ytd
        mom_delta = None if current_month is None or prior_month is None else current_month - prior_month
        rows.append(
            {
                "指标": metric_name,
                "类别": row["类别"],
                "同比本期": _display_yi(current_ytd),
                "同比对比期": _display_yi(prior_year_ytd),
                "同比增减额": _display_yi(yoy_delta),
                "同比增减幅%": _display_number(_safe_pct(yoy_delta, prior_year_ytd)),
                "环比本月": _display_yi(current_month),
                "环比上月": _display_yi(prior_month),
                "环比增减额": _display_yi(mom_delta),
                "环比增减幅%": _display_number(_safe_pct(mom_delta, prior_month)),
                "口径来源": row["口径来源"],
            }
        )

    return _sheet(
        "parent_company_revenue_components",
        "母公司营收分项",
        [
            "指标",
            "类别",
            "同比本期",
            "同比对比期",
            "同比增减额",
            "同比增减幅%",
            "环比本月",
            "环比上月",
            "环比增减额",
            "环比增减幅%",
            "口径来源",
        ],
        rows,
    )


def _parent_company_revenue_raw_rows(merged_data: dict[str, Any]) -> list[dict[str, Any]]:
    if not isinstance(merged_data, dict):
        return []
    components = _interest_component_values(merged_data)
    if components is None:
        return []

    company_loan_income = components["company_loan_income"]
    personal_loan_income = components["personal_loan_income"]
    discount_callback = components["discount_callback"]
    company_deposit_expense = components["company_deposit_expense"]
    savings_deposit_expense = components["savings_deposit_expense"]
    loan_interest_income = company_loan_income + personal_loan_income + discount_callback
    deposit_interest_expense = company_deposit_expense + savings_deposit_expense

    def row(name: str, category: str, amount: Decimal | None, source: str) -> dict[str, Any]:
        return {
            "指标": name,
            "类别": category,
            "年累计金额": amount,
            "口径来源": source,
        }

    return [
        row("贷款利息收入", "利息净收入", loan_interest_income, PARENT_COMPANY_REVENUE_SOURCE),
        row("公司贷款", "贷款利息收入", company_loan_income, PARENT_COMPANY_REVENUE_SOURCE),
        row("个人贷款", "贷款利息收入", personal_loan_income, PARENT_COMPANY_REVENUE_SOURCE),
        row("折现回拨", "贷款利息收入调节项", discount_callback, PARENT_COMPANY_REVENUE_SOURCE),
        row("利息冲减", "贷款利息收入调节项", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("存款利息支出", "利息净收入", deposit_interest_expense, PARENT_COMPANY_REVENUE_SOURCE),
        row("公司存款", "存款利息支出", company_deposit_expense, PARENT_COMPANY_REVENUE_SOURCE),
        row("储蓄存款", "存款利息支出", savings_deposit_expense, PARENT_COMPANY_REVENUE_SOURCE),
        row("金融投资利息收入", "利息净收入", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("同业资产负债利息净收入", "利息净收入", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("利息净收入", "母公司营收", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("非息净收入", "母公司营收", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("中间业务净收入", "非息净收入", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("汇兑损益", "非息净收入", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("估值及投资收益", "非息净收入", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("其他收入", "非息净收入", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
        row("母公司营业收入合计", "母公司营收", None, PARENT_COMPANY_REVENUE_MISSING_SOURCE),
    ]


def _interest_component_values(merged_data: dict[str, Any]) -> dict[str, Decimal] | None:
    rows_3 = {row["科目代码"]: row for row in merged_data.get("3位", [])}
    rows_11 = list(merged_data.get("11位", []))
    if not rows_3 or not rows_11:
        return None

    def value_3d(code: str) -> Decimal:
        row = rows_3.get(code)
        return ZERO if row is None else (_as_decimal(row.get("期末余额")) or ZERO)

    def sum_11d(prefix: str) -> Decimal:
        return sum(
            (_as_decimal(row.get("期末余额")) or ZERO for row in rows_11 if row["科目代码"].startswith(prefix)),
            ZERO,
        )

    def value_11d(code: str) -> Decimal:
        for row in rows_11:
            if row["科目代码"] == code:
                return _as_decimal(row.get("期末余额")) or ZERO
        return ZERO

    return {
        "company_loan_income": -(
            value_3d("501")
            - sum_11d("50109")
            - sum_11d("50110")
            + value_11d("50110000002")
            + sum_11d("50206")
            - sum_11d("50115")
        ),
        "personal_loan_income": -(sum_11d("50109") + sum_11d("50110") - value_11d("50110000002")),
        "discount_callback": -sum_11d("50115"),
        "company_deposit_expense": (
            sum_11d("52101")
            + sum_11d("52102")
            + sum_11d("52110")
            - value_11d("52110000001")
            + value_11d("52105000001")
            + value_11d("52106000002")
        ),
        "savings_deposit_expense": (
            sum_11d("52103")
            + sum_11d("52104")
            + value_11d("52105000002")
            + value_11d("52106000001")
            + value_11d("52110000001")
        ),
    }


def _raw_rows_by_metric(rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {row["指标"]: row for row in rows}


def _month_amount(
    report_month: str | None,
    ytd_amount: Decimal | None,
    prior_month_ytd_amount: Decimal | None,
) -> Decimal | None:
    if ytd_amount is None or report_month is None:
        return None
    if report_month[4:6] == "01":
        return ytd_amount
    if prior_month_ytd_amount is None:
        return None
    return ytd_amount - prior_month_ytd_amount


def _shift_report_month_key(report_month: str, month_delta: int) -> str | None:
    if len(report_month) != 6 or not report_month.isdigit():
        return None
    year = int(report_month[:4])
    month = int(report_month[4:])
    if month < 1 or month > 12:
        return None
    month_index = year * 12 + month - 1 + month_delta
    shifted_year = month_index // 12
    shifted_month = month_index % 12 + 1
    return f"{shifted_year:04d}{shifted_month:02d}"


def _ytd_days(month_key: str) -> int:
    year = int(month_key[:4])
    month = int(month_key[4:6])
    return sum(monthrange(year, month_number)[1] for month_number in range(1, month + 1))


def _annualized_rate(amount: Decimal | None, average_scale: Decimal | None, days: int) -> Decimal | None:
    if amount is None or average_scale is None or average_scale == ZERO or days <= 0:
        return None
    return amount / average_scale * Decimal(365) / Decimal(days)


def _annualized_rate_pct(amount: Decimal | None, average_scale: Decimal | None, days: int) -> Decimal | None:
    rate = _annualized_rate(amount, average_scale, days)
    return None if rate is None else rate * Decimal(100)


def _industry_sheet(key: str, title: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    ordered = sorted(rows, key=lambda row: abs(_as_decimal(row.get("期末余额")) or ZERO), reverse=True)
    return _sheet(key, title, ["行业名称", "期初余额", "期末余额", "变动额", "月日均", "年日均", "偏离额", "偏离%", "趋势%"], [
        {"行业名称": row.get("行业名称", ""), "期初余额": _display_number(_to_yi(_as_decimal(row.get("期初余额")) or ZERO)), "期末余额": _display_number(_to_yi(_as_decimal(row.get("期末余额")) or ZERO)), "变动额": _display_number(_to_yi(_as_decimal(row.get("变动额")) or ZERO)), "月日均": _display_number(_to_yi(_as_decimal(row.get("月日均")) or ZERO)), "年日均": _display_number(_to_yi(_as_decimal(row.get("年日均")) or ZERO)), "偏离额": _display_number(_to_yi(_as_decimal(row.get("偏离额")) or ZERO)), "偏离%": _display_number(_as_decimal(row.get("偏离%"))), "趋势%": _display_number(_as_decimal(row.get("趋势%")))}
        for row in ordered
    ])


def _normalize_amount_row(row: dict[str, Any], *, include_trend: bool) -> dict[str, Any]:
    normalized = {
        "科目代码": row["科目代码"],
        "名称": row.get("名称", ""),
        "期初余额": _display_number(_to_yi(_as_decimal(row.get("期初余额")) or ZERO)),
        "期末余额": _display_number(_to_yi(_as_decimal(row.get("期末余额")) or ZERO)),
        "变动额": _display_number(_to_yi(_as_decimal(row.get("变动额")) or ZERO)),
        "月日均": _display_number(_to_yi(_as_decimal(row.get("月日均")) or ZERO)),
        "年日均": _display_number(_to_yi(_as_decimal(row.get("年日均")) or ZERO)),
        "偏离额": _display_number(_to_yi(_as_decimal(row.get("偏离额")) or ZERO)),
        "偏离%": _display_number(_as_decimal(row.get("偏离%"))),
    }
    if include_trend:
        normalized["趋势额"] = _display_number(_to_yi(_as_decimal(row.get("趋势额")) or ZERO))
        normalized["趋势%"] = _display_number(_as_decimal(row.get("趋势%")))
    return normalized


def _sheet(key: str, title: str, columns: list[str], rows: list[dict[str, Any]]) -> dict[str, Any]:
    return {"key": key, "title": title, "columns": columns, "rows": rows}


def _group_sum(rows: list[dict[str, Any]], *, key: str, include_name: bool = False) -> list[dict[str, Any]]:
    buckets: dict[str, dict[str, Any]] = {}
    for row in rows:
        group_key = row[key]
        bucket = buckets.setdefault(group_key, {"group_key": group_key, "期初余额": ZERO, "期末余额": ZERO, "变动额": ZERO, "本期借方": ZERO, "本期贷方": ZERO})
        for field in ("期初余额", "期末余额", "变动额", "本期借方", "本期贷方"):
            bucket[field] += _as_decimal(row.get(field)) or ZERO
        if include_name and not bucket.get("科目名称"):
            bucket["科目名称"] = row.get("科目名称", "")
    return list(buckets.values())


def _daily_average_level_rows(
    *,
    month_avg: dict[str, Decimal],
    year_avg: dict[str, Decimal],
) -> list[dict[str, Any]]:
    return [
        {
            "科目代码": code,
            "月日均": month_avg.get(code),
            "年日均": year_avg.get(code),
        }
        for code in sorted(set(month_avg) | set(year_avg))
    ]


def _index_amounts(rows: list[dict[str, Any]]) -> dict[str, Decimal]:
    return {row["科目代码"]: _as_decimal(row.get("日均余额")) or ZERO for row in rows}


def _sort_rows(rows: list[dict[str, Any]], field: str) -> list[dict[str, Any]]:
    return sorted(rows, key=lambda row: abs(_as_decimal(row.get(field)) or ZERO), reverse=True)


def _normalize_account_code(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        if not value.is_integer():
            return None
        return str(int(value))
    text = str(value).strip()
    return text or None


def _to_decimal(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _as_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    return _to_decimal(value)


def _safe_pct(numerator: Decimal | None, denominator: Decimal | None) -> Decimal | None:
    if numerator is None or denominator is None or denominator == ZERO:
        return None
    return numerator / abs(denominator) * Decimal("100")


def _pct(numerator: Decimal, denominator: Decimal) -> Decimal:
    return ZERO if denominator == ZERO else numerator / denominator * Decimal("100")


def _to_yi(value: Decimal | None) -> Decimal:
    return (value or ZERO) / ONE_HUNDRED_MILLION


def _display_yi(value: Decimal | None) -> int | float | None:
    return None if value is None else _display_number(_to_yi(value))


def _display_number(value: Decimal | None) -> int | float | None:
    if value is None:
        return None
    normalized = value.quantize(Decimal("0.01"))
    return int(normalized) if normalized == normalized.to_integral_value() else float(normalized)


def _excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value
