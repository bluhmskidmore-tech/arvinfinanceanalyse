from __future__ import annotations

import hashlib
from collections import Counter
from functools import lru_cache
from pathlib import Path

import duckdb
from openpyxl import load_workbook
import xlrd

from backend.app.repositories.duckdb_migrations import apply_pending_migrations_on_connection
from backend.app.repositories.governance_repo import (
    SOURCE_MANIFEST_STREAM,
    GovernanceRepository,
)
from backend.app.schemas.source_preview import (
    PreviewColumn,
    PreviewRowPage,
    PnlPreviewRow,
    RuleTracePage,
    NonstdPnlPreviewRow,
    SourcePreviewHistoryPage,
    SourcePreviewPayload,
    SourcePreviewSummary,
    TywPreviewRow,
    ZqtzPreviewRow,
)
from backend.app.services.source_rules import (
    classify_nonstd_pnl_preview,
    classify_pnl_preview,
    classify_tyw_preview,
    classify_zqtz_preview,
    describe_source_file,
)

RULE_VERSION = "rv_phase1_source_preview_v1"
MANIFEST_ELIGIBLE_STATUSES = {"completed", "rerun"}
PREVIEW_TABLES = (
    "phase1_source_preview_summary",
    "phase1_source_preview_groups",
    "phase1_zqtz_preview_rows",
    "phase1_tyw_preview_rows",
    "phase1_pnl_preview_rows",
    "phase1_nonstd_pnl_preview_rows",
    "phase1_zqtz_rule_traces",
    "phase1_tyw_rule_traces",
    "phase1_pnl_rule_traces",
    "phase1_nonstd_pnl_rule_traces",
)
SUPPORTED_PREVIEW_SOURCE_FAMILIES = frozenset(
    {"zqtz", "tyw", "pnl", "pnl_514", "pnl_516", "pnl_517"}
)

ZQTZ_BOND_CODE = "\u503a\u5238\u4ee3\u53f7"
ZQTZ_BOND_NAME = "\u503a\u5238\u540d\u79f0"
ZQTZ_DATE = "\u65e5\u671f"
ZQTZ_BUSINESS_TYPE1 = "\u4e1a\u52a1\u79cd\u7c7b1"
ZQTZ_ACCOUNT_CATEGORY = "\u8d26\u6237\u7c7b\u522b"

TYW_PRODUCT_TYPE = "\u4ea7\u54c1\u7c7b\u578b"
TYW_COUNTERPARTY_NAME = "\u5bf9\u624b\u65b9\u540d\u79f0"
TYW_INVESTMENT_PORTFOLIO = "\u6295\u8d44\u7ec4\u5408"
TYW_CBIRC_TYPE = "\u4f1a\u8ba1\u7c7b\u578b_\u94f6\u4fdd\u76d1\u4f1a"
TYW_PBOC_TYPE = "\u4f1a\u8ba1\u7c7b\u578b_\u4eba\u884c"
TYW_CORE_CUSTOMER_TYPE = "\u6838\u5fc3\u5ba2\u6237\u7c7b\u578b"
TYW_ACCOUNT_TYPE = "\u8d26\u6237\u7c7b\u578b"
TYW_SPECIAL_ACCOUNT_TYPE = "\u7279\u6b8a\u8d26\u6237\u7c7b\u578b"
TYW_CUSTODY_ACCOUNT_NAME = "\u6258\u7ba1\u8d26\u6237\u540d\u79f0"
TYW_TRACE_FIELDS = {
    TYW_PRODUCT_TYPE,
    TYW_INVESTMENT_PORTFOLIO,
    TYW_CBIRC_TYPE,
    TYW_PBOC_TYPE,
    TYW_CORE_CUSTOMER_TYPE,
    TYW_ACCOUNT_TYPE,
    TYW_SPECIAL_ACCOUNT_TYPE,
    TYW_CUSTODY_ACCOUNT_NAME,
}


def summarize_source_file(path: Path) -> dict[str, object]:
    metadata = describe_source_file(path.name)
    source_version = _build_source_version(path)
    family, report_date, parsed_rows, _ = _parse_source_file(
        path=path,
        ingest_batch_id="preview",
        source_version=source_version,
        source_file_name=path.name,
    )
    return _summarize_rows(
        ingest_batch_id="preview",
        batch_created_at="preview",
        family=family,
        report_date=report_date,
        report_start_date=metadata.report_start_date,
        report_end_date=metadata.report_end_date,
        report_granularity=metadata.report_granularity,
        source_file=str(path),
        source_version=source_version,
        rows=parsed_rows,
    )


def materialize_source_previews(
    duckdb_path: str,
    governance_dir: str | None = None,
    data_root: str | None = None,
    ingest_batch_id: str | None = None,
    source_families: list[str] | None = None,
) -> list[dict[str, object]]:
    manifest_rows = _load_manifest_rows(governance_dir) if governance_dir is not None else []
    selected = _select_manifest_rows(
        manifest_rows,
        ingest_batch_id=ingest_batch_id,
        source_families=source_families,
    )
    summaries: list[dict[str, object]] = []
    row_records: list[dict[str, object]] = []
    trace_records: list[dict[str, object]] = []
    _source_preview_batch_version_cached.cache_clear()

    for manifest_row in selected:
        path = Path(str(manifest_row["archived_path"]))
        metadata = describe_source_file(str(manifest_row.get("source_file") or path.name))
        family, report_date, parsed_rows, parsed_traces = _parse_source_file(
            path=path,
            ingest_batch_id=str(manifest_row["ingest_batch_id"]),
            source_version=str(manifest_row["source_version"]),
            source_file_name=str(manifest_row.get("source_file") or path.name),
        )
        summaries.append(
            _summarize_rows(
                ingest_batch_id=str(manifest_row["ingest_batch_id"]),
                batch_created_at=str(manifest_row.get("created_at", "")),
                family=family,
                report_date=report_date,
                report_start_date=str(manifest_row.get("report_start_date") or metadata.report_start_date or ""),
                report_end_date=str(manifest_row.get("report_end_date") or metadata.report_end_date or ""),
                report_granularity=str(manifest_row.get("report_granularity") or metadata.report_granularity or ""),
                source_file=str(manifest_row["source_file"]),
                source_version=str(manifest_row["source_version"]),
                rows=parsed_rows,
            )
        )
        row_records.extend(parsed_rows)
        trace_records.extend(parsed_traces)

    _write_preview_tables(duckdb_path, summaries, row_records, trace_records)
    return summaries


def snapshot_preview_tables(duckdb_path: str) -> None:
    _source_preview_batch_version_cached.cache_clear()
    conn = duckdb.connect(duckdb_path, read_only=False)
    try:
        for table_name in PREVIEW_TABLES:
            backup_name = f"{table_name}__backup"
            if _table_exists(conn, table_name):
                conn.execute(f"drop table if exists {backup_name}")
                conn.execute(f"create table {backup_name} as select * from {table_name}")
            else:
                conn.execute(f"drop table if exists {backup_name}")
    finally:
        conn.close()


def restore_preview_tables(duckdb_path: str) -> None:
    _source_preview_batch_version_cached.cache_clear()
    conn = duckdb.connect(duckdb_path, read_only=False)
    try:
        for table_name in PREVIEW_TABLES:
            backup_name = f"{table_name}__backup"
            if _table_exists(conn, backup_name):
                conn.execute(f"drop table if exists {table_name}")
                conn.execute(f"create table {table_name} as select * from {backup_name}")
            else:
                conn.execute(f"drop table if exists {table_name}")
    finally:
        conn.close()


def cleanup_preview_backups(duckdb_path: str) -> None:
    _source_preview_batch_version_cached.cache_clear()
    conn = duckdb.connect(duckdb_path, read_only=False)
    try:
        for table_name in PREVIEW_TABLES:
            conn.execute(f"drop table if exists {table_name}__backup")
    finally:
        conn.close()


def clear_preview_tables(duckdb_path: str) -> None:
    _source_preview_batch_version_cached.cache_clear()
    conn = duckdb.connect(duckdb_path, read_only=False)
    try:
        for table_name in PREVIEW_TABLES:
            conn.execute(f"drop table if exists {table_name}")
    finally:
        conn.close()


def load_source_preview_payload(duckdb_path: str) -> SourcePreviewPayload:
    duckdb_file = Path(duckdb_path)
    if not duckdb_file.exists():
        return SourcePreviewPayload(sources=[])

    conn = duckdb.connect(str(duckdb_file), read_only=True)
    try:
        summary_rows = conn.execute(
            """
            with ranked as (
              select ingest_batch_id, batch_created_at, source_family, report_date, report_start_date, report_end_date,
                   report_granularity, source_file, total_rows,
                   manual_review_count, source_version, rule_version, preview_mode,
                   row_number() over (
                     partition by source_family
                     order by batch_created_at desc, ingest_batch_id desc
                   ) as rn
              from phase1_source_preview_summary
            )
            select ingest_batch_id, batch_created_at, source_family, report_date, report_start_date, report_end_date,
                   report_granularity, source_file, total_rows,
                   manual_review_count, source_version, rule_version, preview_mode
            from ranked
            where rn = 1
            order by source_family, report_date, source_file
            """
        ).fetchall()
        group_rows = conn.execute(
            """
            select ingest_batch_id, source_family, group_label, row_count
            from phase1_source_preview_groups
            order by ingest_batch_id, source_family, group_label
            """
        ).fetchall()
    except duckdb.Error:
        return SourcePreviewPayload(sources=[])
    finally:
        conn.close()

    grouped_counts: dict[tuple[str, str], dict[str, int]] = {}
    for ingest_batch_id, source_family, group_label, row_count in group_rows:
        grouped_counts.setdefault((str(ingest_batch_id), str(source_family)), {})[str(group_label)] = int(row_count)

    return SourcePreviewPayload(
        sources=[
            SourcePreviewSummary(
                ingest_batch_id=str(ingest_batch_id) if ingest_batch_id else None,
                batch_created_at=str(batch_created_at) if batch_created_at else None,
                source_family=str(source_family),
                report_date=str(report_date) if report_date else None,
                report_start_date=str(report_start_date) if report_start_date else None,
                report_end_date=str(report_end_date) if report_end_date else None,
                report_granularity=str(report_granularity) if report_granularity else None,
                source_file=str(source_file),
                total_rows=int(total_rows),
                manual_review_count=int(manual_review_count),
                source_version=str(source_version),
                rule_version=str(rule_version),
                group_counts=grouped_counts.get((str(ingest_batch_id), str(source_family)), {}),
                preview_mode=str(preview_mode),
            )
            for (
                ingest_batch_id,
                batch_created_at,
                source_family,
                report_date,
                report_start_date,
                report_end_date,
                report_granularity,
                source_file,
                total_rows,
                manual_review_count,
                source_version,
                rule_version,
                preview_mode,
            ) in summary_rows
        ]
    )


def load_source_preview_history_payload(
    duckdb_path: str,
    limit: int,
    offset: int,
    source_family: str | None = None,
) -> SourcePreviewHistoryPage:
    duckdb_file = Path(duckdb_path)
    if not duckdb_file.exists():
        return SourcePreviewHistoryPage(limit=limit, offset=offset, total_rows=0, rows=[])

    where_clause, params = _history_query_parts(source_family)
    conn = duckdb.connect(str(duckdb_file), read_only=True)
    try:
        total_rows = conn.execute(
            f"select count(*) from phase1_source_preview_summary {where_clause}",
            params,
        ).fetchone()[0]
        rows = conn.execute(
            f"""
            select ingest_batch_id, batch_created_at, source_family, report_date, report_start_date, report_end_date,
                   report_granularity, source_file, total_rows, manual_review_count,
                   source_version, rule_version, preview_mode
            from phase1_source_preview_summary
            {where_clause}
            order by batch_created_at desc, ingest_batch_id desc, source_family asc
            limit ? offset ?
            """,
            [*params, limit, offset],
        ).fetchall()
    except duckdb.Error:
        return SourcePreviewHistoryPage(limit=limit, offset=offset, total_rows=0, rows=[])
    finally:
        conn.close()

    return SourcePreviewHistoryPage(
        limit=limit,
        offset=offset,
        total_rows=int(total_rows),
        rows=[
            SourcePreviewSummary(
                ingest_batch_id=str(ingest_batch_id) if ingest_batch_id else None,
                batch_created_at=str(batch_created_at) if batch_created_at else None,
                source_family=str(source_family),
                report_date=str(report_date) if report_date else None,
                report_start_date=str(report_start_date) if report_start_date else None,
                report_end_date=str(report_end_date) if report_end_date else None,
                report_granularity=str(report_granularity) if report_granularity else None,
                source_file=str(source_file),
                total_rows=int(total_rows_value),
                manual_review_count=int(manual_review_count),
                source_version=str(source_version),
                rule_version=str(rule_version),
                group_counts={},
                preview_mode=str(preview_mode),
            )
            for (
                ingest_batch_id,
                batch_created_at,
                source_family,
                report_date,
                report_start_date,
                report_end_date,
                report_granularity,
                source_file,
                total_rows_value,
                manual_review_count,
                source_version,
                rule_version,
                preview_mode,
            ) in rows
        ],
    )


def load_preview_rows(
    duckdb_path: str,
    source_family: str,
    limit: int,
    offset: int,
    ingest_batch_id: str | None = None,
) -> PreviewRowPage:
    duckdb_file = Path(str(duckdb_path))
    if not duckdb_file.exists():
        return PreviewRowPage(
            source_family=source_family,
            ingest_batch_id=ingest_batch_id,
            limit=limit,
            offset=offset,
            total_rows=0,
            columns=[],
            rows=[],
        )

    table_name = _row_table_name(source_family)
    resolved_batch_id, where_clause, params = _preview_read_scope(
        duckdb_path=str(duckdb_file),
        source_family=source_family,
        ingest_batch_id=ingest_batch_id,
    )

    result = _read_paged_table(
        duckdb_file=duckdb_file,
        table_name=table_name,
        where_clause=where_clause,
        params=params,
        select_clause="select *",
        order_clause="order by ingest_batch_id, row_locator",
        limit=limit,
        offset=offset,
    )
    if result is None:
        return PreviewRowPage(
            source_family=source_family,
            ingest_batch_id=ingest_batch_id,
            limit=limit,
            offset=offset,
            total_rows=0,
            columns=[],
            rows=[],
        )

    total_rows, rows, columns = result

    return PreviewRowPage(
        source_family=source_family,
        ingest_batch_id=resolved_batch_id,
        limit=limit,
        offset=offset,
        total_rows=int(total_rows),
        columns=_build_preview_columns(source_family, columns),
        rows=[dict(zip(columns, row, strict=False)) for row in rows],
    )


def load_rule_traces(
    duckdb_path: str,
    source_family: str,
    limit: int,
    offset: int,
    ingest_batch_id: str | None = None,
) -> RuleTracePage:
    duckdb_file = Path(str(duckdb_path))
    if not duckdb_file.exists():
        return RuleTracePage(
            source_family=source_family,
            ingest_batch_id=ingest_batch_id,
            limit=limit,
            offset=offset,
            total_rows=0,
            columns=[],
            rows=[],
        )

    table_name = _trace_table_name(source_family)
    resolved_batch_id, where_clause, params = _preview_read_scope(
        duckdb_path=str(duckdb_file),
        source_family=source_family,
        ingest_batch_id=ingest_batch_id,
    )

    result = _read_paged_table(
        duckdb_file=duckdb_file,
        table_name=table_name,
        where_clause=where_clause,
        params=params,
        select_clause=(
            "select ingest_batch_id, row_locator, trace_step, field_name, "
            "field_value, derived_label, manual_review_needed"
        ),
        order_clause="order by ingest_batch_id, row_locator, trace_step",
        limit=limit,
        offset=offset,
    )
    if result is None:
        return RuleTracePage(
            source_family=source_family,
            ingest_batch_id=ingest_batch_id,
            limit=limit,
            offset=offset,
            total_rows=0,
            columns=[],
            rows=[],
        )

    total_rows, rows, columns = result

    return RuleTracePage(
        source_family=source_family,
        ingest_batch_id=resolved_batch_id,
        limit=limit,
        offset=offset,
        total_rows=int(total_rows),
        columns=_build_trace_columns(columns),
        rows=[
            {
                "ingest_batch_id": str(batch_id),
                "row_locator": int(row_locator),
                "trace_step": int(trace_step),
                "field_name": str(field_name),
                "field_value": str(field_value),
                "derived_label": str(derived_label),
                "manual_review_needed": bool(manual_review_needed),
            }
            for batch_id, row_locator, trace_step, field_name, field_value, derived_label, manual_review_needed in rows
        ],
    )
def _load_manifest_rows(governance_dir: str) -> list[dict[str, object]]:
    return GovernanceRepository(base_dir=governance_dir).read_all(SOURCE_MANIFEST_STREAM)


def _history_query_parts(source_family: str | None) -> tuple[str, list[object]]:
    if source_family is None:
        return "", []
    return "where source_family = ?", [source_family]


def _preview_read_scope(
    duckdb_path: str,
    source_family: str,
    ingest_batch_id: str | None,
) -> tuple[str | None, str, list[object]]:
    resolved_batch_id = ingest_batch_id
    filters: list[str] = []
    params: list[object] = []

    if resolved_batch_id is None:
        resolved_batch_id = _latest_batch_id_for_family(duckdb_path, source_family)
    if resolved_batch_id is not None:
        filters.append("ingest_batch_id = ?")
        params.append(resolved_batch_id)
    if source_family in {"pnl_514", "pnl_516", "pnl_517"}:
        filters.append("source_family = ?")
        params.append(source_family)

    where_clause = f"where {' and '.join(filters)}" if filters else ""
    return resolved_batch_id, where_clause, params


def _read_paged_table(
    duckdb_file: Path,
    table_name: str,
    where_clause: str,
    params: list[object],
    select_clause: str,
    order_clause: str,
    limit: int,
    offset: int,
) -> tuple[int, list[tuple[object, ...]], list[str]] | None:
    conn = duckdb.connect(str(duckdb_file), read_only=True)
    try:
        total_rows = conn.execute(
            f"select count(*) from {table_name} {where_clause}",
            params,
        ).fetchone()[0]
        rows = conn.execute(
            f"""
            {select_clause}
            from {table_name}
            {where_clause}
            {order_clause}
            limit ? offset ?
            """,
            [*params, limit, offset],
        ).fetchall()
        columns = [item[0] for item in conn.description]
    except duckdb.Error:
        return None
    finally:
        conn.close()

    return int(total_rows), rows, columns


ROW_LABELS_BY_FAMILY: dict[str, dict[str, str]] = {
    "zqtz": {
        "ingest_batch_id": "批次ID",
        "row_locator": "行号",
        "report_date": "报告日期",
        "business_type_primary": "业务种类1",
        "business_type_final": "业务种类2归类",
        "asset_group": "资产分组",
        "instrument_code": "债券代码",
        "instrument_name": "债券名称",
        "account_category": "账户类别",
        "manual_review_needed": "需人工复核",
        "source_version": "数据版本",
        "rule_version": "规则版本",
    },
    "tyw": {
        "ingest_batch_id": "批次ID",
        "row_locator": "行号",
        "report_date": "报告日期",
        "business_type_primary": "业务种类1",
        "product_group": "产品分组",
        "institution_category": "机构类型",
        "special_nature": "特殊性质",
        "counterparty_name": "对手方名称",
        "investment_portfolio": "投资组合",
        "manual_review_needed": "需人工复核",
        "source_version": "数据版本",
        "rule_version": "规则版本",
    },
    "pnl": {
        "source_family": "源类型",
        "ingest_batch_id": "批次ID",
        "row_locator": "行号",
        "report_date": "报告日期",
        "instrument_code": "债券代码",
        "invest_type_raw": "投资类型原值",
        "portfolio_name": "投资组合",
        "cost_center": "成本中心",
        "currency": "币种",
        "manual_review_needed": "需人工复核",
        "source_version": "数据版本",
        "rule_version": "规则版本",
    },
    "pnl_514": {
        "source_family": "源类型",
        "ingest_batch_id": "批次ID",
        "row_locator": "行号",
        "report_date": "报告日期",
        "journal_type": "分录类型",
        "product_type": "产品类型",
        "asset_code": "资产代码",
        "account_code": "科目号",
        "dc_flag_raw": "借贷标识",
        "raw_amount": "原始金额",
        "manual_review_needed": "需人工复核",
        "source_version": "数据版本",
        "rule_version": "规则版本",
    },
    "pnl_516": {
        "source_family": "源类型",
        "ingest_batch_id": "批次ID",
        "row_locator": "行号",
        "report_date": "报告日期",
        "journal_type": "分录类型",
        "product_type": "产品类型",
        "asset_code": "资产代码",
        "account_code": "科目号",
        "dc_flag_raw": "借贷标识",
        "raw_amount": "原始金额",
        "manual_review_needed": "需人工复核",
        "source_version": "数据版本",
        "rule_version": "规则版本",
    },
    "pnl_517": {
        "source_family": "源类型",
        "ingest_batch_id": "批次ID",
        "row_locator": "行号",
        "report_date": "报告日期",
        "journal_type": "分录类型",
        "product_type": "产品类型",
        "asset_code": "资产代码",
        "account_code": "科目号",
        "dc_flag_raw": "借贷标识",
        "raw_amount": "原始金额",
        "manual_review_needed": "需人工复核",
        "source_version": "数据版本",
        "rule_version": "规则版本",
    },
}

TRACE_LABELS: dict[str, str] = {
    "ingest_batch_id": "批次ID",
    "row_locator": "行号",
    "trace_step": "轨迹步骤",
    "field_name": "字段名",
    "field_value": "字段值",
    "derived_label": "归类标签",
    "manual_review_needed": "需人工复核",
}


def _build_preview_columns(source_family: str, columns: list[str]) -> list[PreviewColumn]:
    labels = ROW_LABELS_BY_FAMILY.get(source_family, {})
    return [
        PreviewColumn(
            key=column,
            label=labels.get(column, column),
            type=_preview_column_type(column),
        )
        for column in columns
    ]


def _build_trace_columns(columns: list[str]) -> list[PreviewColumn]:
    return [
        PreviewColumn(
            key=column,
            label=TRACE_LABELS.get(column, column),
            type=_preview_column_type(column),
        )
        for column in columns
    ]


def _preview_column_type(column: str) -> str:
    if column in {"row_locator", "trace_step"}:
        return "number"
    if column in {"manual_review_needed"}:
        return "boolean"
    return "string"


def _select_manifest_rows(
    manifest_rows: list[dict[str, object]],
    ingest_batch_id: str | None = None,
    source_families: list[str] | None = None,
) -> list[dict[str, object]]:
    eligible_rows = [
        row
        for row in manifest_rows
        if str(row.get("status", "")) in MANIFEST_ELIGIBLE_STATUSES
        and row.get("archived_path")
        and Path(str(row["archived_path"])).exists()
    ]
    if source_families is not None:
        allowed = {str(family) for family in source_families}
        eligible_rows = [
            row
            for row in eligible_rows
            if str(row.get("source_family", "")) in allowed
        ]
    if ingest_batch_id is not None:
        return [
            row
            for row in eligible_rows
            if str(row.get("ingest_batch_id", "")) == ingest_batch_id
        ]

    latest_rows: list[dict[str, object]] = []
    families = sorted({str(row.get("source_family", "")) for row in eligible_rows if row.get("source_family")})
    for family in families:
        family_rows = [row for row in eligible_rows if str(row.get("source_family", "")) == family]
        latest_report_date = max(str(row.get("report_date", "")) for row in family_rows)
        bounded_rows = [row for row in family_rows if str(row.get("report_date", "")) == latest_report_date]
        latest_batch_id = max(
            bounded_rows,
            key=lambda item: (
                str(item.get("created_at", "")),
                str(item.get("ingest_batch_id", "")),
            ),
        )["ingest_batch_id"]
        latest_rows.extend(
            sorted(
                [row for row in bounded_rows if str(row.get("ingest_batch_id", "")) == str(latest_batch_id)],
                key=lambda item: str(item.get("archived_path", "")),
            )
        )
    return latest_rows


def _parse_source_file(
    path: Path,
    ingest_batch_id: str,
    source_version: str,
    source_file_name: str | None = None,
) -> tuple[str, str | None, list[dict[str, object]], list[dict[str, object]]]:
    metadata = describe_source_file(source_file_name or path.name)
    source_family = metadata.source_family
    if source_family == "pnl":
        return _parse_pnl_source_file(
            path=path,
            ingest_batch_id=ingest_batch_id,
            source_version=source_version,
            metadata=metadata,
        )

    if source_family in {"pnl_514", "pnl_516", "pnl_517"}:
        return _parse_nonstd_pnl_source_file(
            path=path,
            ingest_batch_id=ingest_batch_id,
            source_version=source_version,
            metadata=metadata,
        )

    if source_family not in {"zqtz", "tyw"}:
        return source_family, metadata.report_date, [], []

    sheet = xlrd.open_workbook(str(path)).sheet_by_index(0)
    headers = [str(sheet.cell_value(1, column)).strip() for column in range(sheet.ncols)]
    rows: list[dict[str, object]] = []
    traces: list[dict[str, object]] = []
    row_locator = 0

    for row_index in range(2, sheet.nrows):
        raw_row = {
            headers[column]: sheet.cell_value(row_index, column)
            for column in range(sheet.ncols)
            if headers[column]
        }
        row_locator += 1
        if source_family == "zqtz":
            preview = classify_zqtz_preview(raw_row)
            row_record = ZqtzPreviewRow(
                ingest_batch_id=ingest_batch_id,
                row_locator=row_locator,
                report_date=metadata.report_date or _text(raw_row, ZQTZ_DATE) or None,
                business_type_primary=str(preview["business_type_primary"]),
                business_type_final=str(preview["business_type_final"]),
                asset_group=str(preview["asset_group"]),
                instrument_code=_text(raw_row, ZQTZ_BOND_CODE),
                instrument_name=_text(raw_row, ZQTZ_BOND_NAME),
                account_category=_text(raw_row, ZQTZ_ACCOUNT_CATEGORY),
                manual_review_needed=bool(preview["manual_review_needed"]),
            ).model_dump(mode="json")
            trace_rows = _zqtz_trace_rows(raw_row, row_record)
        else:
            preview = classify_tyw_preview(raw_row)
            row_record = TywPreviewRow(
                ingest_batch_id=ingest_batch_id,
                row_locator=row_locator,
                report_date=metadata.report_date,
                business_type_primary=str(preview["business_type_primary"]),
                product_group=str(preview["product_group"]),
                institution_category=str(preview["institution_category"]),
                special_nature=str(preview["special_nature"]),
                counterparty_name=_text(raw_row, TYW_COUNTERPARTY_NAME),
                investment_portfolio=str(preview["investment_portfolio"]),
                manual_review_needed=bool(preview["manual_review_needed"]),
            ).model_dump(mode="json")
            trace_rows = _tyw_trace_rows(raw_row, row_record)

        row_record["source_version"] = source_version
        row_record["rule_version"] = RULE_VERSION
        rows.append(row_record)
        traces.extend(trace_rows)

    return source_family, metadata.report_date, rows, traces


def _zqtz_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    rows = [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": int(row_record["row_locator"]),
            "trace_step": 1,
            "field_name": ZQTZ_BUSINESS_TYPE1,
            "field_value": _text(raw_row, ZQTZ_BUSINESS_TYPE1),
            "derived_label": str(row_record["business_type_primary"]),
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "zqtz",
        }
    ]
    if str(row_record["business_type_primary"]) == "\u5176\u4ed6\u503a\u5238":
        rows.append(
            {
                "ingest_batch_id": str(row_record["ingest_batch_id"]),
                "row_locator": int(row_record["row_locator"]),
                "trace_step": 2,
                "field_name": ZQTZ_BOND_CODE,
                "field_value": _text(raw_row, ZQTZ_BOND_CODE),
                "derived_label": str(row_record["business_type_final"]),
                "manual_review_needed": bool(row_record["manual_review_needed"]),
                "source_family": "zqtz",
            }
        )
    rows.append(
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": int(row_record["row_locator"]),
            "trace_step": 3,
            "field_name": "asset_group_map",
            "field_value": str(row_record["business_type_final"]),
            "derived_label": str(row_record["asset_group"]),
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "zqtz",
        }
    )
    return rows


def _tyw_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    trace_spec = [
        (1, TYW_PRODUCT_TYPE, str(row_record["business_type_primary"])),
        (2, TYW_INVESTMENT_PORTFOLIO, str(row_record["product_group"])),
        (3, TYW_CBIRC_TYPE, str(row_record["institution_category"])),
        (4, TYW_PBOC_TYPE, str(row_record["institution_category"])),
        (5, TYW_CORE_CUSTOMER_TYPE, str(row_record["institution_category"])),
        (6, TYW_ACCOUNT_TYPE, str(row_record["special_nature"])),
        (7, TYW_SPECIAL_ACCOUNT_TYPE, str(row_record["special_nature"])),
        (8, TYW_CUSTODY_ACCOUNT_NAME, str(row_record["special_nature"])),
    ]
    return [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": int(row_record["row_locator"]),
            "trace_step": trace_step,
            "field_name": field_name,
            "field_value": _text(raw_row, field_name),
            "derived_label": derived_label,
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "tyw",
        }
        for trace_step, field_name, derived_label in trace_spec
    ]


def _parse_pnl_source_file(
    path: Path,
    ingest_batch_id: str,
    source_version: str,
    metadata,
) -> tuple[str, str | None, list[dict[str, object]], list[dict[str, object]]]:
    sheet = xlrd.open_workbook(str(path)).sheet_by_index(0)
    headers = [str(sheet.cell_value(0, column)).strip() for column in range(sheet.ncols)]
    rows: list[dict[str, object]] = []
    traces: list[dict[str, object]] = []
    row_locator = 0

    for row_index in range(1, sheet.nrows):
        raw_row = {
            headers[column]: sheet.cell_value(row_index, column)
            for column in range(sheet.ncols)
            if headers[column]
        }
        if not _text(raw_row, "\u503a\u5238\u4ee3\u7801"):
            continue

        row_locator += 1
        preview = classify_pnl_preview(raw_row)
        row_record = PnlPreviewRow(
            ingest_batch_id=ingest_batch_id,
            row_locator=row_locator,
            report_date=metadata.report_date,
            instrument_code=str(preview["instrument_code"]),
            invest_type_raw=str(preview["invest_type_raw"]),
            portfolio_name=str(preview["portfolio_name"]),
            cost_center=str(preview["cost_center"]),
            currency=str(preview["currency"]),
            manual_review_needed=bool(preview["manual_review_needed"]),
        ).model_dump(mode="json")
        row_record["source_family"] = "pnl"
        row_record["source_version"] = source_version
        row_record["rule_version"] = RULE_VERSION
        rows.append(row_record)
        traces.extend(_pnl_trace_rows(raw_row, row_record))

    return "pnl", metadata.report_date, rows, traces


def _parse_nonstd_pnl_source_file(
    path: Path,
    ingest_batch_id: str,
    source_version: str,
    metadata,
) -> tuple[str, str | None, list[dict[str, object]], list[dict[str, object]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    headers = [
        "" if value is None else str(value).strip()
        for value in next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True))
    ]
    rows: list[dict[str, object]] = []
    traces: list[dict[str, object]] = []
    row_locator = 0
    bucket = metadata.source_family.removeprefix("pnl_")

    for values in worksheet.iter_rows(min_row=3, values_only=True):
        raw_row = {
            headers[index]: values[index]
            for index in range(min(len(headers), len(values)))
            if headers[index]
        }
        if not _text(raw_row, "\u8d44\u4ea7\u4ee3\u7801"):
            continue

        row_locator += 1
        preview = classify_nonstd_pnl_preview(raw_row, bucket=bucket)
        row_record = NonstdPnlPreviewRow(
            ingest_batch_id=ingest_batch_id,
            row_locator=row_locator,
            report_date=metadata.report_date,
            journal_type=str(preview["journal_type"]),
            product_type=str(preview["product_type"]),
            asset_code=str(preview["asset_code"]),
            account_code=str(preview["account_code"]),
            dc_flag_raw=str(preview["dc_flag_raw"]),
            raw_amount=str(preview["raw_amount"]),
            manual_review_needed=bool(preview["manual_review_needed"]),
        ).model_dump(mode="json")
        row_record["source_family"] = metadata.source_family
        row_record["source_version"] = source_version
        row_record["rule_version"] = RULE_VERSION
        rows.append(row_record)
        traces.extend(_nonstd_pnl_trace_rows(raw_row, row_record))

    return metadata.source_family, metadata.report_date, rows, traces


def _pnl_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    trace_spec = [
        (1, "\u6295\u8d44\u7c7b\u578b", str(row_record["invest_type_raw"])),
        (2, "\u503a\u5238\u4ee3\u7801", str(row_record["instrument_code"])),
        (3, "\u6210\u672c\u4e2d\u5fc3", str(row_record["cost_center"])),
    ]
    return [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": int(row_record["row_locator"]),
            "trace_step": trace_step,
            "field_name": field_name,
            "field_value": _text(raw_row, field_name),
            "derived_label": derived_label,
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": "pnl",
        }
        for trace_step, field_name, derived_label in trace_spec
    ]


def _nonstd_pnl_trace_rows(raw_row: dict[str, object], row_record: dict[str, object]) -> list[dict[str, object]]:
    trace_spec = [
        (1, "科目号", str(row_record["account_code"])),
        (2, "资产代码", str(row_record["asset_code"])),
        (3, "借贷标识", str(row_record["dc_flag_raw"])),
    ]
    return [
        {
            "ingest_batch_id": str(row_record["ingest_batch_id"]),
            "row_locator": int(row_record["row_locator"]),
            "trace_step": trace_step,
            "field_name": field_name,
            "field_value": _text(raw_row, field_name),
            "derived_label": derived_label,
            "manual_review_needed": bool(row_record["manual_review_needed"]),
            "source_family": str(row_record["source_family"]),
        }
        for trace_step, field_name, derived_label in trace_spec
    ]


def _summarize_rows(
    ingest_batch_id: str,
    batch_created_at: str,
    family: str,
    report_date: str | None,
    report_start_date: str | None,
    report_end_date: str | None,
    report_granularity: str | None,
    source_file: str,
    source_version: str,
    rows: list[dict[str, object]],
) -> dict[str, object]:
    group_counts = Counter(_group_label(family, row) for row in rows)
    return {
        "ingest_batch_id": ingest_batch_id,
        "batch_created_at": batch_created_at,
        "source_family": family,
        "report_date": report_date,
        "report_start_date": report_start_date,
        "report_end_date": report_end_date,
        "report_granularity": report_granularity,
        "source_file": source_file,
        "total_rows": len(rows),
        "manual_review_count": sum(int(bool(row["manual_review_needed"])) for row in rows),
        "source_version": source_version,
        "rule_version": RULE_VERSION,
        "group_counts": dict(group_counts),
        "preview_mode": "tabular",
    }


def _group_label(source_family: str, row: dict[str, object]) -> str:
    if source_family == "zqtz":
        return str(row["asset_group"])
    if source_family == "tyw":
        return str(row["product_group"])
    if source_family == "pnl":
        return str(row["invest_type_raw"] or "未标注")
    return str(row["product_type"] or row["journal_type"] or "未标注")


def ensure_source_preview_schema_tables(conn: duckdb.DuckDBPyConnection) -> None:
    """Baseline DDL is versioned in `duckdb_migrations` (also run at API/worker startup)."""
    apply_pending_migrations_on_connection(conn)


def _write_preview_tables(
    duckdb_path: str,
    summaries: list[dict[str, object]],
    row_records: list[dict[str, object]],
    trace_records: list[dict[str, object]],
) -> None:
    conn = duckdb.connect(duckdb_path, read_only=False)
    transaction_started = False
    try:
        ensure_source_preview_schema_tables(conn)
        conn.execute("begin transaction")
        transaction_started = True

        if summaries:
            current_batch_ids = sorted({summary["ingest_batch_id"] for summary in summaries})
            for ingest_batch_id in current_batch_ids:
                conn.execute(
                    "delete from phase1_source_preview_summary where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_source_preview_groups where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_zqtz_preview_rows where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_tyw_preview_rows where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_pnl_preview_rows where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_nonstd_pnl_preview_rows where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_zqtz_rule_traces where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_tyw_rule_traces where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_pnl_rule_traces where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
                conn.execute(
                    "delete from phase1_nonstd_pnl_rule_traces where ingest_batch_id = ?",
                    [ingest_batch_id],
                )
            conn.executemany(
                "insert into phase1_source_preview_summary values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                [
                    (
                        summary["ingest_batch_id"],
                        summary["batch_created_at"],
                        summary["source_family"],
                        summary["report_date"],
                        summary["report_start_date"],
                        summary["report_end_date"],
                        summary["report_granularity"],
                        summary["source_file"],
                        summary["total_rows"],
                        summary["manual_review_count"],
                        summary["source_version"],
                        summary["rule_version"],
                        summary["preview_mode"],
                    )
                    for summary in summaries
                ],
            )
            group_rows = [
                (
                    summary["ingest_batch_id"],
                    summary["source_family"],
                    group_label,
                    row_count,
                    summary["source_version"],
                )
                for summary in summaries
                for group_label, row_count in summary["group_counts"].items()
            ]
            if group_rows:
                conn.executemany(
                    "insert into phase1_source_preview_groups values (?, ?, ?, ?, ?)",
                    group_rows,
                )
        else:
            for table_name in PREVIEW_TABLES:
                conn.execute(f"delete from {table_name}")

        zqtz_rows = [
            (
                row["ingest_batch_id"],
                row["row_locator"],
                row["report_date"],
                row["business_type_primary"],
                row["business_type_final"],
                row["asset_group"],
                row["instrument_code"],
                row["instrument_name"],
                row["account_category"],
                row["manual_review_needed"],
                row["source_version"],
                row["rule_version"],
            )
            for row in row_records
            if "asset_group" in row
        ]
        if zqtz_rows:
            conn.executemany(
                "insert into phase1_zqtz_preview_rows values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                zqtz_rows,
            )

        tyw_rows = [
            (
                row["ingest_batch_id"],
                row["row_locator"],
                row["report_date"],
                row["business_type_primary"],
                row["product_group"],
                row["institution_category"],
                row["special_nature"],
                row["counterparty_name"],
                row["investment_portfolio"],
                row["manual_review_needed"],
                row["source_version"],
                row["rule_version"],
            )
            for row in row_records
            if "product_group" in row
        ]
        if tyw_rows:
            conn.executemany(
                "insert into phase1_tyw_preview_rows values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                tyw_rows,
            )

        pnl_rows = [
            (
                row["source_family"],
                row["ingest_batch_id"],
                row["row_locator"],
                row["report_date"],
                row["instrument_code"],
                row["invest_type_raw"],
                row["portfolio_name"],
                row["cost_center"],
                row["currency"],
                row["manual_review_needed"],
                row["source_version"],
                row["rule_version"],
            )
            for row in row_records
            if "invest_type_raw" in row
        ]
        if pnl_rows:
            conn.executemany(
                "insert into phase1_pnl_preview_rows values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                pnl_rows,
            )

        nonstd_pnl_rows = [
            (
                row["source_family"],
                row["ingest_batch_id"],
                row["row_locator"],
                row["report_date"],
                row["journal_type"],
                row["product_type"],
                row["asset_code"],
                row["account_code"],
                row["dc_flag_raw"],
                row["raw_amount"],
                row["manual_review_needed"],
                row["source_version"],
                row["rule_version"],
            )
            for row in row_records
            if "journal_type" in row
        ]
        if nonstd_pnl_rows:
            conn.executemany(
                "insert into phase1_nonstd_pnl_preview_rows values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                nonstd_pnl_rows,
            )

        zqtz_traces = [
            (
                trace["ingest_batch_id"],
                trace["row_locator"],
                trace["trace_step"],
                trace["field_name"],
                trace["field_value"],
                trace["derived_label"],
                trace["manual_review_needed"],
            )
            for trace in trace_records
            if trace.get("source_family") == "zqtz"
        ]
        if zqtz_traces:
            conn.executemany(
                "insert into phase1_zqtz_rule_traces values (?, ?, ?, ?, ?, ?, ?)",
                zqtz_traces,
            )

        tyw_traces = [
            (
                trace["ingest_batch_id"],
                trace["row_locator"],
                trace["trace_step"],
                trace["field_name"],
                trace["field_value"],
                trace["derived_label"],
                trace["manual_review_needed"],
            )
            for trace in trace_records
            if trace.get("source_family") == "tyw"
        ]
        if tyw_traces:
            conn.executemany(
                "insert into phase1_tyw_rule_traces values (?, ?, ?, ?, ?, ?, ?)",
                tyw_traces,
            )

        pnl_traces = [
            (
                trace["source_family"],
                trace["ingest_batch_id"],
                trace["row_locator"],
                trace["trace_step"],
                trace["field_name"],
                trace["field_value"],
                trace["derived_label"],
                trace["manual_review_needed"],
            )
            for trace in trace_records
            if trace.get("source_family") == "pnl"
        ]
        if pnl_traces:
            conn.executemany(
                "insert into phase1_pnl_rule_traces values (?, ?, ?, ?, ?, ?, ?, ?)",
                pnl_traces,
            )

        nonstd_pnl_traces = [
            (
                trace["source_family"],
                trace["ingest_batch_id"],
                trace["row_locator"],
                trace["trace_step"],
                trace["field_name"],
                trace["field_value"],
                trace["derived_label"],
                trace["manual_review_needed"],
            )
            for trace in trace_records
            if trace.get("source_family") in {"pnl_514", "pnl_516", "pnl_517"}
        ]
        if nonstd_pnl_traces:
            conn.executemany(
                "insert into phase1_nonstd_pnl_rule_traces values (?, ?, ?, ?, ?, ?, ?, ?)",
                nonstd_pnl_traces,
            )
        conn.execute("commit")
        transaction_started = False
    except Exception:
        if transaction_started:
            try:
                conn.execute("rollback")
            except Exception:
                pass
        raise
    finally:
        conn.close()


def _table_exists(conn: duckdb.DuckDBPyConnection, table_name: str) -> bool:
    return bool(
        conn.execute(
            """
            select count(*)
            from information_schema.tables
            where table_name = ?
            """,
            [table_name],
        ).fetchone()[0]
    )


def _row_table_name(source_family: str) -> str:
    if source_family == "zqtz":
        return "phase1_zqtz_preview_rows"
    if source_family == "tyw":
        return "phase1_tyw_preview_rows"
    if source_family == "pnl":
        return "phase1_pnl_preview_rows"
    if source_family in {"pnl_514", "pnl_516", "pnl_517"}:
        return "phase1_nonstd_pnl_preview_rows"
    raise ValueError(f"Unsupported source family: {source_family}")


def _trace_table_name(source_family: str) -> str:
    if source_family == "zqtz":
        return "phase1_zqtz_rule_traces"
    if source_family == "tyw":
        return "phase1_tyw_rule_traces"
    if source_family == "pnl":
        return "phase1_pnl_rule_traces"
    if source_family in {"pnl_514", "pnl_516", "pnl_517"}:
        return "phase1_nonstd_pnl_rule_traces"
    raise ValueError(f"Unsupported source family: {source_family}")


def _text(row: dict[str, object], key: str) -> str:
    value = row.get(key, "")
    if value is None:
        return ""
    return str(value).strip()


def _build_source_version(path: Path) -> str:
    stat = path.stat()
    seed = f"{path.name}:{stat.st_size}:{stat.st_mtime_ns}"
    return f"sv_{hashlib.sha256(seed.encode('utf-8')).hexdigest()[:12]}"


def source_preview_payload_version(payload: SourcePreviewPayload) -> str:
    return _join_source_versions(source.source_version for source in payload.sources)


def source_preview_history_version(payload: SourcePreviewHistoryPage) -> str:
    return _join_source_versions(row.source_version for row in payload.rows)


def _join_source_versions(versions) -> str:
    ordered: list[str] = []
    seen: set[str] = set()
    for version in versions:
        normalized = str(version or "").strip()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        ordered.append(normalized)
    if not ordered:
        return "sv_preview_empty"
    return "__".join(ordered)


def source_preview_batch_version(
    duckdb_path: str,
    source_family: str,
    ingest_batch_id: str | None,
) -> str:
    return _source_preview_batch_version_cached(
        str(duckdb_path),
        str(source_family),
        None if ingest_batch_id is None else str(ingest_batch_id),
    )


@lru_cache(maxsize=1024)
def _source_preview_batch_version_cached(
    duckdb_path: str,
    source_family: str,
    ingest_batch_id: str | None,
) -> str:
    duckdb_file = Path(str(duckdb_path))
    if not duckdb_file.exists():
        return "sv_preview_empty"

    filters = ["source_family = ?"]
    params: list[object] = [source_family]
    if ingest_batch_id is not None:
        filters.append("ingest_batch_id = ?")
        params.append(ingest_batch_id)
    where_clause = f"where {' and '.join(filters)}"

    conn = duckdb.connect(str(duckdb_file), read_only=True)
    try:
        row = conn.execute(
            f"""
            select source_version
            from phase1_source_preview_summary
            {where_clause}
            order by batch_created_at desc, ingest_batch_id desc
            limit 1
            """,
            params,
        ).fetchone()
    except duckdb.Error:
        return "sv_preview_empty"
    finally:
        conn.close()

    if row is None:
        return "sv_preview_empty"
    return str(row[0] or "sv_preview_empty")


def _latest_batch_id_for_family(duckdb_path: str, source_family: str) -> str | None:
    conn = duckdb.connect(duckdb_path, read_only=True)
    try:
        row = conn.execute(
            """
            select ingest_batch_id
            from phase1_source_preview_summary
            where source_family = ?
            order by batch_created_at desc, ingest_batch_id desc
            limit 1
            """,
            [source_family],
        ).fetchone()
    except duckdb.Error:
        return None
    finally:
        conn.close()
    return str(row[0]) if row and row[0] else None


def _direct_source_rows(data_root: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for path in sorted(item for item in data_root.rglob("*") if item.is_file()):
        metadata = describe_source_file(path.name)
        if metadata.source_family not in {"zqtz", "tyw", "pnl", "pnl_514", "pnl_516", "pnl_517"}:
            continue
        rows.append(
            {
                "source_family": metadata.source_family,
                "report_date": metadata.report_date,
                "report_start_date": metadata.report_start_date,
                "report_end_date": metadata.report_end_date,
                "report_granularity": metadata.report_granularity,
                "source_file": path.name,
                "archived_path": str(path),
                "ingest_batch_id": "preview-direct",
                "source_version": _build_source_version(path),
                "status": "completed",
                "created_at": path.stat().st_mtime_ns,
            }
        )
    return rows
