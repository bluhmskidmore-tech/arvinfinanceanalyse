from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from tests.helpers import ROOT, load_module


LEDGER_HEADERS = [
    "组合科目代码",
    "组合科目名称",
    "币种",
    "期初余额",
    "本期借方",
    "本期贷方",
    "期末余额",
]

AVERAGE_BLOCK_HEADER = ["币种", "科目", "科目日均余额", None]


def test_discover_qdb_gl_baseline_bindings_groups_ledger_and_average_sources(tmp_path):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    _write_qdb_gl_ledger_workbook(tmp_path / "总账对账202602.xlsx")
    _write_qdb_gl_average_workbook(tmp_path / "日均202602.xlsx")
    (tmp_path / "README.txt").write_text("ignore me", encoding="utf-8")

    bindings = module.discover_qdb_gl_baseline_bindings(tmp_path)

    assert [(binding.source_kind, binding.report_month, binding.path.name) for binding in bindings] == [
        ("average_balance", "202602", "日均202602.xlsx"),
        ("ledger_reconciliation", "202602", "总账对账202602.xlsx"),
    ]
    assert all(binding.source_version.startswith("sv_qdb_gl_") for binding in bindings)


def test_validate_qdb_gl_ledger_source_emits_pass_evidence_with_lineage(tmp_path):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    path = tmp_path / "总账对账202602.xlsx"
    _write_qdb_gl_ledger_workbook(path)

    evidence = module.validate_qdb_gl_baseline_source(path)
    checks = {check.check_id: check for check in evidence.checks}

    assert evidence.binding_status == "bound"
    assert evidence.source_kind == "ledger_reconciliation"
    assert evidence.report_month == "202602"
    assert evidence.admissible is True
    assert evidence.status_label == "pass"
    assert evidence.source_version.startswith("sv_qdb_gl_")
    assert evidence.rule_version == module.RULE_VERSION
    assert evidence.trace_id.startswith("tr_qdb_gl_ledger_reconciliation_")
    assert evidence.bound_currency_groups == ["CNX", "CNY"]
    assert evidence.lineage.source_file == path.name
    assert evidence.lineage.sheet_names == ["综本", "人民币"]
    assert checks["header_row"].status_label == "pass"
    assert checks["row_shape"].status_label == "pass"
    assert checks["required_raw_fields"].status_label == "pass"
    assert checks["account_code_text_preserved"].status_label == "pass"
    assert checks["currency_grouping"].status_label == "pass"
    assert checks["reconciliation_contract"].status_label == "pass"


def test_validate_qdb_gl_average_source_emits_pass_evidence_with_lineage(tmp_path):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    path = tmp_path / "日均202602.xlsx"
    _write_qdb_gl_average_workbook(path)

    evidence = module.validate_qdb_gl_baseline_source(path)
    checks = {check.check_id: check for check in evidence.checks}

    assert evidence.binding_status == "bound"
    assert evidence.source_kind == "average_balance"
    assert evidence.report_month == "202602"
    assert evidence.admissible is True
    assert evidence.status_label == "pass"
    assert evidence.bound_currency_groups == ["CNX", "CNY"]
    assert evidence.lineage.sheet_names == ["年", "月"]
    assert checks["header_row"].status_label == "pass"
    assert checks["row_shape"].status_label == "pass"
    assert checks["required_raw_fields"].status_label == "pass"
    assert checks["account_code_text_preserved"].status_label == "pass"
    assert checks["currency_grouping"].status_label == "pass"
    assert checks["reconciliation_contract"].status_label == "not_applicable"


def test_validate_qdb_gl_source_rejects_unrecognized_file_name(tmp_path):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    path = tmp_path / "未知输入202602.xlsx"
    _write_qdb_gl_ledger_workbook(path)

    evidence = module.validate_qdb_gl_baseline_source(path)
    checks = {check.check_id: check for check in evidence.checks}

    assert evidence.binding_status == "rejected"
    assert evidence.admissible is False
    assert evidence.status_label == "fail"
    assert evidence.source_kind == "unknown"
    assert checks["source_binding"].status_label == "fail"


def test_validate_qdb_gl_source_binding_fails_when_canonical_sheet_is_missing(tmp_path):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    path = tmp_path / "总账对账202602.xlsx"
    _write_qdb_gl_ledger_workbook(path)

    workbook = load_workbook(path)
    del workbook["人民币"]
    workbook.save(path)
    workbook.close()

    evidence = module.validate_qdb_gl_baseline_source(path)
    checks = {check.check_id: check for check in evidence.checks}

    assert evidence.admissible is False
    assert evidence.status_label == "fail"
    assert checks["source_binding"].status_label == "fail"
    assert any(item.sheet_name == "人民币" for item in checks["source_binding"].findings)


def test_validate_qdb_gl_ledger_source_flags_header_violations(tmp_path):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    path = tmp_path / "总账对账202602.xlsx"
    _write_qdb_gl_ledger_workbook(path)

    workbook = load_workbook(path)
    worksheet = workbook["综本"]
    worksheet["A6"] = "错误表头"
    workbook.save(path)
    workbook.close()

    evidence = module.validate_qdb_gl_baseline_source(path)
    checks = {check.check_id: check for check in evidence.checks}

    assert evidence.admissible is False
    assert evidence.status_label == "fail"
    assert checks["header_row"].status_label == "fail"
    assert any(item.sheet_name == "综本" and item.cell_ref == "A6" for item in checks["header_row"].findings)


def test_validate_qdb_gl_average_source_flags_row_shape_violations(tmp_path):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    path = tmp_path / "日均202602.xlsx"
    _write_qdb_gl_average_workbook(path)

    workbook = load_workbook(path)
    worksheet = workbook["年"]
    worksheet["D4"] = "unexpected"
    workbook.save(path)
    workbook.close()

    evidence = module.validate_qdb_gl_baseline_source(path)
    checks = {check.check_id: check for check in evidence.checks}

    assert evidence.admissible is False
    assert evidence.status_label == "fail"
    assert checks["row_shape"].status_label == "fail"
    assert any(item.sheet_name == "年" and item.cell_ref == "D4" for item in checks["row_shape"].findings)


def test_validate_qdb_gl_ledger_source_flags_required_field_account_code_currency_and_reconciliation_failures(
    tmp_path,
):
    module = load_module(
        "backend.app.services.qdb_gl_input_validation_service",
        "backend/app/services/qdb_gl_input_validation_service.py",
    )

    path = tmp_path / "总账对账202602.xlsx"
    _write_qdb_gl_ledger_workbook(path)

    workbook = load_workbook(path)
    worksheet = workbook["综本"]
    worksheet["A7"] = "1.0101E+10"
    worksheet["B7"] = None
    worksheet["C7"] = "USD"
    worksheet["G7"] = 999
    workbook.save(path)
    workbook.close()

    evidence = module.validate_qdb_gl_baseline_source(path)
    checks = {check.check_id: check for check in evidence.checks}

    assert evidence.admissible is False
    assert evidence.status_label == "fail"
    assert checks["required_raw_fields"].status_label == "fail"
    assert checks["row_shape"].status_label == "fail"
    assert checks["account_code_text_preserved"].status_label == "fail"
    assert checks["currency_grouping"].status_label == "fail"
    assert checks["reconciliation_contract"].status_label == "fail"
    assert any(item.sheet_name == "综本" and item.row_locator == 7 for item in checks["required_raw_fields"].findings)
    assert any(item.sheet_name == "综本" and item.row_locator == 7 for item in checks["row_shape"].findings)
    assert any(item.sheet_name == "综本" and item.row_locator == 7 for item in checks["account_code_text_preserved"].findings)
    assert any(item.sheet_name == "综本" and item.row_locator == 7 for item in checks["currency_grouping"].findings)
    assert any(item.sheet_name == "综本" and item.row_locator == 7 for item in checks["reconciliation_contract"].findings)


def test_gl_rules_spec_exists_and_stays_assembly_only():
    path = ROOT / "docs" / "gl_rules_spec.md"

    assert path.exists(), "docs/gl_rules_spec.md must exist for the QDB GL owner-doc stack."

    text = path.read_text(encoding="utf-8")

    assert "assembly" in text.lower()
    assert "index" in text.lower()
    assert "must not become" in text.lower()
    assert "second normative truth source" in text.lower()
    assert "product_category_*" in text


def _write_qdb_gl_ledger_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, company_name, currency in (
        ("综本", "（199200）青岛银行", "CNX"),
        ("人民币", "（199200）青岛银行", "CNY"),
    ):
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(["总账（对账）"])
        worksheet.append([None])
        worksheet.append(["分类账： QDB_Priledger 来源： "])
        worksheet.append([f"公司： {company_name} 类别： "])
        worksheet.append([f"会计期间： 2026-02-01--2026-02-28 币种： {currency}"])
        worksheet.append([*LEDGER_HEADERS])
        worksheet.append(["00101010001", "业务库存现金", currency, 100, 20, 5, 115])
        worksheet.append([10101000002, "ATM库存现金", currency, 50, 10, 4, 56])
    workbook.save(path)


def _write_qdb_gl_average_workbook(path: Path) -> None:
    workbook = Workbook()
    year_sheet = workbook.active
    year_sheet.title = "年"
    month_sheet = workbook.create_sheet(title="月")

    for worksheet, date_line in (
        (year_sheet, "日期： 2026-01-01 至 2026-02-28"),
        (month_sheet, "日期： 2026-02-01 至 2026-02-28"),
    ):
        worksheet.append(["机构： 199200青岛银行", None, None, None, "机构： 199200青岛银行", None, None, None])
        worksheet.append([date_line, None, None, None, date_line, None, None, None])
        worksheet.append([*AVERAGE_BLOCK_HEADER, *AVERAGE_BLOCK_HEADER])
        worksheet.append(["CNX", "00101010001", 115, None, "CNY", "00101010001", 95, None])
        worksheet.append(["CNX", 10101000002, 56, None, "CNY", 10101000002, 46, None])
    workbook.save(path)
