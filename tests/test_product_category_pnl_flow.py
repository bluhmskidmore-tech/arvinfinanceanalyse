from __future__ import annotations

import csv
from decimal import Decimal
from pathlib import Path
import sys

from fastapi.testclient import TestClient
from openpyxl import Workbook

from backend.app.governance.settings import get_settings
from backend.app.repositories.governance_repo import (
    CACHE_BUILD_RUN_STREAM,
    GovernanceRepository,
)
from backend.app.schemas.materialize import CacheBuildRunRecord
from backend.app.tasks.product_category_pnl import PRODUCT_CATEGORY_ADJUSTMENT_STREAM
from tests.helpers import load_module


LEDGER_PREFIX = "\u603b\u8d26\u5bf9\u8d26"
AVG_PREFIX = "\u65e5\u5747"


def test_product_category_materialize_and_api_flow(tmp_path, monkeypatch):
    data_root = tmp_path / "data_input"
    source_dir = data_root / "pnl_\u603b\u8d26\u5bf9\u8d26-\u65e5\u5747"
    source_dir.mkdir(parents=True)

    _write_month_pair(source_dir, "202601", january=True)
    _write_month_pair(source_dir, "202602", january=False)

    duckdb_path = tmp_path / "moss.duckdb"
    governance_dir = tmp_path / "governance"
    monkeypatch.setenv("MOSS_DUCKDB_PATH", str(duckdb_path))
    monkeypatch.setenv("MOSS_PRODUCT_CATEGORY_SOURCE_DIR", str(source_dir))
    monkeypatch.setenv("MOSS_GOVERNANCE_PATH", str(governance_dir))
    get_settings.cache_clear()

    GovernanceRepository(base_dir=governance_dir).append(
        PRODUCT_CATEGORY_ADJUSTMENT_STREAM,
        {
            "report_date": "2026-02-28",
            "operator": "DELTA",
            "approval_status": "approved",
            "account_code": "13304010001",
            "currency": "CNX",
            "monthly_pnl": "5",
        },
    )

    task_module = sys.modules.get("backend.app.tasks.product_category_pnl")
    if task_module is None:
        task_module = load_module(
            "backend.app.tasks.product_category_pnl",
            "backend/app/tasks/product_category_pnl.py",
        )
    payload = task_module.materialize_product_category_pnl.fn(
        duckdb_path=str(duckdb_path),
        source_dir=str(source_dir),
        governance_dir=str(governance_dir),
    )

    assert payload["status"] == "completed"
    assert payload["month_count"] == 2
    assert payload["report_dates"] == ["2026-01-31", "2026-02-28"]

    main_module = load_module("backend.app.main", "backend/app/main.py")
    client = TestClient(main_module.app)

    dates_response = client.get("/ui/pnl/product-category/dates")
    assert dates_response.status_code == 200
    assert dates_response.json()["result"]["report_dates"] == ["2026-02-28", "2026-01-31"]

    refresh_response = client.post("/ui/pnl/product-category/refresh")
    assert refresh_response.status_code == 200
    refresh_payload = refresh_response.json()
    assert refresh_payload["status"] in {"queued", "completed"}
    assert refresh_payload["job_name"] == "product_category_pnl"
    assert refresh_payload["trigger_mode"] in {"async", "sync-fallback"}
    if refresh_payload["status"] == "completed":
        assert refresh_payload["report_dates"] == ["2026-01-31", "2026-02-28"]
    else:
        assert refresh_payload["cache_key"] == "product_category_pnl.formal"
        queued_status = client.get(
            "/ui/pnl/product-category/refresh-status",
            params={"run_id": refresh_payload["run_id"]},
        )
        assert queued_status.status_code == 200
        assert queued_status.json()["status"] in {"queued", "running"}

    manual_adjustment_response = client.post(
        "/ui/pnl/product-category/manual-adjustments",
        json={
            "report_date": "2026-02-28",
            "operator": "DELTA",
            "approval_status": "approved",
            "account_code": "13304010001",
            "currency": "CNX",
            "account_name": "测试科目",
            "monthly_pnl": "6",
        },
    )
    assert manual_adjustment_response.status_code == 200
    manual_adjustment_payload = manual_adjustment_response.json()
    assert manual_adjustment_payload["stream"] == "product_category_pnl_adjustments"
    assert manual_adjustment_payload["account_code"] == "13304010001"
    appended_adjustments = GovernanceRepository(base_dir=governance_dir).read_all(
        PRODUCT_CATEGORY_ADJUSTMENT_STREAM
    )
    assert any(
        row.get("adjustment_id") == manual_adjustment_payload["adjustment_id"]
        for row in appended_adjustments
    )

    monthly_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-01-31", "view": "monthly"},
    )
    assert monthly_response.status_code == 200
    monthly_payload = monthly_response.json()
    assert monthly_payload["result_meta"]["basis"] == "formal"
    assert monthly_payload["result_meta"]["scenario_flag"] is False
    assert monthly_payload["result"]["view"] == "monthly"

    rows = monthly_payload["result"]["rows"]
    bond_parent = next(row for row in rows if row["category_id"] == "bond_investment")
    assert bond_parent["children"] == [
        "bond_tpl",
        "bond_ac",
        "bond_ac_other",
        "bond_fvoci",
        "bond_valuation_spread",
    ]
    assert Decimal(str(bond_parent["cnx_scale"])) > 0
    assert Decimal(str(bond_parent["business_net_income"])) > 0

    qtd_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-02-28", "view": "qtd"},
    )
    assert qtd_response.status_code == 200
    qtd_payload = qtd_response.json()
    assert qtd_payload["result"]["view"] == "qtd"
    lending = next(
        row for row in qtd_payload["result"]["rows"] if row["category_id"] == "interbank_lending_assets"
    )
    assert Decimal(str(lending["cnx_scale"])) > 50

    feb_monthly_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-02-28", "view": "monthly"},
    )
    assert feb_monthly_response.status_code == 200
    feb_monthly_payload = feb_monthly_response.json()
    feb_bond = next(
        row for row in feb_monthly_payload["result"]["rows"] if row["category_id"] == "bond_tpl"
    )
    assert Decimal(str(feb_bond["cnx_cash"])) > 0

    scenario_response = client.get(
        "/ui/pnl/product-category",
        params={
            "report_date": "2026-02-28",
            "view": "monthly",
            "scenario_rate_pct": "2.5",
        },
    )
    assert scenario_response.status_code == 200
    scenario_payload = scenario_response.json()
    assert scenario_payload["result_meta"]["basis"] == "scenario"
    assert scenario_payload["result_meta"]["scenario_flag"] is True
    scenario_asset_total = scenario_payload["result"]["asset_total"]
    baseline_asset_total = feb_monthly_payload["result"]["asset_total"]
    assert Decimal(str(scenario_asset_total["cny_ftp"])) != Decimal(str(baseline_asset_total["cny_ftp"]))
    get_settings.cache_clear()


def test_product_category_refresh_queue_and_status_flow(tmp_path, monkeypatch):
    data_root = tmp_path / "data_input"
    source_dir = data_root / "pnl_总账对账-日均"
    source_dir.mkdir(parents=True)

    _write_month_pair(source_dir, "202601", january=True)

    duckdb_path = tmp_path / "moss.duckdb"
    governance_dir = tmp_path / "governance"
    monkeypatch.setenv("MOSS_DUCKDB_PATH", str(duckdb_path))
    monkeypatch.setenv("MOSS_PRODUCT_CATEGORY_SOURCE_DIR", str(source_dir))
    monkeypatch.setenv("MOSS_GOVERNANCE_PATH", str(governance_dir))
    get_settings.cache_clear()

    task_module = sys.modules.get("backend.app.tasks.product_category_pnl")
    if task_module is None:
        task_module = load_module(
            "backend.app.tasks.product_category_pnl",
            "backend/app/tasks/product_category_pnl.py",
        )

    queued_messages: list[dict[str, object]] = []

    def fake_send(**kwargs):
        queued_messages.append(kwargs)
        return None

    monkeypatch.setattr(
        "backend.app.services.product_category_pnl_service.materialize_product_category_pnl.send",
        fake_send,
    )

    main_module = load_module("backend.app.main", "backend/app/main.py")
    client = TestClient(main_module.app)

    refresh_response = client.post("/ui/pnl/product-category/refresh")
    assert refresh_response.status_code == 200
    refresh_payload = refresh_response.json()
    assert refresh_payload["status"] == "queued"
    assert refresh_payload["job_name"] == "product_category_pnl"
    assert refresh_payload["trigger_mode"] == "async"
    assert queued_messages[0]["run_id"] == refresh_payload["run_id"]

    queued_status = client.get(
        "/ui/pnl/product-category/refresh-status",
        params={"run_id": refresh_payload["run_id"]},
    )
    assert queued_status.status_code == 200
    assert queued_status.json()["status"] == "queued"

    GovernanceRepository(base_dir=governance_dir).append(
        CACHE_BUILD_RUN_STREAM,
        CacheBuildRunRecord(
            run_id=refresh_payload["run_id"],
            job_name="product_category_pnl",
            status="completed",
            cache_key="product_category_pnl.formal",
            lock="lock:duckdb:product-category-pnl",
            source_version="sv_product_category_test",
            vendor_version="vv_none",
        ).model_dump(),
    )

    completed_status = client.get(
        "/ui/pnl/product-category/refresh-status",
        params={"run_id": refresh_payload["run_id"]},
    )
    assert completed_status.status_code == 200
    completed_payload = completed_status.json()
    assert completed_payload["status"] == "completed"
    assert completed_payload["run_id"] == refresh_payload["run_id"]
    get_settings.cache_clear()


def test_manual_adjustment_changes_read_model_and_can_be_revoked(tmp_path, monkeypatch):
    data_root = tmp_path / "data_input"
    source_dir = data_root / "pnl_总账对账-日均"
    source_dir.mkdir(parents=True)

    _write_month_pair(source_dir, "202602", january=False)

    duckdb_path = tmp_path / "moss.duckdb"
    governance_dir = tmp_path / "governance"
    monkeypatch.setenv("MOSS_DUCKDB_PATH", str(duckdb_path))
    monkeypatch.setenv("MOSS_PRODUCT_CATEGORY_SOURCE_DIR", str(source_dir))
    monkeypatch.setenv("MOSS_GOVERNANCE_PATH", str(governance_dir))
    get_settings.cache_clear()

    task_module = sys.modules.get("backend.app.tasks.product_category_pnl")
    if task_module is None:
        task_module = load_module(
            "backend.app.tasks.product_category_pnl",
            "backend/app/tasks/product_category_pnl.py",
        )
    task_module.materialize_product_category_pnl.fn(
        duckdb_path=str(duckdb_path),
        source_dir=str(source_dir),
        governance_dir=str(governance_dir),
    )

    monkeypatch.setattr(
        "backend.app.services.product_category_pnl_service.materialize_product_category_pnl.send",
        lambda **_: (_ for _ in ()).throw(RuntimeError("queue disabled")),
    )

    main_module = load_module("backend.app.main", "backend/app/main.py")
    client = TestClient(main_module.app)

    baseline_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-02-28", "view": "monthly"},
    )
    assert baseline_response.status_code == 200
    baseline_rows = baseline_response.json()["result"]["rows"]
    baseline_tpl = next(row for row in baseline_rows if row["category_id"] == "bond_tpl")
    baseline_cash = Decimal(str(baseline_tpl["cnx_cash"]))

    create_response = client.post(
        "/ui/pnl/product-category/manual-adjustments",
        json={
            "report_date": "2026-02-28",
            "operator": "DELTA",
            "approval_status": "approved",
            "account_code": "51402010001",
            "currency": "CNX",
            "account_name": "测试科目",
            "monthly_pnl": "10",
        },
    )
    assert create_response.status_code == 200
    adjustment_id = create_response.json()["adjustment_id"]

    list_response = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={"report_date": "2026-02-28"},
    )
    assert list_response.status_code == 200
    listed_payload = list_response.json()
    listed = listed_payload["adjustments"]
    assert any(item["adjustment_id"] == adjustment_id for item in listed)
    assert [event["event_type"] for event in listed_payload["events"]] == ["created"]

    refresh_response = client.post("/ui/pnl/product-category/refresh")
    assert refresh_response.status_code == 200
    assert refresh_response.json()["status"] == "completed"

    adjusted_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-02-28", "view": "monthly"},
    )
    assert adjusted_response.status_code == 200
    adjusted_rows = adjusted_response.json()["result"]["rows"]
    adjusted_tpl = next(row for row in adjusted_rows if row["category_id"] == "bond_tpl")
    adjusted_cash = Decimal(str(adjusted_tpl["cnx_cash"]))
    assert adjusted_cash == baseline_cash + Decimal("10")

    edit_response = client.post(
        f"/ui/pnl/product-category/manual-adjustments/{adjustment_id}/edit",
        json={
            "report_date": "2026-02-28",
            "operator": "DELTA",
            "approval_status": "approved",
            "account_code": "51402010001",
            "currency": "CNX",
            "account_name": "测试科目",
            "monthly_pnl": "12",
        },
    )
    assert edit_response.status_code == 200
    assert edit_response.json()["event_type"] == "edited"

    refresh_after_edit = client.post("/ui/pnl/product-category/refresh")
    assert refresh_after_edit.status_code == 200
    assert refresh_after_edit.json()["status"] == "completed"

    edited_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-02-28", "view": "monthly"},
    )
    assert edited_response.status_code == 200
    edited_rows = edited_response.json()["result"]["rows"]
    edited_tpl = next(row for row in edited_rows if row["category_id"] == "bond_tpl")
    edited_cash = Decimal(str(edited_tpl["cnx_cash"]))
    assert edited_cash == baseline_cash + Decimal("12")

    revoke_response = client.post(
        f"/ui/pnl/product-category/manual-adjustments/{adjustment_id}/revoke",
    )
    assert revoke_response.status_code == 200
    assert revoke_response.json()["approval_status"] == "rejected"

    refresh_after_revoke = client.post("/ui/pnl/product-category/refresh")
    assert refresh_after_revoke.status_code == 200
    assert refresh_after_revoke.json()["status"] == "completed"

    reverted_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-02-28", "view": "monthly"},
    )
    assert reverted_response.status_code == 200
    reverted_rows = reverted_response.json()["result"]["rows"]
    reverted_tpl = next(row for row in reverted_rows if row["category_id"] == "bond_tpl")
    reverted_cash = Decimal(str(reverted_tpl["cnx_cash"]))
    assert reverted_cash == baseline_cash

    restore_response = client.post(
        f"/ui/pnl/product-category/manual-adjustments/{adjustment_id}/restore",
    )
    assert restore_response.status_code == 200
    assert restore_response.json()["event_type"] == "restored"
    assert restore_response.json()["approval_status"] == "approved"

    refresh_after_restore = client.post("/ui/pnl/product-category/refresh")
    assert refresh_after_restore.status_code == 200
    assert refresh_after_restore.json()["status"] == "completed"

    restored_response = client.get(
        "/ui/pnl/product-category",
        params={"report_date": "2026-02-28", "view": "monthly"},
    )
    assert restored_response.status_code == 200
    restored_rows = restored_response.json()["result"]["rows"]
    restored_tpl = next(row for row in restored_rows if row["category_id"] == "bond_tpl")
    restored_cash = Decimal(str(restored_tpl["cnx_cash"]))
    assert restored_cash == baseline_cash + Decimal("12")

    history_after_restore = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={"report_date": "2026-02-28"},
    )
    assert history_after_restore.status_code == 200
    history_events = [event["event_type"] for event in history_after_restore.json()["events"]]
    assert history_events == ["restored", "revoked", "edited", "created"]

    filtered_history = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "adjustment_id": adjustment_id,
            "event_type": "edited",
            "limit": 1,
            "offset": 0,
        },
    )
    assert filtered_history.status_code == 200
    filtered_payload = filtered_history.json()
    assert filtered_payload["adjustment_count"] == 1
    assert filtered_payload["event_total"] == 1
    assert filtered_payload["event_limit"] == 1
    assert filtered_payload["event_offset"] == 0
    assert filtered_payload["adjustments"][0]["event_type"] == "restored"
    assert [event["event_type"] for event in filtered_payload["events"]] == ["edited"]
    get_settings.cache_clear()


def test_manual_adjustment_list_supports_exact_id_and_current_state_paging(tmp_path, monkeypatch):
    data_root = tmp_path / "data_input"
    source_dir = data_root / "pnl_总账对账-日均"
    source_dir.mkdir(parents=True)

    _write_month_pair(source_dir, "202602", january=False)

    duckdb_path = tmp_path / "moss.duckdb"
    governance_dir = tmp_path / "governance"
    monkeypatch.setenv("MOSS_DUCKDB_PATH", str(duckdb_path))
    monkeypatch.setenv("MOSS_PRODUCT_CATEGORY_SOURCE_DIR", str(source_dir))
    monkeypatch.setenv("MOSS_GOVERNANCE_PATH", str(governance_dir))
    get_settings.cache_clear()

    task_module = sys.modules.get("backend.app.tasks.product_category_pnl")
    if task_module is None:
        task_module = load_module(
            "backend.app.tasks.product_category_pnl",
            "backend/app/tasks/product_category_pnl.py",
        )
    task_module.materialize_product_category_pnl.fn(
        duckdb_path=str(duckdb_path),
        source_dir=str(source_dir),
        governance_dir=str(governance_dir),
    )

    main_module = load_module("backend.app.main", "backend/app/main.py")
    client = TestClient(main_module.app)

    first_response = client.post(
        "/ui/pnl/product-category/manual-adjustments",
        json={
            "report_date": "2026-02-28",
            "operator": "DELTA",
            "approval_status": "approved",
            "account_code": "51402010001",
            "currency": "CNX",
            "account_name": "测试科目1",
            "monthly_pnl": "10",
        },
    )
    second_response = client.post(
        "/ui/pnl/product-category/manual-adjustments",
        json={
            "report_date": "2026-02-28",
            "operator": "DELTA",
            "approval_status": "pending",
            "account_code": "51403010001",
            "currency": "CNY",
            "account_name": "测试科目2",
            "monthly_pnl": "6",
        },
    )
    assert first_response.status_code == 200
    assert second_response.status_code == 200

    first_id = first_response.json()["adjustment_id"]
    second_id = second_response.json()["adjustment_id"]

    paged_current = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "adjustment_limit": 1,
            "adjustment_offset": 1,
        },
    )
    assert paged_current.status_code == 200
    paged_payload = paged_current.json()
    assert paged_payload["adjustment_count"] == 2
    assert paged_payload["adjustment_limit"] == 1
    assert paged_payload["adjustment_offset"] == 1
    assert len(paged_payload["adjustments"]) == 1

    fuzzy_fragment = first_id.split("-")[1][:8]
    fuzzy_match = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "adjustment_id": fuzzy_fragment,
        },
    )
    assert fuzzy_match.status_code == 200
    fuzzy_payload = fuzzy_match.json()
    assert fuzzy_payload["adjustment_count"] == 1
    assert fuzzy_payload["adjustments"][0]["adjustment_id"] == first_id

    exact_miss = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "adjustment_id": fuzzy_fragment,
            "adjustment_id_exact": "true",
        },
    )
    assert exact_miss.status_code == 200
    assert exact_miss.json()["adjustment_count"] == 0

    exact_hit = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "adjustment_id": second_id,
            "adjustment_id_exact": "true",
        },
    )
    assert exact_hit.status_code == 200
    exact_payload = exact_hit.json()
    assert exact_payload["adjustment_count"] == 1
    assert exact_payload["adjustments"][0]["adjustment_id"] == second_id

    export_response = client.get(
        "/ui/pnl/product-category/manual-adjustments/export",
        params={
            "report_date": "2026-02-28",
        },
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith("text/csv")
    export_text = export_response.text
    assert "Current State" in export_text
    assert "Event Timeline" in export_text
    assert first_id in export_text
    assert second_id in export_text
    get_settings.cache_clear()


def test_manual_adjustment_query_contract_defaults_and_sorting(tmp_path, monkeypatch):
    client, governance_dir = _build_product_category_client(tmp_path, monkeypatch)

    repo = GovernanceRepository(base_dir=governance_dir)
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="created",
        created_at="2026-02-01T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="300",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-b",
        event_type="created",
        created_at="2026-02-02T00:00:00Z",
        report_date="2026-02-28",
        approval_status="pending",
        account_code="100",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-c",
        event_type="created",
        created_at="2026-02-03T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="200",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="edited",
        created_at="2026-02-07T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="300",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-c",
        event_type="revoked",
        created_at="2026-02-08T00:00:00Z",
        report_date="2026-02-28",
        approval_status="rejected",
        account_code="200",
    )

    default_response = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={"report_date": "2026-02-28"},
    )
    assert default_response.status_code == 200
    default_payload = default_response.json()
    assert [item["adjustment_id"] for item in default_payload["adjustments"]] == [
        "adj-c",
        "adj-a",
        "adj-b",
    ]
    assert [item["event_type"] for item in default_payload["events"]] == [
        "revoked",
        "edited",
        "created",
        "created",
        "created",
    ]

    sorted_response = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "current_sort_field": "account_code",
            "current_sort_dir": "asc",
            "event_sort_field": "adjustment_id",
            "event_sort_dir": "asc",
        },
    )
    assert sorted_response.status_code == 200
    sorted_payload = sorted_response.json()
    assert [item["adjustment_id"] for item in sorted_payload["adjustments"]] == [
        "adj-b",
        "adj-c",
        "adj-a",
    ]
    assert [item["adjustment_id"] for item in sorted_payload["events"]] == [
        "adj-a",
        "adj-a",
        "adj-b",
        "adj-c",
        "adj-c",
    ]
    get_settings.cache_clear()


def test_manual_adjustment_query_contract_created_at_range_is_inclusive_and_stateful(
    tmp_path, monkeypatch
):
    client, governance_dir = _build_product_category_client(tmp_path, monkeypatch)

    repo = GovernanceRepository(base_dir=governance_dir)
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="created",
        created_at="2026-02-01T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="300",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-b",
        event_type="created",
        created_at="2026-02-02T00:00:00Z",
        report_date="2026-02-28",
        approval_status="pending",
        account_code="100",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="edited",
        created_at="2026-02-07T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="300",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-c",
        event_type="created",
        created_at="2026-02-08T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="200",
    )

    response = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "created_at_from": "2026-02-01T00:00:00Z",
            "created_at_to": "2026-02-07T00:00:00Z",
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert [item["adjustment_id"] for item in payload["adjustments"]] == [
        "adj-a",
        "adj-b",
    ]
    assert [(item["adjustment_id"], item["event_type"]) for item in payload["events"]] == [
        ("adj-a", "edited"),
        ("adj-b", "created"),
        ("adj-a", "created"),
    ]
    get_settings.cache_clear()


def test_manual_adjustment_query_contract_returns_422_for_invalid_sort_values(
    tmp_path, monkeypatch
):
    client, _ = _build_product_category_client(tmp_path, monkeypatch)

    invalid_field = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "current_sort_field": "operator",
        },
    )
    assert invalid_field.status_code == 422

    invalid_direction = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "event_sort_dir": "down",
        },
    )
    assert invalid_direction.status_code == 422

    invalid_report_date = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-01,2026-02-28",
        },
    )
    assert invalid_report_date.status_code == 422
    get_settings.cache_clear()


def test_manual_adjustment_query_contract_composes_filters(tmp_path, monkeypatch):
    client, governance_dir = _build_product_category_client(tmp_path, monkeypatch)

    repo = GovernanceRepository(base_dir=governance_dir)
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="created",
        created_at="2026-02-02T00:00:00Z",
        report_date="2026-02-28",
        approval_status="pending",
        account_code="100",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="edited",
        created_at="2026-02-05T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="100",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-b",
        event_type="created",
        created_at="2026-02-04T00:00:00Z",
        report_date="2026-02-28",
        approval_status="pending",
        account_code="100",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-c",
        event_type="created",
        created_at="2026-02-04T00:00:00Z",
        report_date="2026-02-28",
        approval_status="pending",
        account_code="200",
    )

    response = client.get(
        "/ui/pnl/product-category/manual-adjustments",
        params={
            "report_date": "2026-02-28",
            "account_code": "100",
            "approval_status": "pending",
            "event_type": "created",
            "created_at_from": "2026-02-02T00:00:00Z",
            "created_at_to": "2026-02-04T00:00:00Z",
            "current_sort_field": "adjustment_id",
            "current_sort_dir": "asc",
            "event_sort_field": "created_at",
            "event_sort_dir": "asc",
        },
    )
    assert response.status_code == 200
    payload = response.json()
    assert [item["adjustment_id"] for item in payload["adjustments"]] == ["adj-b"]
    assert [(item["adjustment_id"], item["event_type"]) for item in payload["events"]] == [
        ("adj-a", "created"),
        ("adj-b", "created"),
    ]
    get_settings.cache_clear()


def test_manual_adjustment_export_locks_csv_order_under_sort_and_range(tmp_path, monkeypatch):
    client, governance_dir = _build_product_category_client(tmp_path, monkeypatch)

    repo = GovernanceRepository(base_dir=governance_dir)
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="created",
        created_at="2026-02-02T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="300",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-a",
        event_type="edited",
        created_at="2026-02-05T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="300",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-b",
        event_type="created",
        created_at="2026-02-04T00:00:00Z",
        report_date="2026-02-28",
        approval_status="pending",
        account_code="100",
    )
    _append_adjustment_event(
        repo,
        adjustment_id="adj-c",
        event_type="created",
        created_at="2026-02-04T00:00:00Z",
        report_date="2026-02-28",
        approval_status="approved",
        account_code="200",
    )

    export_response = client.get(
        "/ui/pnl/product-category/manual-adjustments/export",
        params={
            "report_date": "2026-02-28",
            "created_at_from": "2026-02-02T00:00:00Z",
            "created_at_to": "2026-02-04T00:00:00Z",
            "current_sort_field": "account_code",
            "current_sort_dir": "asc",
            "event_sort_field": "adjustment_id",
            "event_sort_dir": "desc",
        },
    )
    assert export_response.status_code == 200

    current_rows, event_rows = _parse_adjustment_csv_sections(export_response.text)

    assert [row["adjustment_id"] for row in current_rows] == [
        "adj-b",
        "adj-c",
    ]
    assert [row["account_code"] for row in current_rows] == ["100", "200"]
    assert [row["adjustment_id"] for row in event_rows] == [
        "adj-c",
        "adj-b",
        "adj-a",
    ]
    assert [row["event_type"] for row in event_rows] == [
        "created",
        "created",
        "created",
    ]
    get_settings.cache_clear()


def _build_product_category_client(tmp_path, monkeypatch) -> tuple[TestClient, Path]:
    governance_dir = tmp_path / "governance"
    monkeypatch.setenv("MOSS_DUCKDB_PATH", str(tmp_path / "moss.duckdb"))
    monkeypatch.setenv("MOSS_PRODUCT_CATEGORY_SOURCE_DIR", str(tmp_path / "data_input"))
    monkeypatch.setenv("MOSS_GOVERNANCE_PATH", str(governance_dir))
    get_settings.cache_clear()

    main_module = load_module("backend.app.main", "backend/app/main.py")
    return TestClient(main_module.app), governance_dir


def _append_adjustment_event(
    repo: GovernanceRepository,
    *,
    adjustment_id: str,
    event_type: str,
    created_at: str,
    report_date: str,
    approval_status: str,
    account_code: str,
) -> None:
    repo.append(
        PRODUCT_CATEGORY_ADJUSTMENT_STREAM,
        {
            "adjustment_id": adjustment_id,
            "event_type": event_type,
            "created_at": created_at,
            "report_date": report_date,
            "operator": "DELTA",
            "approval_status": approval_status,
            "account_code": account_code,
            "currency": "CNX",
            "account_name": f"Account {account_code}",
            "monthly_pnl": "1",
        },
    )


def _parse_adjustment_csv_sections(content: str) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    lines = content.splitlines()
    current_start = lines.index("Current State")
    event_start = lines.index("Event Timeline")

    current_header = lines[current_start + 1].split(",")
    current_data_lines = [line for line in lines[current_start + 2 : event_start] if line]
    event_header = lines[event_start + 1].split(",")
    event_data_lines = [line for line in lines[event_start + 2 :] if line]

    def build_rows(header: list[str], data_lines: list[str]) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for line in data_lines:
            values = next(csv.reader([line]))
            rows.append(dict(zip(header, values, strict=True)))
        return rows

    return build_rows(current_header, current_data_lines), build_rows(event_header, event_data_lines)


def _write_month_pair(target_dir: Path, month_key: str, *, january: bool) -> None:
    report_date = f"{month_key[:4]}-{month_key[4:]}-{'31' if january else '28'}"
    ledger_path = target_dir / f"{LEDGER_PREFIX}{month_key}.xlsx"
    avg_path = target_dir / f"{AVG_PREFIX}{month_key}.xlsx"
    _write_ledger_workbook(ledger_path, report_date, january=january)
    _write_average_workbook(avg_path, report_date, january=january)


def _write_ledger_workbook(path: Path, report_date: str, *, january: bool) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name, currency in (("\u7efc\u672c", "CNX"), ("\u4eba\u6c11\u5e01", "CNY")):
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(["\u603b\u8d26\uff08\u5bf9\u8d26\uff09"])
        worksheet.append([None])
        worksheet.append(["\u5206\u7c7b\u8d26\uff1a"])
        worksheet.append(["\u516c\u53f8\uff1a"])
        worksheet.append([f"\u4f1a\u8ba1\u671f\u95f4\uff1a  {report_date[:8]}01--{report_date}  \u5e01\u79cd\uff1a {currency}"])
        worksheet.append(
            [
                "\u7ec4\u5408\u79d1\u76ee\u4ee3\u7801",
                "\u7ec4\u5408\u79d1\u76ee\u540d\u79f0",
                "\u5e01\u79cd",
                "\u671f\u521d\u4f59\u989d",
                "\u672c\u671f\u501f\u65b9",
                "\u672c\u671f\u8d37\u65b9",
                "\u671f\u672b\u4f59\u989d",
            ]
        )
        for row in _ledger_rows(currency, january=january):
            worksheet.append(row)
    workbook.save(path)


def _write_average_workbook(path: Path, report_date: str, *, january: bool) -> None:
    workbook = Workbook()
    annual = workbook.active
    annual.title = "\u5e74"
    monthly = workbook.create_sheet(title="\u6708")
    for worksheet in (annual, monthly):
        worksheet.append([f"\u673a\u6784\uff1a 199200\u9752\u5c9b\u94f6\u884c"])
        worksheet.append([f"\u65e5\u671f\uff1a {report_date[:8]}01 \u81f3 {report_date}"])
        worksheet.append(["\u5e01\u79cd", "\u79d1\u76ee", "\u79d1\u76ee\u65e5\u5747\u4f59\u989d", None])
        for row in _avg_rows(monthly=(worksheet.title == "\u6708"), january=january):
            worksheet.append(row)
    workbook.save(path)


def _ledger_rows(currency: str, *, january: bool) -> list[list[object]]:
    cny = currency == "CNY"
    scale = lambda a, b: b if cny else a
    pnl = lambda a, b: b if cny else a

    rows: list[list[object]] = []

    rows.extend(
        [
            _balance_row("12000010005", "\u62c6\u653e\u57ce\u5e02\u5546\u4e1a\u94f6\u884c", currency, scale(55, 45)),
            _balance_row("12100010001", "\u5176\u4ed6\u62c6\u653e\u540c\u4e1a", currency, scale(12, 9)),
            _balance_row("14000010001", "\u4e70\u5165\u8fd4\u552e\u8bc1\u5238", currency, scale(230, 230)),
            _balance_row("14004000001", "\u4e70\u5165\u8fd4\u552e\u91d1\u878d\u8d44\u4ea7-\u6263\u51cf\u79d1\u76ee1", currency, scale(10, 10)),
            _balance_row("14005000001", "\u4e70\u5165\u8fd4\u552e\u91d1\u878d\u8d44\u4ea7-\u6263\u51cf\u79d1\u76ee2", currency, scale(5, 5)),
            _balance_row("14101010001", "\u516c\u5141\u4ef7\u503c\u53d8\u52a8\u8ba1\u5165\u635f\u76ca\u7684\u91d1\u878d\u8d44\u4ea7-\u6210\u672c", currency, scale(200 if january else 210, 190 if january else 200)),
            _balance_row("14201010001", "\u4ee5\u644a\u4f59\u6210\u672c\u8ba1\u91cf\u7684\u503a\u5238\u6295\u8d44-\u6210\u672c", currency, scale(280 if january else 310, 230 if january else 260)),
            _balance_row("14301040001", "\u4ee5\u644a\u4f59\u6210\u672c\u8ba1\u91cf\u7684\u5176\u4ed6\u6295\u8d44-\u6210\u672c", currency, scale(145 if january else 150, 135 if january else 140)),
            _balance_row("14301010001", "\u4e2d\u56fd\u653f\u5e9c\u503a\u5238", currency, scale(12, 12)),
            _balance_row("14301010002", "\u4e2d\u56fd\u653f\u5e9c\u503a\u5238-\u5e94\u8ba1\u5229\u606f", currency, scale(3, 3)),
            _balance_row("14401010001", "\u516c\u5141\u4ef7\u503c\u53d8\u52a8\u8ba1\u5165\u7efc\u5408\u6536\u76ca\u7684\u91d1\u878d\u8d44\u4ea7-\u6210\u672c", currency, scale(170 if january else 180, 150 if january else 160)),
            _balance_row("43001020003", "\u5229\u7387\u5de5\u5177\u884d\u751f\u91d1\u878d\u8d44\u4ea7", currency, scale(18 if january else 22, 12 if january else 15)),
            _balance_row("23401000001", "\u540c\u4e1a\u94f6\u884c\u7c7b\u6d3b\u671f\u5b58\u6b3e", currency, scale(250 if january else 260, 200 if january else 210)),
            _balance_row("23501000001", "\u540c\u4e1a\u94f6\u884c\u7c7b\u5b9a\u671f\u5b58\u6b3e", currency, scale(40 if january else 42, 35 if january else 36)),
            _balance_row("24100010004", "\u80a1\u4efd\u5236\u5546\u4e1a\u94f6\u884c\u62c6\u5165", currency, scale(80 if january else 90, 60 if january else 70)),
            _balance_row("24200010001", "\u540c\u4e1a\u62c6\u5165-\u5176\u4ed6", currency, scale(8, 6)),
            _balance_row("25500010001", "\u5356\u51fa\u56de\u8d2d", currency, scale(40 if january else 45, 30 if january else 35)),
            _balance_row("27205000001", "\u5e94\u4ed8\u540c\u4e1a\u5b58\u5355-\u9762\u503c", currency, scale(130 if january else 150, 120 if january else 140)),
            _balance_row("27206000001", "\u5e94\u4ed8\u540c\u4e1a\u5b58\u5355-\u5229\u606f\u8c03\u6574", currency, scale(6, 6)),
            _balance_row("27201010001", "\u53d1\u884c\u91d1\u878d\u503a\u5238\u9762\u503c", currency, scale(90 if january else 95, 90 if january else 95)),
            _balance_row("24501000004", "\u4fe1\u7528\u8054\u7ed3\u7968\u636e-\u6210\u672c", currency, scale(18.51, 18.51)),
            _balance_row("24501000005", "\u4fe1\u7528\u8054\u7ed3\u7968\u636e-\u516c\u5141\u4ef7\u503c\u53d8\u52a8", currency, scale(0.5, 0.5)),
        ]
    )

    rows.extend(
        [
            _pnl_row("13302000016", "\u5176\u5b83\u5e94\u8ba1\u5229\u606f-\u5b58\u653e\u5883\u5185\u540c\u4e1a\uff08\u81ea\u8425\uff09", currency, pnl(11 if january else 12, 9 if january else 10)),
            _pnl_row("13302000018", "\u5176\u5b83\u5e94\u8ba1\u5229\u606f-\u62c6\u653e\u5883\u5185\u540c\u4e1a\uff08\u81ea\u8425\uff09", currency, pnl(7 if january else 8, 6 if january else 7)),
            _pnl_row("13302000032", "\u5176\u5b83\u5e94\u8ba1\u5229\u606f-\u540c\u4e1a\u501f\u6b3e\u5229\u606f\u6536\u5165", currency, pnl(2 if january else 2, 2 if january else 1)),
            _pnl_row("13302000022", "\u5176\u5b83\u5e94\u8ba1\u5229\u606f-\u8d28\u62bc\u5f0f\u8bc1\u5238\u56de\u8d2d\u534f\u8bae\u501f\u51fa\u6b3e", currency, pnl(5 if january else 6, 4 if january else 5)),
            _pnl_row("51402010001", "\u4ee5\u516c\u5141\u4ef7\u503c\u8ba1\u91cf\u4e14\u5176\u53d8\u52a8\u8ba1\u5165\u5f53\u671f\u635f\u76ca\u7684\u91d1\u878d\u503a\u5229\u606f\u6536\u5165", currency, pnl(10 if january else 11, 9 if january else 10)),
            _pnl_row("51402010004", "\u4ee5\u516c\u5141\u4ef7\u503c\u8ba1\u91cf\u4e14\u5176\u53d8\u52a8\u8ba1\u5165\u5f53\u671f\u635f\u76ca\u7684\u4fe1\u7528\u503a\u5229\u606f\u6536\u5165", currency, pnl(9 if january else 10, 8 if january else 9)),
            _pnl_row("51601010001", "\u516c\u5141\u4ef7\u503c\u53d8\u52a8\u8ba1\u5165\u5f53\u671f\u635f\u76ca", currency, pnl(1.2 if january else 1.5, 1.1 if january else 1.4)),
            _pnl_row("51701010001", "\u516c\u5141\u4ef7\u503c\u53d8\u52a8\u8ba1\u5165\u635f\u76ca\u7684\u91d1\u878d\u8d44\u4ea7\u5dee\u4ef7\u6536\u76ca", currency, pnl(4 if january else 5, 3 if january else 4)),
            _pnl_row("51701010002", "\u516c\u5141\u4ef7\u503c\u53d8\u52a8\u8ba1\u5165\u635f\u76ca\u7684\u91d1\u878d\u8d44\u4ea7\u6301\u6709\u671f\u95f4\u6536\u76ca", currency, pnl(3 if january else 4, 2 if january else 3)),
            _pnl_row("51701010004", "\u4ea4\u6613\u6027\u8d44\u4ea7\u6295\u8d44\u6536\u76ca4", currency, pnl(1 if january else 1.2, 1 if january else 1.1)),
            _pnl_row("51701010006", "\u4ea4\u6613\u6027\u8d44\u4ea7\u6295\u8d44\u6536\u76ca6", currency, pnl(0.8 if january else 0.9, 0.7 if january else 0.8)),
            _pnl_row("51404010001", "\u4ee5\u644a\u4f59\u6210\u672c\u8ba1\u91cf\u7684\u91d1\u878d\u503a\u5229\u606f\u6536\u5165", currency, pnl(16 if january else 17, 15 if january else 16)),
            _pnl_row("51404010004", "\u4ee5\u644a\u4f59\u6210\u672c\u8ba1\u91cf\u7684\u4fe1\u7528\u503a\u5229\u606f\u6536\u5165", currency, pnl(14 if january else 15, 13 if january else 14)),
            _pnl_row("51401000004", "\u4ee5\u644a\u4f59\u6210\u672c\u8ba1\u91cf\u5176\u4ed6\u6295\u8d44\u5229\u606f\u6536\u5165", currency, pnl(6 if january else 7, 5 if january else 6)),
            _pnl_row("51401000001", "\u5176\u4ed6\u6295\u8d44\u6263\u51cf1", currency, pnl(1, 1)),
            _pnl_row("51401000002", "\u5176\u4ed6\u6295\u8d44\u6263\u51cf2", currency, pnl(1, 1)),
            _pnl_row("51403010001", "\u4ee5\u516c\u5141\u4ef7\u503c\u8ba1\u91cf\u4e14\u5176\u53d8\u52a8\u8ba1\u5165\u5176\u4ed6\u7efc\u5408\u6536\u76ca\u7684\u91d1\u878d\u503a\u5229\u606f\u6536\u5165", currency, pnl(11 if january else 12, 10 if january else 11)),
            _pnl_row("51403010004", "\u4ee5\u516c\u5141\u4ef7\u503c\u8ba1\u91cf\u4e14\u5176\u53d8\u52a8\u8ba1\u5165\u5176\u4ed6\u7efc\u5408\u6536\u76ca\u7684\u4fe1\u7528\u503a\u5229\u606f\u6536\u5165", currency, pnl(10 if january else 11, 9 if january else 10)),
            _pnl_row("51702010001", "\u516c\u5141\u4ef7\u503c\u8ba1\u5165\u7efc\u5408\u6536\u76ca\u7684\u91d1\u878d\u8d44\u4ea7\u5dee\u4ef7\u6536\u76ca", currency, pnl(2.1 if january else 2.2, 1.9 if january else 2.0)),
            _pnl_row("51703010001", "\u5176\u4ed6\u6295\u8d44\u4ef0\u503c\u53ca\u4e70\u5356\u4ef7\u5dee", currency, pnl(0.5, 0.5)),
            _pnl_row("51603010005", "\u975e\u5957\u671f\u4fdd\u503c\u7c7b\u884d\u751f\u5de5\u5177\u516c\u5141\u4ef7\u503c\u53d8\u52a8\u635f\u76ca", currency, pnl(3 if january else -1, 2 if january else -1)),
            _pnl_row("51102000004", "\u4e2d\u95f4\u4e1a\u52a1\u6536\u5165-1", currency, pnl(0.12, 0.11)),
            _pnl_row("51203010001", "\u4e2d\u95f4\u4e1a\u52a1\u6536\u5165-2", currency, pnl(0.14, 0.13)),
            _pnl_row("52206000001", "\u540c\u4e1a\u5b58\u653e\u5229\u606f\u652f\u51fa-\u5883\u5185", currency, pnl(-6 if january else -7, -5 if january else -6)),
            _pnl_row("52204000001", "\u540c\u4e1a\u62c6\u5165\u5229\u606f\u652f\u51fa", currency, pnl(-4 if january else -5, -3 if january else -4)),
            _pnl_row("52208000001", "\u5356\u51fa\u56de\u8d2d\u5229\u606f\u652f\u51fa1", currency, pnl(-2 if january else -3, -2 if january else -2.5)),
            _pnl_row("52210000001", "\u5356\u51fa\u56de\u8d2d\u5229\u606f\u652f\u51fa2", currency, pnl(-1, -1)),
            _pnl_row("52300030001", "\u53d1\u884c\u540c\u4e1a\u5b58\u5355\u5229\u606f\u652f\u51fa", currency, pnl(-8 if january else -9, -7 if january else -8)),
            _pnl_row("51605010002", "\u4fe1\u7528\u8054\u7ed3\u7968\u636e\u516c\u5141\u4ef7\u503c\u53d8\u52a8", currency, pnl(0.02, 0.02)),
            _pnl_row("51710020002", "\u4fe1\u7528\u8054\u7ed3\u7968\u636e\u6295\u8d44\u6536\u76ca", currency, pnl(0.02, 0.02)),
        ]
    )
    return rows


def _avg_rows(*, monthly: bool, january: bool) -> list[list[object]]:
    scale = {
        "120": (55, 45, 60, 50),
        "121": (12, 9, 13, 10),
        "140": (230, 230, 240, 240),
        "14004": (10, 10, 10, 10),
        "14005": (5, 5, 5, 5),
        "141": (200 if january else 210, 190 if january else 200, 210 if january else 220, 200 if january else 210),
        "142": (280 if january else 310, 230 if january else 260, 300 if january else 330, 250 if january else 280),
        "143": (160 if january else 165, 150 if january else 155, 165 if january else 170, 155 if january else 160),
        "14301010001": (12, 12, 12, 12),
        "14301010002": (3, 3, 3, 3),
        "144": (170 if january else 180, 150 if january else 160, 180 if january else 190, 160 if january else 170),
        "234": (250 if january else 260, 200 if january else 210, 250 if january else 260, 200 if january else 210),
        "235": (40 if january else 42, 35 if january else 36, 40 if january else 42, 35 if january else 36),
        "241": (80 if january else 90, 60 if january else 70, 80 if january else 90, 60 if january else 70),
        "242": (8, 6, 8, 6),
        "255": (40 if january else 45, 30 if january else 35, 40 if january else 45, 30 if january else 35),
        "27205": (130 if january else 150, 120 if january else 140, 130 if january else 150, 120 if january else 140),
        "27205000001": (130 if january else 150, 120 if january else 140, 130 if january else 150, 120 if january else 140),
        "27206000001": (6, 6, 6, 6),
        "24501000004": (18.51, 18.51, 18.51, 18.51),
        "24501000005": (0.5, 0.5, 0.5, 0.5),
    }
    rows: list[list[object]] = []
    for account_code, values in scale.items():
        cnx_monthly, cny_monthly, cnx_annual, cny_annual = values
        cnx_balance = cnx_monthly if monthly else cnx_annual
        cny_balance = cny_monthly if monthly else cny_annual
        rows.append(["CNX", int(account_code), cnx_balance, None])
        rows.append(["CNY", int(account_code), cny_balance, None])
    return rows


def _balance_row(account_code: str, account_name: str, currency: str, ending_balance: float) -> list[object]:
    return [int(account_code), account_name, currency, ending_balance, 0, 0, ending_balance]


def _pnl_row(account_code: str, account_name: str, currency: str, monthly_pnl: float) -> list[object]:
    if monthly_pnl >= 0:
        return [int(account_code), account_name, currency, 0, 0, monthly_pnl, -monthly_pnl]
    return [int(account_code), account_name, currency, 0, abs(monthly_pnl), 0, abs(monthly_pnl)]
