from __future__ import annotations

import io
import json
from pathlib import Path

import duckdb
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.governance.settings import get_settings
from tests.helpers import load_module
from tests.test_ledger_import_flow import (
    _configure_ledger_import_env,
    _ledger_csv_bytes,
    _ledger_row_values,
    _pack_sample,
)


def test_fastapi_application_registers_ledger_analytics_routes(tmp_path, monkeypatch):
    _configure_ledger_import_env(tmp_path, monkeypatch)

    app = load_module("backend.app.main", "backend/app/main.py").app
    paths = {route.path for route in app.routes}

    assert "/api/ledger/dates" in paths
    assert "/api/ledger/dashboard" in paths
    assert "/api/ledger/positions" in paths
    assert "/api/ledger/export/positions" in paths
    get_settings.cache_clear()


def test_ledger_dates_dashboard_positions_and_export_use_imported_snapshot(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    _import_two_position_fixture(duckdb_path)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    dates = client.get("/api/ledger/dates")
    assert dates.status_code == 200
    assert dates.json()["data"]["items"] == ["2026-03-17"]
    assert dates.json()["metadata"]["batch_id"] == 1
    assert dates.json()["metadata"]["no_data"] is False

    dashboard = client.get("/api/ledger/dashboard", params={"as_of_date": "2026-03-17"})
    assert dashboard.status_code == 200
    dashboard_payload = dashboard.json()
    assert dashboard_payload["data"] == {
        "as_of_date": "2026-03-17",
        "asset_face_amount": 1.0,
        "liability_face_amount": 0.5,
        "net_face_exposure": 0.5,
        "alert_count": 0,
    }
    assert dashboard_payload["metadata"]["stale"] is False
    assert dashboard_payload["metadata"]["fallback"] is False
    assert dashboard_payload["trace"]["requested_as_of_date"] == "2026-03-17"
    assert dashboard_payload["trace"]["resolved_as_of_date"] == "2026-03-17"

    positions = client.get(
        "/api/ledger/positions",
        params={
            "as_of_date": "2026-03-17",
            "direction": "ASSET",
            "account_category_std": "银行账户",
            "page": 1,
            "page_size": 10,
        },
    )
    assert positions.status_code == 200
    positions_payload = positions.json()
    assert positions_payload["data"]["total"] == 1
    assert positions_payload["metadata"]["no_data"] is False
    item = positions_payload["data"]["items"][0]
    assert item["direction"] == "ASSET"
    assert item["batch_id"] == 1
    assert item["row_no"] == 1
    assert item["trace"] == {
        "position_key": item["position_key"],
        "batch_id": 1,
        "row_no": 1,
    }
    assert positions_payload["trace"]["filters"]["account_category_std"] == "银行账户"

    exported = client.get(
        "/api/ledger/export/positions",
        params={"as_of_date": "2026-03-17", "direction": "ASSET"},
    )
    assert exported.status_code == 200
    assert exported.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert exported.headers["x-ledger-batch-id"] == "1"
    workbook = load_workbook(io.BytesIO(exported.content), read_only=True)
    try:
        rows = list(workbook["positions"].iter_rows(values_only=True))
        metadata = dict(workbook["metadata"].iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()

    assert rows[0][:3] == ("position_key", "batch_id", "row_no")
    assert len(rows) == 2
    assert rows[1][1:3] == (1, 1)
    assert metadata["as_of_date"] == "2026-03-17"
    assert metadata["requested_as_of_date"] == "2026-03-17"
    assert metadata["resolved_as_of_date"] == "2026-03-17"
    assert metadata["total"] == 1
    assert json.loads(metadata["filters"]) == {
        "account_category_std": None,
        "asset_class_std": None,
        "bond_code": None,
        "cost_center": None,
        "direction": "ASSET",
        "portfolio": None,
    }
    get_settings.cache_clear()


def test_ledger_dashboard_uses_existing_zqtz_snapshot_source(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    _insert_zqtz_snapshot_fixture(duckdb_path)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    dates = client.get("/api/ledger/dates")
    assert dates.status_code == 200
    assert dates.json()["data"]["items"] == ["2026-03-31"]
    assert dates.json()["metadata"]["batch_id"] == "ib-existing-zqtz"
    assert dates.json()["metadata"]["source_version"] == "sv-existing-zqtz"

    dashboard = client.get("/api/ledger/dashboard", params={"as_of_date": "2026-03-31"})
    assert dashboard.status_code == 200
    dashboard_payload = dashboard.json()
    assert dashboard_payload["data"] == {
        "as_of_date": "2026-03-31",
        "asset_face_amount": 1.0,
        "liability_face_amount": 0.5,
        "net_face_exposure": 0.5,
        "alert_count": 0,
    }
    assert dashboard_payload["metadata"]["rule_version"] == "rv_snapshot_zqtz_tyw_v1"

    positions = client.get(
        "/api/ledger/positions",
        params={"as_of_date": "2026-03-31", "direction": "LIABILITY", "page": 1, "page_size": 10},
    )
    assert positions.status_code == 200
    positions_payload = positions.json()
    assert positions_payload["data"]["total"] == 1
    item = positions_payload["data"]["items"][0]
    assert item["batch_id"] == "ib-existing-zqtz"
    assert item["direction"] == "LIABILITY"
    assert item["bond_code"] == "LIAB-ZQTZ-001"
    assert item["face_amount"] == 50000000.0
    assert item["trace"]["ingest_batch_id"] == "ib-existing-zqtz"
    get_settings.cache_clear()


def test_ledger_imports_and_dates_reject_unknown_query_parameters(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    _import_two_position_fixture(duckdb_path)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    imports = client.get("/api/ledger/imports", params={"foo": "bar"})
    dates = client.get("/api/ledger/dates", params={"foo": "bar"})

    assert imports.status_code == 400
    assert imports.json()["error"]["code"] == "LEDGER_IMPORTS_INVALID_REQUEST"
    assert dates.status_code == 400
    assert dates.json()["error"]["code"] == "LEDGER_DATES_INVALID_REQUEST"
    get_settings.cache_clear()


def test_ledger_import_rejects_unknown_query_parameters(tmp_path, monkeypatch):
    _configure_ledger_import_env(tmp_path, monkeypatch)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    response = client.post("/api/ledger/import", params={"foo": "bar"}, content=b"")

    assert response.status_code == 400
    assert response.json()["error"]["code"] == "LEDGER_IMPORT_INVALID_REQUEST"
    assert "Unsupported query parameter" in response.json()["error"]["message"]
    get_settings.cache_clear()


def test_ledger_dashboard_falls_back_to_latest_snapshot(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    _import_two_position_fixture(duckdb_path)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    response = client.get("/api/ledger/dashboard", params={"as_of_date": "2026-03-18"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["data"]["as_of_date"] == "2026-03-17"
    assert payload["metadata"]["stale"] is True
    assert payload["metadata"]["fallback"] is True
    assert payload["trace"]["requested_as_of_date"] == "2026-03-18"
    assert payload["trace"]["resolved_as_of_date"] == "2026-03-17"
    get_settings.cache_clear()


def test_ledger_export_fallback_metadata_preserves_requested_and_resolved_dates(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    _import_two_position_fixture(duckdb_path)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    response = client.get(
        "/api/ledger/export/positions",
        params={"as_of_date": "2026-03-18", "direction": "ASSET"},
    )

    assert response.status_code == 200
    assert response.headers["x-ledger-fallback"] == "true"
    workbook = load_workbook(io.BytesIO(response.content), read_only=True)
    try:
        metadata = dict(workbook["metadata"].iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()
    assert metadata["requested_as_of_date"] == "2026-03-18"
    assert metadata["resolved_as_of_date"] == "2026-03-17"
    assert metadata["fallback"] is True
    get_settings.cache_clear()


def test_ledger_dashboard_keeps_missing_side_null_but_net_uses_known_side(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    service_mod = load_module(
        "backend.app.services.ledger_import_service",
        "backend/app/services/ledger_import_service.py",
    )
    csv_bytes = _ledger_csv_bytes(
        service_mod,
        [
            _ledger_row_values(
                service_mod,
                bond_code="LIABILITY-ONLY",
                account_category="发行类债券",
                asset_class="发行类债券",
                face_amount="100000000",
                as_of_date="2026-03-17",
            ),
        ],
    )
    service_mod.LedgerImportService(str(duckdb_path)).import_file(
        file_name="ZQTZSHOW-20260317.csv",
        content=csv_bytes,
    )
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    response = client.get("/api/ledger/dashboard", params={"as_of_date": "2026-03-17"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["data"]["asset_face_amount"] is None
    assert payload["data"]["liability_face_amount"] == 1.0
    assert payload["data"]["net_face_exposure"] == -1.0
    get_settings.cache_clear()


def test_ledger_dashboard_no_data_keeps_all_kpis_null(tmp_path, monkeypatch):
    _configure_ledger_import_env(tmp_path, monkeypatch)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    response = client.get("/api/ledger/dashboard", params={"as_of_date": "2026-03-17"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["data"] == {
        "as_of_date": None,
        "asset_face_amount": None,
        "liability_face_amount": None,
        "net_face_exposure": None,
        "alert_count": None,
    }
    assert payload["metadata"]["no_data"] is True
    get_settings.cache_clear()


def test_ledger_batch2_rejects_non_contract_query_aliases_and_bad_dates(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    _import_two_position_fixture(duckdb_path)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    date_alias = client.get("/api/ledger/dashboard", params={"date": "2026-03-17"})
    bad_date = client.get("/api/ledger/dashboard", params={"as_of_date": "20260317"})
    account_alias = client.get(
        "/api/ledger/positions",
        params={"as_of_date": "2026-03-17", "account_category": "银行账户"},
    )
    asset_alias = client.get(
        "/api/ledger/export/positions",
        params={"as_of_date": "2026-03-17", "asset_class": "持有至到期类资产"},
    )

    assert date_alias.status_code == 400
    assert "Unsupported query parameter" in date_alias.json()["error"]["message"]
    assert bad_date.status_code == 400
    assert "YYYY-MM-DD" in bad_date.json()["error"]["message"]
    assert account_alias.status_code == 400
    assert "account_category" in account_alias.json()["error"]["message"]
    assert asset_alias.status_code == 400
    assert "asset_class" in asset_alias.json()["error"]["message"]
    get_settings.cache_clear()


def test_ledger_positions_no_data_and_invalid_direction_are_explicit(tmp_path, monkeypatch):
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    _import_two_position_fixture(duckdb_path)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    no_data = client.get(
        "/api/ledger/positions",
        params={"as_of_date": "2026-03-17", "direction": "ASSET", "portfolio": "NO-SUCH"},
    )
    assert no_data.status_code == 200
    assert no_data.json()["data"]["items"] == []
    assert no_data.json()["data"]["total"] == 0
    assert no_data.json()["metadata"]["no_data"] is True

    invalid = client.get(
        "/api/ledger/positions",
        params={"as_of_date": "2026-03-17", "direction": "SIDEWAYS"},
    )
    assert invalid.status_code == 400
    assert invalid.json()["error"]["code"] == "LEDGER_POSITIONS_INVALID_REQUEST"
    get_settings.cache_clear()


def test_ledger_dashboard_real_pack_20260317_golden_kpis(tmp_path, monkeypatch):
    sample = _pack_sample("ZQTZSHOW-20260317.xls")
    if sample is None:
        pytest.skip("bank ledger pack sample ZQTZSHOW-20260317.xls is not available")
    duckdb_path = _configure_ledger_import_env(tmp_path, monkeypatch)
    service_mod = load_module(
        "backend.app.services.ledger_import_service",
        "backend/app/services/ledger_import_service.py",
    )
    service_mod.LedgerImportService(str(duckdb_path)).import_file(
        file_name=Path(sample).name,
        content=Path(sample).read_bytes(),
    )
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    response = client.get("/api/ledger/dashboard", params={"as_of_date": "2026-03-17"})

    assert response.status_code == 200
    payload = response.json()
    assert payload["data"]["asset_face_amount"] == 3289.07
    assert payload["data"]["liability_face_amount"] == 1231.77
    assert payload["data"]["net_face_exposure"] == 2057.31
    assert payload["metadata"]["source_version"].startswith("sv_ledger_")
    assert payload["metadata"]["rule_version"] == "position_key_contract_v1"
    get_settings.cache_clear()


def test_ledger_schema_registry_adds_position_snapshot_agg():
    loader = load_module(
        "backend.app.schema_registry.duckdb_loader",
        "backend/app/schema_registry/duckdb_loader.py",
    )
    conn = duckdb.connect(":memory:")
    try:
        loader.apply_registry_sql(conn)
        tables = {
            row[0]
            for row in conn.execute(
                """
                select table_name
                from information_schema.tables
                where table_schema = 'main'
                """
            ).fetchall()
        }
    finally:
        conn.close()

    assert "position_snapshot_agg" in tables


def _import_two_position_fixture(duckdb_path: Path) -> None:
    service_mod = load_module(
        "backend.app.services.ledger_import_service",
        "backend/app/services/ledger_import_service.py",
    )
    csv_bytes = _ledger_csv_bytes(
        service_mod,
        [
            _ledger_row_values(
                service_mod,
                bond_code="ASSET-001",
                account_category="银行账户",
                asset_class="持有至到期类资产",
                face_amount="100000000",
                as_of_date="2026-03-17",
            ),
            _ledger_row_values(
                service_mod,
                bond_code="LIABILITY-001",
                account_category="发行类债券",
                asset_class="发行类债券",
                face_amount="50000000",
                as_of_date="2026-03-17",
            ),
        ],
    )
    service_mod.LedgerImportService(str(duckdb_path)).import_file(
        file_name="ZQTZSHOW-20260317.csv",
        content=csv_bytes,
    )


def _insert_zqtz_snapshot_fixture(duckdb_path: Path) -> None:
    snapshot_mod = load_module(
        "backend.app.repositories.snapshot_repo",
        "backend/app/repositories/snapshot_repo.py",
    )
    conn = duckdb.connect(str(duckdb_path), read_only=False)
    try:
        snapshot_mod.ensure_snapshot_tables(conn)
        rows = [
            ("ASSET-ZQTZ-001", "Asset bond", False, "asset-trace", 100000000),
            ("LIAB-ZQTZ-001", "Issued bond", True, "liability-trace", 50000000),
        ]
        conn.executemany(
            """
            insert into zqtz_bond_daily_snapshot (
              report_date, instrument_code, instrument_name, portfolio_name, cost_center,
              account_category, asset_class, bond_type, business_type_primary, issuer_name,
              industry_name, rating, currency_code, face_value_native, market_value_native,
              amortized_cost_native, accrued_interest_native, coupon_rate, ytm_value,
              maturity_date, next_call_date, overdue_days, is_issuance_like, interest_mode,
              source_version, rule_version, ingest_batch_id, trace_id, value_date, customer_attribute
            ) values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                [
                    "2026-03-31",
                    code,
                    name,
                    "BANK-BOOK",
                    "5010",
                    "ISSUANCE" if issuance else "BANK",
                    "ISSUANCE" if issuance else "HOLD_TO_MATURITY",
                    "TEST",
                    "TEST_BUSINESS",
                    "Issuer",
                    "Banking",
                    "AAA",
                    "CNY",
                    face,
                    face,
                    face,
                    0,
                    "0.03",
                    "0.031",
                    "2030-03-31",
                    None,
                    0,
                    issuance,
                    "FIXED",
                    "sv-existing-zqtz",
                    "rv_snapshot_zqtz_tyw_v1",
                    "ib-existing-zqtz",
                    trace_id,
                    "2026-03-31",
                    "CORP",
                ]
                for code, name, issuance, trace_id, face in rows
            ],
        )
    finally:
        conn.close()
