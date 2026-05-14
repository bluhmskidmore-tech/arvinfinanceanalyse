from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from tests.helpers import load_module


def test_pnl_source_rules_detect_all_planned_pnl_source_families():
    module = load_module(
        "backend.app.services.source_rules",
        "backend/app/services/source_rules.py",
    )

    cases = [
        ("FI损益202512.xls", "pnl"),
        ("FI损益202602.xls", "pnl"),
        ("非标514-20250101-1231.xlsx", "pnl_514"),
        ("非标516-20250101-1231.xlsx", "pnl_516"),
        ("非标517-20250101-1231.xlsx", "pnl_517"),
    ]

    for file_name, expected_family in cases:
        assert module.detect_source_family(file_name) == expected_family


def test_pnl_source_rules_extract_report_dates_from_fi_and_nonstd_file_names():
    module = load_module(
        "backend.app.services.source_rules",
        "backend/app/services/source_rules.py",
    )

    cases = [
        ("FI损益202512.xls", "2025-12-31"),
        ("FI损益202602.xls", "2026-02-28"),
        ("非标514-20250101-1231.xlsx", "2025-12-31"),
        ("非标516-20250101-1231.xlsx", "2025-12-31"),
        ("非标517-20260101-0228.xlsx", "2026-02-28"),
    ]

    for file_name, expected_report_date in cases:
        assert module.extract_report_date_from_name(file_name) == expected_report_date


def test_ingest_scan_enriches_manifest_rows_for_all_planned_pnl_source_families(tmp_path):
    ingest_module = load_module(
        "backend.app.services.ingest_service",
        "backend/app/services/ingest_service.py",
    )

    files = [
        ("pnl", "FI损益202512.xls"),
        ("pnl_514", "非标514-20250101-1231.xlsx"),
        ("pnl_516", "非标516-20250101-1231.xlsx"),
        ("pnl_517", "非标517-20250101-1231.xlsx"),
    ]

    data_root = tmp_path / "data_input"
    for directory, file_name in files:
        path = data_root / directory / file_name
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(b"fixture")

    service = ingest_module.IngestService(data_root=data_root)
    rows = sorted(
        service.scan(),
        key=lambda row: Path(row["file_path"]).parent.name,
    )

    assert [row["source_family"] for row in rows] == ["pnl", "pnl_514", "pnl_516", "pnl_517"]
    assert [row["report_date"] for row in rows] == [
        "2025-12-31",
        "2025-12-31",
        "2025-12-31",
        "2025-12-31",
    ]


def test_load_latest_pnl_refresh_input_keeps_manifest_report_date_for_archived_fi_names(tmp_path):
    governance_module = load_module(
        "backend.app.repositories.governance_repo",
        "backend/app/repositories/governance_repo.py",
    )
    source_module = load_module(
        "backend.app.services.pnl_source_service",
        "backend/app/services/pnl_source_service.py",
    )

    archive_dir = tmp_path / "archive"
    archive_dir.mkdir(parents=True, exist_ok=True)
    archived_fi = archive_dir / "FI_202602__archived.xls"
    source_fi = Path(__file__).resolve().parents[1] / "data_input" / "pnl" / "FI损益202602.xls"
    archived_fi.write_bytes(source_fi.read_bytes())

    governance_repo = governance_module.GovernanceRepository(base_dir=tmp_path / "governance")
    governance_repo.append(
        governance_module.SOURCE_MANIFEST_STREAM,
        {
            "source_family": "pnl",
            "report_date": "2026-02-28",
            "source_file": "FI损益202602.xls",
            "source_version": "sv_manifest_fi",
            "ingest_batch_id": "ib_manifest_fi",
            "created_at": "2026-04-12T00:00:00+00:00",
            "status": "completed",
            "archived_path": str(archived_fi),
        },
    )

    refresh = source_module.load_latest_pnl_refresh_input(
        governance_dir=tmp_path / "governance",
        data_root=tmp_path / "missing-data-root",
        report_date="2026-02-28",
    )

    assert refresh.report_date == "2026-02-28"
    assert refresh.fi_rows
    assert refresh.fi_rows[0]["report_date"] == "2026-02-28"


def test_pnl_refresh_discovers_processed_fi_month_files(tmp_path):
    source_module = load_module(
        "backend.app.services.pnl_source_service",
        "backend/app/services/pnl_source_service.py",
    )

    processed_dir = tmp_path / "data_input" / "pnl" / "processed"
    processed_dir.mkdir(parents=True)
    source_fi = Path(__file__).resolve().parents[1] / "data_input" / "pnl" / "FI损益202602.xls"
    processed_fi = processed_dir / "FI损益202502.xls"
    processed_fi.write_bytes(source_fi.read_bytes())

    report_dates = source_module.list_pnl_refresh_report_dates(
        governance_dir=tmp_path / "governance",
        data_root=tmp_path / "data_input",
    )
    assert "2025-02-28" in report_dates

    refresh = source_module.load_latest_pnl_refresh_input(
        governance_dir=tmp_path / "governance",
        data_root=tmp_path / "data_input",
        report_date="2025-02-28",
    )

    assert refresh.fi_rows
    assert refresh.fi_rows[0]["report_date"] == "2025-02-28"


def test_load_latest_pnl_refresh_input_marks_usd_rows_for_fx_conversion(tmp_path) -> None:
    source_module = load_module(
        "backend.app.services.pnl_source_service",
        "backend/app/services/pnl_source_service.py",
    )

    data_root = tmp_path / "data_input" / "pnl"
    data_root.mkdir(parents=True, exist_ok=True)
    source_fi = Path(__file__).resolve().parents[1] / "data_input" / "pnl" / "FI损益202602.xls"
    target_fi = data_root / source_fi.name
    target_fi.write_bytes(source_fi.read_bytes())

    refresh = source_module.load_latest_pnl_refresh_input(
        governance_dir=tmp_path / "governance",
        data_root=tmp_path / "data_input",
        report_date="2026-02-28",
    )

    assert refresh.fi_rows
    usd_rows = [row for row in refresh.fi_rows if row.get("fx_base_currency") == "USD"]
    assert usd_rows
    assert all(row["currency_basis"] == "CNY" for row in usd_rows)


def test_nonstd_pnl_parser_skips_blank_pivot_sheet_before_detail_rows(tmp_path) -> None:
    source_module = load_module(
        "backend.app.services.pnl_source_service",
        "backend/app/services/pnl_source_service.py",
    )

    workbook = Workbook()
    blank_sheet = workbook.active
    blank_sheet.title = "Sheet5"
    detail_sheet = workbook.create_sheet("Sheet1")
    detail_sheet.append(["会计分录详情表"])
    detail_sheet.append(
        [
            "账务流水号",
            "账务日期",
            "会计事件",
            "成本中心",
            "投资组合",
            "资产代码",
            "借贷标识",
            "科目号",
            "金额",
        ]
    )
    detail_sheet.append(
        [
            "1",
            "2026-04-30",
            "公允价值变动",
            "5010",
            "FIOA",
            "BOND-516",
            "贷",
            "51601010005",
            "123.45",
        ]
    )
    path = tmp_path / "非标516-20260101-0430.xlsx"
    workbook.save(path)

    snapshot = source_module.PnlSourceSnapshot(
        source_family="pnl_516",
        report_date="2026-04-30",
        path=path,
        source_version="sv-test-516",
        ingest_batch_id="ib-test-516",
        created_at="2026-05-14T00:00:00+00:00",
    )

    rows = source_module._parse_nonstd_rows(snapshot, bucket="516")

    assert len(rows) == 1
    assert rows[0]["voucher_date"] == "2026-04-30"
    assert rows[0]["asset_code"] == "BOND-516"
    assert rows[0]["account_code"] == "51601010005"
    assert rows[0]["raw_amount"] == source_module.Decimal("123.45")


def test_nonstd_pnl_parser_prefers_signed_amount_column_when_abs_amount_exists(tmp_path) -> None:
    source_module = load_module(
        "backend.app.services.pnl_source_service",
        "backend/app/services/pnl_source_service.py",
    )

    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "Sheet1"
    detail_sheet.append(["会计分录详情表"])
    detail_sheet.append(
        [
            "账务日期",
            "会计事件",
            "成本中心",
            "投资组合",
            "资产代码",
            "借贷标识",
            "科目号",
            "金额",
            "amount",
        ]
    )
    detail_sheet.append(
        [
            "2026-04-30",
            "冲销前一日估值",
            "5010",
            "FIOA",
            "BOND-516",
            "贷",
            "51601010004",
            "-287599.70",
            "287599.70",
        ]
    )
    path = tmp_path / "非标516-20260101-0430.xlsx"
    workbook.save(path)

    snapshot = source_module.PnlSourceSnapshot(
        source_family="pnl_516",
        report_date="2026-04-30",
        path=path,
        source_version="sv-test-516-signed",
        ingest_batch_id="ib-test-516-signed",
        created_at="2026-05-14T00:00:00+00:00",
    )

    rows = source_module._parse_nonstd_rows(snapshot, bucket="516")

    assert len(rows) == 1
    assert rows[0]["raw_amount"] == source_module.Decimal("-287599.70")
