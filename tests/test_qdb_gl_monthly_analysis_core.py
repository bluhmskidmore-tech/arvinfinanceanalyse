from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tests.helpers import ROOT, load_module


B = 100_000_000


def test_parse_daily_avg_coerces_codes_and_builds_expected_groups(tmp_path):
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    avg_path, _ledger_path = _write_month_pair(tmp_path, "202602")

    parsed = module.parse_daily_avg(avg_path)

    assert sorted(parsed) == [
        "年日均_CNX_11d",
        "年日均_CNX_3d",
        "年日均_CNX_5d",
        "年日均_CNX_7d",
        "年日均_CNY_11d",
        "年日均_CNY_3d",
        "年日均_CNY_5d",
        "年日均_CNY_7d",
        "月日均_CNX_11d",
        "月日均_CNX_3d",
        "月日均_CNX_5d",
        "月日均_CNX_7d",
        "月日均_CNY_11d",
        "月日均_CNY_3d",
        "月日均_CNY_5d",
        "月日均_CNY_7d",
    ]
    month_11 = parsed["月日均_CNX_11d"]
    assert "12301000001" in {row["科目代码"] for row in month_11}
    assert next(row["日均余额"] for row in month_11 if row["科目代码"] == "14001000001") == 230 * B


def test_build_workbook_payload_computes_metrics_gap_alerts_and_foreign_split(tmp_path):
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    avg_path, ledger_path = _write_month_pair(tmp_path, "202602")

    parsed_avg = module.parse_daily_avg(avg_path)
    parsed_ledger = module.parse_general_ledger(ledger_path)
    merged = module.merge_all(parsed_ledger, parsed_avg)
    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202602",
        merged_data=merged,
    )

    sheet_titles = [sheet["title"] for sheet in workbook["sheets"]]
    assert sheet_titles == [
        "经营概览",
        "3位科目总览",
        "资产结构",
        "负债结构",
        "贷款行业",
        "存款行业_活期",
        "存款行业_定期",
        "行业存贷差",
        "11位偏离TOP",
        "异动预警",
        "外币分析",
        "分部基础规模",
        "公司规模",
        "零售规模",
        "金融市场规模",
        "收益率分析（总账可复算）",
    ]

    overview = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "overview")
    overview_rows = {row["指标"]: row["值"] for row in overview["rows"]}
    assert overview_rows["贷款总额(亿)"] == 1086
    assert overview_rows["存款总额(亿)"] == 1600
    assert overview_rows["存贷比%"] == 67.88
    assert overview_rows["贷款减值准备率%"] == 2.76

    gap_sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "industry_gap")
    agriculture = next(row for row in gap_sheet["rows"] if row["行业"] == "农林牧渔")
    assert agriculture["存贷差_时点"] == -600
    assert agriculture["存贷差_日均"] == -400

    alerts_sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "alerts")
    alerts = {(row["科目代码"], row["异动类型"]) for row in alerts_sheet["rows"]}
    assert ("14001000001", "同业业务月末操纵嫌疑") in alerts
    assert ("13001000001", "逾期贷款异动") in alerts
    assert ("23401000001", "同业业务月末操纵嫌疑") in alerts

    foreign_sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "foreign_currency")
    foreign_rows = {row["科目代码"]: row for row in foreign_sheet["rows"]}
    assert foreign_rows["14201000001"]["外币部分"] == 50
    assert foreign_rows["14401000001"]["外币部分"] == 30


def test_qdb_gl_position_vs_ledger_check_uses_reconciliation_helper():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    rows_3d = [
        {"科目代码": "101", "期末余额": 100},
        {"科目代码": "201", "期末余额": -40},
    ]

    checks = module._qdb_gl_position_vs_ledger_check(
        position_totals={"总资产": 100, "总负债": 40},
        ledger_rows_3d=rows_3d,
    )

    assert checks == [
        {
            "dimension": "total_assets",
            "position_value": 100.0,
            "ledger_value": 100.0,
            "diff": 0.0,
            "breached": False,
        },
        {
            "dimension": "total_liabilities",
            "position_value": 40.0,
            "ledger_value": 40.0,
            "diff": 0.0,
            "breached": False,
        },
        {
            "dimension": "net_assets",
            "position_value": 60.0,
            "ledger_value": 60.0,
            "diff": 0.0,
            "breached": False,
        },
    ]


def test_asset_liability_overview_uses_financial_indicator_subject_scope():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )

    metrics = module.compute_asset_liability_structure(
        [
            {"科目代码": "101", "期末余额": 100 * B},
            {"科目代码": "123", "期末余额": 40 * B},
            {"科目代码": "129", "期末余额": 5 * B},
            {"科目代码": "130", "期末余额": 4 * B},
            {"科目代码": "132", "期末余额": 3 * B},
            {"科目代码": "133", "期末余额": 999 * B},
            {"科目代码": "136", "期末余额": 2 * B},
            {"科目代码": "131", "期末余额": -10 * B},
            {"科目代码": "201", "期末余额": -50 * B},
            {"科目代码": "202", "期末余额": -6 * B},
            {"科目代码": "216", "期末余额": -7 * B},
            {"科目代码": "217", "期末余额": -8 * B},
            {"科目代码": "601", "期末余额": 700 * B},
        ],
    )

    assert metrics["总资产"] == 1143 * B
    assert metrics["贷款总额"] == 54 * B
    assert metrics["存款总额"] == 71 * B
    assert metrics["贷款减值准备率%"] == module.Decimal("18.51851851851851851851851852")


def test_qdb_gl_workbook_alerts_position_vs_ledger_mismatch():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    merged_data = {
        "3位": [
            {"科目代码": "101", "名称": "现金", "期初余额": 100, "期末余额": 100, "变动额": 0, "月日均": 100, "年日均": 100},
            {"科目代码": "201", "名称": "单位活期存款", "期初余额": -40, "期末余额": -40, "变动额": 0, "月日均": -40, "年日均": -40},
        ],
        "11位": [],
        "5位_公司贷款": [],
        "5位_活期存款": [],
        "5位_定期存款": [],
        "外币分析": [],
    }
    original = module.compute_asset_liability_structure

    def mismatched_metrics(rows_3d):
        metrics = original(rows_3d)
        metrics["总资产"] = metrics["总资产"] + 1
        return metrics

    module.compute_asset_liability_structure = mismatched_metrics
    try:
        workbook = module.build_qdb_gl_monthly_analysis_workbook(
            report_month="202602",
            merged_data=merged_data,
        )
    finally:
        module.compute_asset_liability_structure = original

    alerts_sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "alerts")
    assert any(
        row["异动类型"] == "position_vs_ledger_reconciliation"
        and row["科目代码"] == "total_assets"
        for row in alerts_sheet["rows"]
    )


def test_exported_workbook_contains_all_required_sheets(tmp_path):
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    avg_path, ledger_path = _write_month_pair(tmp_path, "202602")
    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202602",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    content = module.export_qdb_gl_monthly_analysis_workbook_xlsx_bytes(workbook)
    exported = load_workbook(BytesIO(content))
    assert "经营概览" in exported.sheetnames
    assert "异动预警" in exported.sheetnames
    assert "外币分析" in exported.sheetnames


def test_real_202603_qdb_gl_overview_matches_first_wave_financial_indicator_rules():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()
    avg_path = _real_month_source(source_dir, "\u65e5\u5747", "202603")
    ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", "202603")

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    overview = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "overview")
    value_key = overview["columns"][1]

    assert [row[value_key] for row in overview["rows"]] == [
        8144.05,
        7651.02,
        4189.47,
        5115.96,
        3386.16,
        81.89,
        2.7,
        62.27,
        24.2,
        4.97,
    ]


def test_real_202603_qdb_gl_daily_average_deviation_outputs_keep_summary_and_top_fields():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()
    avg_path = _real_month_source(source_dir, "\u65e5\u5747", "202603")
    ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", "202603")

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    summary = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "summary_3d")
    assert summary["columns"] == [
        "科目代码",
        "名称",
        "期初余额",
        "期末余额",
        "变动额",
        "月日均",
        "年日均",
        "偏离额",
        "偏离%",
        "趋势额",
        "趋势%",
    ]
    company_loan = next(row for row in summary["rows"] if row["科目代码"] == "123")
    assert company_loan == {
        "科目代码": "123",
        "名称": "公司贷款",
        "期初余额": 2993.86,
        "期末余额": 3044.84,
        "变动额": 50.97,
        "月日均": 3009.19,
        "年日均": 2969.67,
        "偏离额": 35.65,
        "偏离%": 1.18,
        "趋势额": 39.52,
        "趋势%": 1.33,
    }

    top_11d = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "top_11d")
    assert top_11d["columns"] == [
        "科目代码",
        "科目名称",
        "期末余额",
        "月日均",
        "年日均",
        "偏离额",
        "偏离%",
        "趋势额",
        "趋势%",
    ]
    assert top_11d["rows"][0] == {
        "科目代码": "40700030046",
        "科目名称": "大额支付来账过渡户",
        "期末余额": 25.2,
        "月日均": 7.77,
        "年日均": 6.08,
        "偏离额": 17.44,
        "偏离%": 224.56,
        "趋势额": 1.69,
        "趋势%": 27.8,
    }


def test_real_202603_qdb_gl_workbook_includes_segment_base_scale_sheet_from_financial_indicator_sample():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()
    avg_path = _real_month_source(source_dir, "\u65e5\u5747", "202603")
    ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", "202603")

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    indicator_key = "\u6307\u6807"
    spot_key = "\u65f6\u70b9\u4f59\u989d"
    year_avg_key = "\u5e74\u65e5\u5747"
    month_avg_key = "\u6708\u65e5\u5747"
    source_key = "\u53e3\u5f84\u6765\u6e90"
    source_value = "\u5206\u90e8\u57fa\u7840\u6570\u636e\uff082026\uff09"

    segment_sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "segment_base_scale")
    assert segment_sheet["title"] == "\u5206\u90e8\u57fa\u7840\u89c4\u6a21"
    assert segment_sheet["columns"] == [
        indicator_key,
        spot_key,
        year_avg_key,
        month_avg_key,
        source_key,
    ]

    rows = {row[indicator_key]: row for row in segment_sheet["rows"]}
    expected_rows = {
        "\u516c\u53f8\u5b58\u6b3e\u5408\u8ba1": (2511.93, 2518.24, 2497.6),
        "\u50a8\u84c4\u5b58\u6b3e\u5408\u8ba1": (2604.03, 2548.79, 2574.13),
        "\u516c\u53f8\u8d37\u6b3e\u5408\u8ba1": (3458.61, 3339.26, 3391.97),
        "\u4e2a\u4eba\u8d37\u6b3e\u5408\u8ba1": (730.86, 732.14, 729.47),
        "\u4fe1\u7528\u5361": (71.36, 74.1, 72.09),
        "\u5b58\u6b3e\u5408\u8ba1": (5115.96, 5067.03, 5071.72),
    }
    for metric_name, (spot, year_avg, month_avg) in expected_rows.items():
        assert rows[metric_name] == {
            indicator_key: metric_name,
            spot_key: spot,
            year_avg_key: year_avg,
            month_avg_key: month_avg,
            source_key: source_value,
        }

    micro_loan = rows["\u5fae\u8d37\u4e2d\u5fc3"]
    assert micro_loan[indicator_key] == "\u5fae\u8d37\u4e2d\u5fc3"
    assert micro_loan[spot_key] is None
    assert micro_loan[year_avg_key] is None
    assert micro_loan[month_avg_key] is None
    assert str(micro_loan[source_key]).startswith("source_missing:")
    assert "80297" in str(micro_loan[source_key])


def test_real_202603_qdb_gl_workbook_includes_segment_scale_yoy_mom_compare_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()

    def merged(month_key: str):
        avg_path = _real_month_source(source_dir, "\u65e5\u5747", month_key)
        ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", month_key)
        return module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        )

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=merged("202603"),
        comparison_data={
            "prior_month": merged("202602"),
            "prior_year": merged("202503"),
        },
    )

    segment_sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "segment_scale_compare")
    assert segment_sheet["title"] == "\u5206\u90e8\u89c4\u6a21\u540c\u6bd4\u73af\u6bd4"
    assert segment_sheet["columns"] == [
        "\u6307\u6807",
        "\u53e3\u5f84",
        "\u672c\u671f",
        "\u5bf9\u6bd4\u671f",
        "\u589e\u51cf\u989d",
        "\u589e\u51cf\u5e45%",
        "\u53e3\u5f84\u6765\u6e90",
    ]
    assert len(segment_sheet["rows"]) == 28

    rows = {
        (row["\u6307\u6807"], row["\u53e3\u5f84"]): row
        for row in segment_sheet["rows"]
    }
    assert rows[("\u516c\u53f8\u5b58\u6b3e\u5408\u8ba1", "\u65f6\u70b9\u540c\u6bd4")] == {
        "\u6307\u6807": "\u516c\u53f8\u5b58\u6b3e\u5408\u8ba1",
        "\u53e3\u5f84": "\u65f6\u70b9\u540c\u6bd4",
        "\u672c\u671f": 2511.93,
        "\u5bf9\u6bd4\u671f": 2183.87,
        "\u589e\u51cf\u989d": 328.06,
        "\u589e\u51cf\u5e45%": 15.02,
        "\u53e3\u5f84\u6765\u6e90": "\u6708\u5ea6\u5206\u6790-\u5206\u90e8\u60c5\u51b5\uff1a\u603b\u8d26\u5bf9\u8d26+\u65e5\u5747\u540c\u6e90\u5386\u53f2\u6708\u91cd\u5efa",
    }
    assert rows[("\u516c\u53f8\u8d37\u6b3e\u5408\u8ba1", "\u65f6\u70b9\u73af\u6bd4")] == {
        "\u6307\u6807": "\u516c\u53f8\u8d37\u6b3e\u5408\u8ba1",
        "\u53e3\u5f84": "\u65f6\u70b9\u73af\u6bd4",
        "\u672c\u671f": 3458.61,
        "\u5bf9\u6bd4\u671f": 3378.6,
        "\u589e\u51cf\u989d": 80.01,
        "\u589e\u51cf\u5e45%": 2.37,
        "\u53e3\u5f84\u6765\u6e90": "\u6708\u5ea6\u5206\u6790-\u5206\u90e8\u60c5\u51b5\uff1a\u603b\u8d26\u5bf9\u8d26+\u65e5\u5747\u540c\u6e90\u5386\u53f2\u6708\u91cd\u5efa",
    }
    assert rows[("\u5b58\u6b3e\u5408\u8ba1", "\u5e74\u65e5\u5747\u540c\u6bd4")] == {
        "\u6307\u6807": "\u5b58\u6b3e\u5408\u8ba1",
        "\u53e3\u5f84": "\u5e74\u65e5\u5747\u540c\u6bd4",
        "\u672c\u671f": 5067.03,
        "\u5bf9\u6bd4\u671f": 4351.14,
        "\u589e\u51cf\u989d": 715.89,
        "\u589e\u51cf\u5e45%": 16.45,
        "\u53e3\u5f84\u6765\u6e90": "\u6708\u5ea6\u5206\u6790-\u5206\u90e8\u60c5\u51b5\uff1a\u603b\u8d26\u5bf9\u8d26+\u65e5\u5747\u540c\u6e90\u5386\u53f2\u6708\u91cd\u5efa",
    }

    micro_loan = rows[("\u5fae\u8d37\u4e2d\u5fc3", "\u6708\u65e5\u5747\u73af\u6bd4")]
    assert micro_loan["\u672c\u671f"] is None
    assert micro_loan["\u5bf9\u6bd4\u671f"] is None
    assert micro_loan["\u589e\u51cf\u989d"] is None
    assert micro_loan["\u589e\u51cf\u5e45%"] is None
    assert str(micro_loan["\u53e3\u5f84\u6765\u6e90"]).startswith("source_missing:")
    assert "80297" in str(micro_loan["\u53e3\u5f84\u6765\u6e90"])


def test_real_202603_qdb_gl_workbook_includes_company_scale_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()
    avg_path = _real_month_source(source_dir, "\u65e5\u5747", "202603")
    ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", "202603")

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "company_scale")
    assert sheet["title"] == "公司规模"
    assert sheet["columns"] == ["指标", "时点余额", "年日均", "月日均", "口径来源"]

    rows = {row["指标"]: row for row in sheet["rows"]}
    source_value = "公司规模：总账对账+日均同源科目重建"
    expected_rows = {
        "公司存款-活期": (923.15, 935.21, 920.29),
        "公司存款-定期": (1427.21, 1420.43, 1418.53),
        "公司存款-结构性": (161.57, 162.6, 158.78),
        "公司存款合计": (2511.93, 2518.24, 2497.6),
        "公司贷款-一般贷款": (3212.35, 3106.51, 3161.44),
        "公司贷款-票据": (246.26, 232.75, 230.53),
        "公司贷款合计": (3458.61, 3339.26, 3391.97),
    }
    for metric_name, (spot, year_avg, month_avg) in expected_rows.items():
        assert rows[metric_name] == {
            "指标": metric_name,
            "时点余额": spot,
            "年日均": year_avg,
            "月日均": month_avg,
            "口径来源": source_value,
        }


def test_real_202603_qdb_gl_workbook_includes_company_scale_compare_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()

    def merged(month_key: str):
        avg_path = _real_month_source(source_dir, "\u65e5\u5747", month_key)
        ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", month_key)
        return module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        )

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=merged("202603"),
        comparison_data={
            "prior_month": merged("202602"),
            "prior_year": merged("202503"),
        },
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "company_scale_compare")
    assert sheet["title"] == "公司规模同比环比"
    assert sheet["columns"] == ["指标", "口径", "本期", "对比期", "增减额", "增减幅%", "口径来源"]
    assert len(sheet["rows"]) == 28

    rows = {(row["指标"], row["口径"]): row for row in sheet["rows"]}
    source_value = "月度分析-公司板块：总账对账+日均同源历史月重建"
    assert rows[("公司存款-活期", "年日均同比")] == {
        "指标": "公司存款-活期",
        "口径": "年日均同比",
        "本期": 935.21,
        "对比期": 828.65,
        "增减额": 106.56,
        "增减幅%": 12.86,
        "口径来源": source_value,
    }
    assert rows[("公司贷款-票据", "月日均环比")] == {
        "指标": "公司贷款-票据",
        "口径": "月日均环比",
        "本期": 230.53,
        "对比期": 228.03,
        "增减额": 2.5,
        "增减幅%": 1.1,
        "口径来源": source_value,
    }
    assert rows[("公司贷款合计", "时点环比")] == {
        "指标": "公司贷款合计",
        "口径": "时点环比",
        "本期": 3458.61,
        "对比期": 3378.6,
        "增减额": 80.01,
        "增减幅%": 2.37,
        "口径来源": source_value,
    }


def test_real_202603_qdb_gl_workbook_includes_retail_scale_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()
    avg_path = _real_month_source(source_dir, "\u65e5\u5747", "202603")
    ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", "202603")

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "retail_scale")
    assert sheet["title"] == "零售规模"
    assert sheet["columns"] == ["指标", "时点余额", "年日均", "月日均", "口径来源"]

    rows = {row["指标"]: row for row in sheet["rows"]}
    source_value = "零售规模：总账对账+日均同源科目重建"
    expected_rows = {
        "零售存款-活期": (327.8, 318.83, 315.56),
        "零售存款-定期": (2211.59, 2165.89, 2195.43),
        "零售存款-结构性": (64.63, 64.07, 63.14),
        "零售存款合计": (2604.03, 2548.79, 2574.13),
        "参考：信用卡": (71.36, 74.1, 72.09),
        "参考：个人贷款合计": (730.86, 732.14, 729.47),
    }
    for metric_name, (spot, year_avg, month_avg) in expected_rows.items():
        assert rows[metric_name] == {
            "指标": metric_name,
            "时点余额": spot,
            "年日均": year_avg,
            "月日均": month_avg,
            "口径来源": source_value,
        }

    branch_loan = rows["零售贷款-分支行个贷"]
    assert branch_loan["时点余额"] is None
    assert branch_loan["年日均"] is None
    assert branch_loan["月日均"] is None
    assert str(branch_loan["口径来源"]).startswith("source_missing:")
    assert "80297" in str(branch_loan["口径来源"])
    micro_loan = rows["参考：微贷中心"]
    assert micro_loan["时点余额"] is None
    assert micro_loan["年日均"] is None
    assert micro_loan["月日均"] is None
    assert str(micro_loan["口径来源"]).startswith("source_missing:")
    assert "80297" in str(micro_loan["口径来源"])


def test_real_202603_qdb_gl_workbook_includes_retail_scale_compare_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()

    def merged(month_key: str):
        avg_path = _real_month_source(source_dir, "\u65e5\u5747", month_key)
        ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", month_key)
        return module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        )

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=merged("202603"),
        comparison_data={
            "prior_month": merged("202602"),
            "prior_year": merged("202503"),
        },
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "retail_scale_compare")
    assert sheet["title"] == "零售规模同比环比"
    assert sheet["columns"] == ["指标", "口径", "本期", "对比期", "增减额", "增减幅%", "口径来源"]
    assert len(sheet["rows"]) == 32

    rows = {(row["指标"], row["口径"]): row for row in sheet["rows"]}
    source_value = "月度分析-零售板块：总账对账+日均同源历史月重建"
    assert rows[("零售存款合计", "时点环比")] == {
        "指标": "零售存款合计",
        "口径": "时点环比",
        "本期": 2604.03,
        "对比期": 2566.14,
        "增减额": 37.88,
        "增减幅%": 1.48,
        "口径来源": source_value,
    }
    assert rows[("参考：信用卡", "月日均环比")] == {
        "指标": "参考：信用卡",
        "口径": "月日均环比",
        "本期": 72.09,
        "对比期": 74.47,
        "增减额": -2.38,
        "增减幅%": -3.19,
        "口径来源": source_value,
    }
    assert rows[("参考：个人贷款合计", "年日均同比")] == {
        "指标": "参考：个人贷款合计",
        "口径": "年日均同比",
        "本期": 732.14,
        "对比期": 791.26,
        "增减额": -59.12,
        "增减幅%": -7.47,
        "口径来源": source_value,
    }

    branch_loan = rows[("零售贷款-分支行个贷", "月日均环比")]
    assert branch_loan["本期"] is None
    assert branch_loan["对比期"] is None
    assert branch_loan["增减额"] is None
    assert branch_loan["增减幅%"] is None
    assert str(branch_loan["口径来源"]).startswith("source_missing:")
    assert "80297" in str(branch_loan["口径来源"])


def test_real_202603_qdb_gl_workbook_includes_financial_market_scale_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()
    avg_path = _real_month_source(source_dir, "\u65e5\u5747", "202603")
    ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", "202603")

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "financial_market_scale")
    assert sheet["title"] == "\u91d1\u878d\u5e02\u573a\u89c4\u6a21"
    assert sheet["columns"] == [
        "\u6307\u6807",
        "\u65f6\u70b9\u4f59\u989d",
        "\u5e74\u65e5\u5747",
        "\u6708\u65e5\u5747",
        "\u53e3\u5f84\u6765\u6e90",
    ]

    rows = {row["\u6307\u6807"]: row for row in sheet["rows"]}
    source_value = "\u91d1\u878d\u5e02\u573a\u89c4\u6a21\uff1a\u603b\u8d26\u5bf9\u8d26+\u65e5\u5747\u540c\u6e90\u79d1\u76ee\u91cd\u5efa"
    assert rows["\u751f\u606f\u503a\u5238\u6295\u8d44"] == {
        "\u6307\u6807": "\u751f\u606f\u503a\u5238\u6295\u8d44",
        "\u65f6\u70b9\u4f59\u989d": 2541.38,
        "\u5e74\u65e5\u5747": 2458.66,
        "\u6708\u65e5\u5747": 2522.26,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows["FVTPL"] == {
        "\u6307\u6807": "FVTPL",
        "\u65f6\u70b9\u4f59\u989d": 824.2,
        "\u5e74\u65e5\u5747": 835.62,
        "\u6708\u65e5\u5747": 832.53,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows["\u540c\u4e1a\u8d44\u4ea7"] == {
        "\u6307\u6807": "\u540c\u4e1a\u8d44\u4ea7",
        "\u65f6\u70b9\u4f59\u989d": 198.25,
        "\u5e74\u65e5\u5747": 418.89,
        "\u6708\u65e5\u5747": 389.95,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows["\u540c\u4e1a\u8d1f\u503a"] == {
        "\u6307\u6807": "\u540c\u4e1a\u8d1f\u503a",
        "\u65f6\u70b9\u4f59\u989d": 1613.37,
        "\u5e74\u65e5\u5747": 1702.73,
        "\u6708\u65e5\u5747": 1756.73,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }


def test_real_202603_qdb_gl_workbook_includes_financial_market_scale_compare_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()

    def merged(month_key: str):
        avg_path = _real_month_source(source_dir, "\u65e5\u5747", month_key)
        ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", month_key)
        return module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        )

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=merged("202603"),
        comparison_data={
            "prior_month": merged("202602"),
            "prior_year": merged("202503"),
        },
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "financial_market_scale_compare")
    assert sheet["title"] == "\u91d1\u878d\u5e02\u573a\u89c4\u6a21\u540c\u6bd4\u73af\u6bd4"
    assert sheet["columns"] == [
        "\u6307\u6807",
        "\u53e3\u5f84",
        "\u672c\u671f",
        "\u5bf9\u6bd4\u671f",
        "\u589e\u51cf\u989d",
        "\u589e\u51cf\u5e45%",
        "\u53e3\u5f84\u6765\u6e90",
    ]
    assert len(sheet["rows"]) == 16

    rows = {(row["\u6307\u6807"], row["\u53e3\u5f84"]): row for row in sheet["rows"]}
    source_value = "\u6708\u5ea6\u5206\u6790-\u91d1\u878d\u5e02\u573a\uff1a\u603b\u8d26\u5bf9\u8d26+\u65e5\u5747\u540c\u6e90\u5386\u53f2\u6708\u91cd\u5efa"
    assert rows[("\u751f\u606f\u503a\u5238\u6295\u8d44", "\u65f6\u70b9\u540c\u6bd4")] == {
        "\u6307\u6807": "\u751f\u606f\u503a\u5238\u6295\u8d44",
        "\u53e3\u5f84": "\u65f6\u70b9\u540c\u6bd4",
        "\u672c\u671f": 2541.38,
        "\u5bf9\u6bd4\u671f": 1940.85,
        "\u589e\u51cf\u989d": 600.52,
        "\u589e\u51cf\u5e45%": 30.94,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows[("FVTPL", "\u65f6\u70b9\u73af\u6bd4")] == {
        "\u6307\u6807": "FVTPL",
        "\u53e3\u5f84": "\u65f6\u70b9\u73af\u6bd4",
        "\u672c\u671f": 824.2,
        "\u5bf9\u6bd4\u671f": 875.48,
        "\u589e\u51cf\u989d": -51.28,
        "\u589e\u51cf\u5e45%": -5.86,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows[("\u540c\u4e1a\u8d44\u4ea7", "\u5e74\u65e5\u5747\u540c\u6bd4")] == {
        "\u6307\u6807": "\u540c\u4e1a\u8d44\u4ea7",
        "\u53e3\u5f84": "\u5e74\u65e5\u5747\u540c\u6bd4",
        "\u672c\u671f": 418.89,
        "\u5bf9\u6bd4\u671f": 238.11,
        "\u589e\u51cf\u989d": 180.78,
        "\u589e\u51cf\u5e45%": 75.93,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows[("\u540c\u4e1a\u8d1f\u503a", "\u6708\u65e5\u5747\u73af\u6bd4")] == {
        "\u6307\u6807": "\u540c\u4e1a\u8d1f\u503a",
        "\u53e3\u5f84": "\u6708\u65e5\u5747\u73af\u6bd4",
        "\u672c\u671f": 1756.73,
        "\u5bf9\u6bd4\u671f": 1710.07,
        "\u589e\u51cf\u989d": 46.66,
        "\u589e\u51cf\u5e45%": 2.73,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }


def test_real_202603_qdb_gl_workbook_includes_income_rate_analysis_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()
    avg_path = _real_month_source(source_dir, "\u65e5\u5747", "202603")
    ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", "202603")

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        ),
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "income_rate_analysis")
    assert sheet["title"] == "\u6536\u76ca\u7387\u5206\u6790\uff08\u603b\u8d26\u53ef\u590d\u7b97\uff09"
    assert sheet["columns"] == [
        "\u6307\u6807",
        "\u677f\u5757",
        "\u6536\u76ca\u7c7b\u522b",
        "\u5e74\u65e5\u5747\u89c4\u6a21",
        "\u603b\u8d26\u6536\u76ca/\u652f\u51fa",
        "\u5e74\u5316\u6536\u76ca\u7387/\u4ed8\u606f\u7387%",
        "\u53e3\u5f84\u6765\u6e90",
    ]

    rows = {row["\u6307\u6807"]: row for row in sheet["rows"]}
    source_value = "\u6536\u76ca\u7387\u5206\u6790\uff1a\u603b\u8d26\u6536\u76ca\u79d1\u76ee+\u65e5\u5747\u89c4\u6a21\u91cd\u5efa"
    assert rows["\u516c\u53f8\u8d37\u6b3e\u5229\u606f\u6536\u5165"] == {
        "\u6307\u6807": "\u516c\u53f8\u8d37\u6b3e\u5229\u606f\u6536\u5165",
        "\u677f\u5757": "\u516c\u53f8\u677f\u5757",
        "\u6536\u76ca\u7c7b\u522b": "\u8d37\u6b3e\u5229\u606f\u6536\u5165",
        "\u5e74\u65e5\u5747\u89c4\u6a21": 3339.26,
        "\u603b\u8d26\u6536\u76ca/\u652f\u51fa": 32.4,
        "\u5e74\u5316\u6536\u76ca\u7387/\u4ed8\u606f\u7387%": 3.93,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    personal_loan_income = rows["\u4e2a\u4eba\u8d37\u6b3e\u5229\u606f\u6536\u5165"]
    assert personal_loan_income["\u5e74\u65e5\u5747\u89c4\u6a21"] is None
    assert personal_loan_income["\u603b\u8d26\u6536\u76ca/\u652f\u51fa"] == 6.53
    assert personal_loan_income["\u5e74\u5316\u6536\u76ca\u7387/\u4ed8\u606f\u7387%"] is None
    assert str(personal_loan_income["\u53e3\u5f84\u6765\u6e90"]).startswith("source_missing:")
    assert "\u4fe1\u7528\u5361\u751f\u606f\u89c4\u6a21" in str(personal_loan_income["\u53e3\u5f84\u6765\u6e90"])
    assert rows["\u516c\u53f8\u5b58\u6b3e\u5229\u606f\u652f\u51fa"] == {
        "\u6307\u6807": "\u516c\u53f8\u5b58\u6b3e\u5229\u606f\u652f\u51fa",
        "\u677f\u5757": "\u516c\u53f8\u677f\u5757",
        "\u6536\u76ca\u7c7b\u522b": "\u5b58\u6b3e\u5229\u606f\u652f\u51fa",
        "\u5e74\u65e5\u5747\u89c4\u6a21": 2518.24,
        "\u603b\u8d26\u6536\u76ca/\u652f\u51fa": 7.67,
        "\u5e74\u5316\u6536\u76ca\u7387/\u4ed8\u606f\u7387%": 1.24,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows["\u91d1\u878d\u6295\u8d44\u5229\u606f\u6536\u5165"]["\u603b\u8d26\u6536\u76ca/\u652f\u51fa"] is None
    assert str(rows["\u91d1\u878d\u6295\u8d44\u5229\u606f\u6536\u5165"]["\u53e3\u5f84\u6765\u6e90"]).startswith("source_missing:")


def test_real_202603_qdb_gl_workbook_includes_income_rate_attribution_sheet():
    module = load_module(
        "backend.app.core_finance.qdb_gl_monthly_analysis",
        "backend/app/core_finance/qdb_gl_monthly_analysis.py",
    )
    source_dir = _real_qdb_gl_source_dir()

    def merged(month_key: str):
        avg_path = _real_month_source(source_dir, "\u65e5\u5747", month_key)
        ledger_path = _real_month_source(source_dir, "\u603b\u8d26\u5bf9\u8d26", month_key)
        return module.merge_all(
            module.parse_general_ledger(ledger_path),
            module.parse_daily_avg(avg_path),
        )

    workbook = module.build_qdb_gl_monthly_analysis_workbook(
        report_month="202603",
        merged_data=merged("202603"),
        comparison_data={
            "prior_year": merged("202503"),
            "prior_month": merged("202602"),
        },
    )

    sheet = next(sheet for sheet in workbook["sheets"] if sheet["key"] == "income_rate_attribution")
    assert sheet["title"] == "\u6536\u76ca\u91cf\u4ef7\u5f52\u56e0\uff08\u5e74\u7d2f\u8ba1\u540c\u6bd4\uff09"
    assert sheet["columns"] == [
        "\u6307\u6807",
        "\u677f\u5757",
        "\u672c\u671f\u6536\u76ca/\u652f\u51fa",
        "\u5bf9\u6bd4\u671f\u6536\u76ca/\u652f\u51fa",
        "\u589e\u51cf\u989d",
        "\u89c4\u6a21\u8d21\u732e",
        "\u5229\u7387\u8d21\u732e",
        "\u6821\u9a8c\u5dee\u5f02",
        "\u53e3\u5f84\u6765\u6e90",
    ]

    rows = {row["\u6307\u6807"]: row for row in sheet["rows"]}
    source_value = "\u6536\u76ca\u91cf\u4ef7\u5f52\u56e0\uff1a\u603b\u8d26\u6536\u76ca\u79d1\u76ee+\u65e5\u5747\u89c4\u6a21\u6309\u5e74\u7d2f\u8ba1\u540c\u6bd4\u62c6\u89e3"
    assert rows["\u516c\u53f8\u8d37\u6b3e\u5229\u606f\u6536\u5165"] == {
        "\u6307\u6807": "\u516c\u53f8\u8d37\u6b3e\u5229\u606f\u6536\u5165",
        "\u677f\u5757": "\u516c\u53f8\u677f\u5757",
        "\u672c\u671f\u6536\u76ca/\u652f\u51fa": 32.4,
        "\u5bf9\u6bd4\u671f\u6536\u76ca/\u652f\u51fa": 29.15,
        "\u589e\u51cf\u989d": 3.24,
        "\u89c4\u6a21\u8d21\u732e": 6.12,
        "\u5229\u7387\u8d21\u732e": -2.88,
        "\u6821\u9a8c\u5dee\u5f02": 0,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    assert rows["\u50a8\u84c4\u5b58\u6b3e\u5229\u606f\u652f\u51fa"] == {
        "\u6307\u6807": "\u50a8\u84c4\u5b58\u6b3e\u5229\u606f\u652f\u51fa",
        "\u677f\u5757": "\u96f6\u552e\u5b58\u6b3e",
        "\u672c\u671f\u6536\u76ca/\u652f\u51fa": 11.13,
        "\u5bf9\u6bd4\u671f\u6536\u76ca/\u652f\u51fa": 12.36,
        "\u589e\u51cf\u989d": -1.23,
        "\u89c4\u6a21\u8d21\u732e": 1.43,
        "\u5229\u7387\u8d21\u732e": -2.65,
        "\u6821\u9a8c\u5dee\u5f02": 0,
        "\u53e3\u5f84\u6765\u6e90": source_value,
    }
    personal_loan_income = rows["\u4e2a\u4eba\u8d37\u6b3e\u5229\u606f\u6536\u5165"]
    assert personal_loan_income["\u672c\u671f\u6536\u76ca/\u652f\u51fa"] == 6.53
    assert personal_loan_income["\u5bf9\u6bd4\u671f\u6536\u76ca/\u652f\u51fa"] == 7.68
    assert personal_loan_income["\u589e\u51cf\u989d"] == -1.15
    assert personal_loan_income["\u89c4\u6a21\u8d21\u732e"] is None
    assert personal_loan_income["\u5229\u7387\u8d21\u732e"] is None
    assert personal_loan_income["\u6821\u9a8c\u5dee\u5f02"] is None
    assert str(personal_loan_income["\u53e3\u5f84\u6765\u6e90"]).startswith("source_missing:")
    assert "\u4fe1\u7528\u5361\u751f\u606f\u89c4\u6a21" in str(personal_loan_income["\u53e3\u5f84\u6765\u6e90"])
    assert rows["\u91d1\u878d\u6295\u8d44\u5229\u606f\u6536\u5165"]["\u589e\u51cf\u989d"] is None
    assert str(rows["\u91d1\u878d\u6295\u8d44\u5229\u606f\u6536\u5165"]["\u53e3\u5f84\u6765\u6e90"]).startswith("source_missing:")


def _real_qdb_gl_source_dir() -> Path:
    ledger_token = "\u603b\u8d26\u5bf9\u8d26"
    average_token = "\u65e5\u5747"
    return next(
        path
        for path in (ROOT / "data_input").iterdir()
        if path.is_dir() and ledger_token in path.name and average_token in path.name
    )


def _real_month_source(source_dir: Path, token: str, month_key: str) -> Path:
    return next(path for path in source_dir.glob(f"*{month_key}.xlsx") if token in path.name)


def _write_month_pair(target_dir: Path, month_key: str) -> tuple[Path, Path]:
    avg_path = target_dir / f"日均{month_key}.xlsx"
    ledger_path = target_dir / f"总账对账{month_key}.xlsx"
    _write_average_workbook(avg_path)
    _write_ledger_workbook(ledger_path)
    return avg_path, ledger_path


def _write_average_workbook(path: Path) -> None:
    workbook = Workbook()
    year_sheet = workbook.active
    year_sheet.title = "年"
    month_sheet = workbook.create_sheet(title="月")

    for worksheet in (year_sheet, month_sheet):
        worksheet.append(["机构：199200青岛银行"])
        worksheet.append(["日期：2026-02-01 至 2026-02-28"])
        worksheet.append([None] * 31)

    month_rows = [
        _avg_row(cnx_3="101", cnx_3_val=18 * B, cny_3="101", cny_3_val=18 * B),
        _avg_row(cnx_3="110", cnx_3_val=28 * B, cny_3="110", cny_3_val=28 * B),
        _avg_row(cnx_3="114", cnx_3_val=20 * B, cny_3="114", cny_3_val=20 * B),
        _avg_row(cnx_3="123", cnx_3_val=900 * B, cnx_5="12301", cnx_5_val=1100 * B, cnx_11="12301000001", cnx_11_val=900 * B,
                 cny_3="123", cny_3_val=900 * B, cny_5="12301", cny_5_val=1100 * B, cny_11="12301000001", cny_11_val=900 * B),
        _avg_row(cnx_3="130", cnx_3_val=8 * B, cnx_11="13001000001", cnx_11_val=8 * B, cny_3="130", cny_3_val=8 * B, cny_11="13001000001", cny_11_val=8 * B),
        _avg_row(cnx_3="131", cnx_3_val=-28 * B, cnx_11="13101000001", cnx_11_val=-28 * B, cny_3="131", cny_3_val=-28 * B, cny_11="13101000001", cny_11_val=-28 * B),
        _avg_row(cnx_3="140", cnx_3_val=230 * B, cnx_11="14001000001", cnx_11_val=230 * B, cny_3="140", cny_3_val=230 * B, cny_11="14001000001", cny_11_val=230 * B),
        _avg_row(cnx_3="142", cnx_3_val=280 * B, cnx_11="14201000001", cnx_11_val=280 * B, cny_3="142", cny_3_val=260 * B, cny_11="14201000001", cny_11_val=260 * B),
        _avg_row(cnx_3="144", cnx_3_val=170 * B, cnx_11="14401000001", cnx_11_val=170 * B, cny_3="144", cny_3_val=150 * B, cny_11="14401000001", cny_11_val=150 * B),
        _avg_row(cnx_3="201", cnx_3_val=-650 * B, cnx_5="20101", cnx_5_val=-650 * B, cnx_11="20101000001", cnx_11_val=-650 * B,
                 cny_3="201", cny_3_val=-650 * B, cny_5="20101", cny_5_val=-650 * B, cny_11="20101000001", cny_11_val=-650 * B),
        _avg_row(cnx_3="205", cnx_3_val=-850 * B, cnx_5="20501", cnx_5_val=-850 * B, cnx_11="20501000001", cnx_11_val=-850 * B,
                 cny_3="205", cny_3_val=-850 * B, cny_5="20501", cny_5_val=-850 * B, cny_11="20501000001", cny_11_val=-850 * B),
        _avg_row(cnx_3="234", cnx_3_val=-269 * B, cnx_11="23401000001", cnx_11_val=-269 * B, cny_3="234", cny_3_val=-269 * B, cny_11="23401000001", cny_11_val=-269 * B),
        _avg_row(cnx_3="255", cnx_3_val=-197 * B, cnx_11="25501000001", cnx_11_val=-197 * B, cny_3="255", cny_3_val=-197 * B, cny_11="25501000001", cny_11_val=-197 * B),
    ]
    year_rows = [
        _avg_row(cnx_3="101", cnx_3_val=17 * B, cny_3="101", cny_3_val=17 * B),
        _avg_row(cnx_3="110", cnx_3_val=26 * B, cny_3="110", cny_3_val=26 * B),
        _avg_row(cnx_3="114", cnx_3_val=18 * B, cny_3="114", cny_3_val=18 * B),
        _avg_row(cnx_3="123", cnx_3_val=850 * B, cnx_5="12301", cnx_5_val=900 * B, cnx_11="12301000001", cnx_11_val=850 * B,
                 cny_3="123", cny_3_val=850 * B, cny_5="12301", cny_5_val=900 * B, cny_11="12301000001", cny_11_val=850 * B),
        _avg_row(cnx_3="130", cnx_3_val=5 * B, cnx_11="13001000001", cnx_11_val=5 * B, cny_3="130", cny_3_val=5 * B, cny_11="13001000001", cny_11_val=5 * B),
        _avg_row(cnx_3="131", cnx_3_val=-25 * B, cnx_11="13101000001", cnx_11_val=-25 * B, cny_3="131", cny_3_val=-25 * B, cny_11="13101000001", cny_11_val=-25 * B),
        _avg_row(cnx_3="140", cnx_3_val=220 * B, cnx_11="14001000001", cnx_11_val=220 * B, cny_3="140", cny_3_val=220 * B, cny_11="14001000001", cny_11_val=220 * B),
        _avg_row(cnx_3="142", cnx_3_val=260 * B, cnx_11="14201000001", cnx_11_val=260 * B, cny_3="142", cny_3_val=240 * B, cny_11="14201000001", cny_11_val=240 * B),
        _avg_row(cnx_3="144", cnx_3_val=160 * B, cnx_11="14401000001", cnx_11_val=160 * B, cny_3="144", cny_3_val=140 * B, cny_11="14401000001", cny_11_val=140 * B),
        _avg_row(cnx_3="201", cnx_3_val=-620 * B, cnx_5="20101", cnx_5_val=-600 * B, cnx_11="20101000001", cnx_11_val=-620 * B,
                 cny_3="201", cny_3_val=-620 * B, cny_5="20101", cny_5_val=-600 * B, cny_11="20101000001", cny_11_val=-620 * B),
        _avg_row(cnx_3="205", cnx_3_val=-830 * B, cnx_5="20501", cnx_5_val=-800 * B, cnx_11="20501000001", cnx_11_val=-830 * B,
                 cny_3="205", cny_3_val=-830 * B, cny_5="20501", cny_5_val=-800 * B, cny_11="20501000001", cny_11_val=-830 * B),
        _avg_row(cnx_3="234", cnx_3_val=-250 * B, cnx_11="23401000001", cnx_11_val=-250 * B, cny_3="234", cny_3_val=-250 * B, cny_11="23401000001", cny_11_val=-250 * B),
        _avg_row(cnx_3="255", cnx_3_val=-180 * B, cnx_11="25501000001", cnx_11_val=-180 * B, cny_3="255", cny_3_val=-180 * B, cny_11="25501000001", cny_11_val=-180 * B),
    ]

    for row in year_rows:
        year_sheet.append(row)
    for row in month_rows:
        month_sheet.append(row)

    workbook.save(path)


def _avg_row(
    *,
    cnx_3=None,
    cnx_3_val=None,
    cnx_5=None,
    cnx_5_val=None,
    cnx_11=None,
    cnx_11_val=None,
    cny_3=None,
    cny_3_val=None,
    cny_5=None,
    cny_5_val=None,
    cny_11=None,
    cny_11_val=None,
):
    row = [None] * 31
    if cnx_3 is not None:
        row[0] = "CNX"
        row[1] = float(cnx_3)
        row[2] = cnx_3_val
    if cnx_5 is not None:
        row[4] = "CNX"
        row[5] = float(cnx_5)
        row[6] = cnx_5_val
    if cnx_11 is not None:
        row[12] = "CNX"
        row[13] = float(cnx_11)
        row[14] = cnx_11_val
    if cny_3 is not None:
        row[16] = "CNY"
        row[17] = float(cny_3)
        row[18] = cny_3_val
    if cny_5 is not None:
        row[20] = "CNY"
        row[21] = float(cny_5)
        row[22] = cny_5_val
    if cny_11 is not None:
        row[28] = "CNY"
        row[29] = float(cny_11)
        row[30] = cny_11_val
    return row


def _write_ledger_workbook(path: Path) -> None:
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    for sheet_name, currency in (("综本", "CNX"), ("人民币", "CNY")):
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(["总账对账"])
        worksheet.append([None])
        worksheet.append(["分类账来源"])
        worksheet.append(["公司：青岛银行"])
        worksheet.append(["会计期间：2026-02-01--2026-02-28"])
        worksheet.append(["组合科目代码", "组合科目名称", "币种", "期初余额", "本期借方", "本期贷方", "期末余额"])
        for row in _ledger_rows(currency):
            worksheet.append(row)

    workbook.save(path)


def _ledger_rows(currency: str):
    return [
        _balanced_ledger_row("10101000001", "现金", currency, 18 * B, 20 * B),
        _balanced_ledger_row("11001000001", "存放央行", currency, 28 * B, 30 * B),
        _balanced_ledger_row("11401000001", "存放同业", currency, 20 * B, 25 * B),
        _balanced_ledger_row("12301000001", "公司贷款-农林牧渔", currency, 900 * B, 1000 * B),
        _balanced_ledger_row("13001000001", "逾期贷款", currency, 50 * B, 86 * B),
        _balanced_ledger_row("13101000001", "贷款减值准备", currency, -28 * B, -30 * B),
        _balanced_ledger_row("14001000001", "买入返售", currency, 230 * B, 58 * B),
        _balanced_ledger_row("14201000001", "AC债券投资", currency, (280 if currency == "CNX" else 240) * B, (310 if currency == "CNX" else 260) * B),
        _balanced_ledger_row("14401000001", "FVOCI金融资产", currency, (170 if currency == "CNX" else 140) * B, (180 if currency == "CNX" else 150) * B),
        _balanced_ledger_row("20101000001", "单位活期存款-农林牧渔", currency, -650 * B, -700 * B),
        _balanced_ledger_row("20501000001", "单位定期存款-农林牧渔", currency, -850 * B, -900 * B),
        _balanced_ledger_row("23401000001", "同业存放", currency, -269 * B, -359 * B),
        _balanced_ledger_row("25501000001", "卖出回购", currency, -197 * B, -60 * B),
    ]


def _balanced_ledger_row(
    account_code: str,
    account_name: str,
    currency: str,
    opening_balance: float,
    ending_balance: float,
):
    delta = ending_balance - opening_balance
    debit = delta if delta >= 0 else 0
    credit = abs(delta) if delta < 0 else 0
    return (account_code, account_name, currency, opening_balance, debit, credit, ending_balance)
