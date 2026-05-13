from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.governance.settings import get_settings
from tests.helpers import load_module
from tests.test_balance_analysis_api import _configure_and_materialize

REPORT_DATE = "2025-12-31"


def _build_client() -> TestClient:
    return TestClient(load_module("backend.app.main", "backend/app/main.py").app)


def test_balance_analysis_workbook_response_is_gzipped_when_client_accepts_it(tmp_path, monkeypatch):
    _configure_and_materialize(tmp_path, monkeypatch)
    client = _build_client()

    response = client.get(
        "/ui/balance-analysis/workbook",
        params={
            "report_date": REPORT_DATE,
            "position_scope": "all",
            "currency_basis": "CNY",
        },
        headers={"Accept-Encoding": "gzip"},
    )

    assert response.status_code == 200
    assert response.headers["content-encoding"] == "gzip"
    assert len(response.content) > 1024
    assert response.json()["result_meta"]["result_kind"] == "balance-analysis.workbook"

    get_settings.cache_clear()


def test_health_live_response_stays_uncompressed_when_small(tmp_path, monkeypatch):
    _configure_and_materialize(tmp_path, monkeypatch)
    client = _build_client()

    response = client.get(
        "/health/live",
        headers={"Accept-Encoding": "gzip"},
    )

    assert response.status_code == 200
    assert "content-encoding" not in response.headers
    assert response.json() == {"status": "ok"}

    get_settings.cache_clear()


def test_balance_analysis_workbook_export_remains_readable_with_gzip_enabled(tmp_path, monkeypatch):
    _configure_and_materialize(tmp_path, monkeypatch)
    client = _build_client()

    response = client.get(
        "/ui/balance-analysis/workbook/export",
        params={
            "report_date": REPORT_DATE,
            "position_scope": "all",
            "currency_basis": "CNY",
        },
        headers={"Accept-Encoding": "gzip"},
    )

    assert response.status_code == 200
    assert response.headers["content-encoding"] == "gzip"
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "balance-analysis-workbook-2025-12-31.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames

    get_settings.cache_clear()
