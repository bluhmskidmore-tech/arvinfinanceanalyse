from __future__ import annotations

from io import BytesIO
from urllib.parse import quote

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.governance.settings import get_settings
from tests.helpers import load_module
from tests.test_qdb_gl_monthly_analysis_core import _write_month_pair


def test_export_returns_valid_xlsx_with_required_sheets(tmp_path, monkeypatch):
    source_dir = tmp_path / "data_input" / "pnl_总账对账-日均"
    source_dir.mkdir(parents=True)
    _write_month_pair(source_dir, "202602")

    monkeypatch.setenv("MOSS_PRODUCT_CATEGORY_SOURCE_DIR", str(source_dir))
    get_settings.cache_clear()

    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)
    response = client.get(
        "/ui/qdb-gl-monthly-analysis/workbook/export",
        params={"report_month": "202602"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        "attachment; "
        "filename=qdb-gl-monthly-analysis-202602.xlsx; "
        f"filename*=UTF-8''{quote('analysis_report_202602.xlsx')}"
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [
        "经营概览",
        "3位科目总览",
        "资产结构",
        "负债结构",
        "贷款行业",
        "存款行业_活期",
        "存款行业_定期",
        "行业存贷差",
        "11位偏离TOP",
        "异动预警",
        "外币分析",
    ]

    overview = workbook["经营概览"]
    assert overview["A1"].value == "指标"
    assert overview["B1"].value == "值"
    assert overview["A2"].value == "总资产(亿)"

    get_settings.cache_clear()
