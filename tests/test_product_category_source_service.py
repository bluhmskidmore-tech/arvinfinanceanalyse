from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from backend.app.services.product_category_source_service import (
    _parse_average_workbook,
    _parse_ledger_workbook,
)


def test_product_category_average_workbook_with_single_sheet_is_treated_as_partial_input(tmp_path: Path):
    avg_path = tmp_path / "日均202401.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "年"
    sheet["A4"] = "CNX"
    sheet["B4"] = "13304010001"
    sheet["C4"] = 123
    workbook.save(avg_path)

    annual_rows, monthly_rows = _parse_average_workbook(avg_path)

    assert annual_rows[("13304010001", "CNX")] == Decimal("123")
    assert monthly_rows == {}


def test_product_category_average_workbook_ignores_misaligned_non_currency_blocks(tmp_path: Path):
    avg_path = tmp_path / "鏃ュ潎202401.xlsx"
    workbook = Workbook()
    year_sheet = workbook.active
    year_sheet.title = "annual"
    month_sheet = workbook.create_sheet(title="monthly")

    for sheet in (year_sheet, month_sheet):
        row = [None] * 20
        row[12] = "CNX"
        row[13] = "23401000001"
        row[14] = 100
        row[16] = "23402000001"
        row[17] = -20
        row[18] = 20
        sheet.append(["header"])
        sheet.append(["header"])
        sheet.append(["header"])
        sheet.append(row)

    workbook.save(avg_path)

    annual_rows, monthly_rows = _parse_average_workbook(avg_path)

    assert annual_rows[("23401000001", "CNX")] == Decimal("100")
    assert monthly_rows[("23401000001", "CNX")] == Decimal("100")
    assert all(currency in {"CNX", "CNY"} for _code, currency in annual_rows)
    assert all(currency in {"CNX", "CNY"} for _code, currency in monthly_rows)
    assert ("-20", "23402000001") not in annual_rows
    assert ("-20", "23402000001") not in monthly_rows


def test_product_category_ledger_workbook_accepts_optional_leading_prefix_column(tmp_path: Path):
    ledger_path = tmp_path / "总账对账202401.xlsx"
    workbook = Workbook()
    cnx = workbook.active
    cnx.title = "综本"
    cny = workbook.create_sheet(title="人民币")

    for worksheet in (cnx, cny):
        worksheet.append(["header"])
        worksheet.append([None])
        worksheet.append(["header"])
        worksheet.append(["header"])
        worksheet.append(["header"])

    cnx.append(["组合科目代码", "组合科目名称", "币种", "期初余额", "本期借方", "本期贷方", "期末余额"])
    cnx.append([10101000001, "业务库存现金", "CNX", 10, 20, 5, 25])

    cny.append(["seicd", "组合科目代码", "组合科目名称", "币种", "期初余额", "本期借方", "本期贷方", "期末余额"])
    cny.append(["101", 10101000002, "ATM库存现金", "CNY", 30, 50, 15, 65])
    workbook.save(ledger_path)

    rows = _parse_ledger_workbook(ledger_path)

    assert rows[("10101000001", "CNX")]["beginning_balance"] == Decimal("10")
    assert rows[("10101000001", "CNX")]["ending_balance"] == Decimal("25")
    assert rows[("10101000001", "CNX")]["monthly_pnl"] == Decimal("-15")
    assert rows[("10101000002", "CNY")]["beginning_balance"] == Decimal("30")
    assert rows[("10101000002", "CNY")]["ending_balance"] == Decimal("65")
    assert rows[("10101000002", "CNY")]["monthly_pnl"] == Decimal("-35")
