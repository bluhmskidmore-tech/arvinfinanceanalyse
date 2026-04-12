from __future__ import annotations

from io import BytesIO
from urllib.parse import quote

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from backend.app.governance.settings import get_settings
from tests.helpers import load_module
from tests.test_balance_analysis_api import _configure_and_materialize


REPORT_DATE = "2025-12-31"
EXPECTED_SHEETS = ["概览", "债券持仓", "同业持仓", "期限分布", "利率分布"]
TABLE_KEY_CANDIDATES = {
    "债券持仓": ("zqtz_balance", "bond_business_types"),
    "同业持仓": ("tyw_balance", "counterparty_types"),
    "期限分布": ("maturity_distribution", "maturity_gap"),
    "利率分布": ("rate_distribution",),
}


def _load_balance_service():
    return load_module(
        "backend.app.services.balance_analysis_service",
        "backend/app/services/balance_analysis_service.py",
    )


def _load_workbook_payload(*, duckdb_path: str, governance_dir: str) -> dict[str, object]:
    service_mod = _load_balance_service()
    return service_mod.balance_analysis_workbook_envelope(
        duckdb_path=duckdb_path,
        governance_dir=governance_dir,
        report_date=REPORT_DATE,
        position_scope="all",
        currency_basis="CNY",
    )["result"]


def _export_workbook_bytes(*, duckdb_path: str, governance_dir: str) -> tuple[str, bytes]:
    service_mod = _load_balance_service()
    return service_mod.export_balance_analysis_workbook_xlsx(
        duckdb_path=duckdb_path,
        governance_dir=governance_dir,
        report_date=REPORT_DATE,
        position_scope="all",
        currency_basis="CNY",
    )


def _pick_table(payload: dict[str, object], *keys: str) -> dict[str, object]:
    table_map = {table["key"]: table for table in payload["tables"]}
    for key in keys:
        if key in table_map:
            return table_map[key]
    raise AssertionError(f"Missing workbook table, expected one of: {keys!r}")


def test_export_returns_valid_xlsx_bytes(tmp_path, monkeypatch):
    _configure_and_materialize(tmp_path, monkeypatch)
    client = TestClient(load_module("backend.app.main", "backend/app/main.py").app)

    response = client.get(
        "/ui/balance-analysis/workbook/export",
        params={
            "report_date": REPORT_DATE,
            "position_scope": "all",
            "currency_basis": "CNY",
        },
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"] == (
        "attachment; "
        f"filename=balance-analysis-workbook-{REPORT_DATE}.xlsx; "
        f"filename*=UTF-8''{quote(f'资产负债分析_{REPORT_DATE}.xlsx')}"
    )

    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == EXPECTED_SHEETS

    get_settings.cache_clear()


def test_export_has_expected_sheets(tmp_path, monkeypatch):
    duckdb_path, governance_dir, _task_mod = _configure_and_materialize(tmp_path, monkeypatch)

    filename, content = _export_workbook_bytes(
        duckdb_path=str(duckdb_path),
        governance_dir=str(governance_dir),
    )

    assert filename == f"资产负债分析_{REPORT_DATE}.xlsx"

    workbook = load_workbook(BytesIO(content))
    assert workbook.sheetnames == EXPECTED_SHEETS

    get_settings.cache_clear()


def test_export_header_row_matches_columns(tmp_path, monkeypatch):
    duckdb_path, governance_dir, _task_mod = _configure_and_materialize(tmp_path, monkeypatch)
    payload = _load_workbook_payload(
        duckdb_path=str(duckdb_path),
        governance_dir=str(governance_dir),
    )
    _filename, content = _export_workbook_bytes(
        duckdb_path=str(duckdb_path),
        governance_dir=str(governance_dir),
    )
    workbook = load_workbook(BytesIO(content))

    for sheet_name, key_candidates in TABLE_KEY_CANDIDATES.items():
        table = _pick_table(payload, *key_candidates)
        expected_headers = [column["label"] for column in table["columns"]]
        actual_headers = [
            workbook[sheet_name].cell(row=1, column=index).value
            for index in range(1, len(expected_headers) + 1)
        ]
        assert actual_headers == expected_headers

    get_settings.cache_clear()
