from __future__ import annotations

import importlib
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

from backend.app.core_finance.balance_analysis import (
    FormalTywBalanceFactRow,
    FormalZqtzBalanceFactRow,
)
from backend.app.repositories.balance_analysis_repo import BalanceAnalysisRepository

EXPORT_WORKBOOK_TABLES = (
    ("\u503a\u5238\u6301\u4ed3", ("zqtz_balance", "bond_business_types")),
    ("\u540c\u4e1a\u6301\u4ed3", ("tyw_balance", "counterparty_types")),
    ("\u671f\u9650\u5206\u5e03", ("maturity_distribution", "maturity_gap")),
    ("\u5229\u7387\u5206\u5e03", ("rate_distribution",)),
)
EXCEL_HEADER_FONT = Font(bold=True)
EXCEL_RIGHT_ALIGNMENT = Alignment(horizontal="right")
EXCEL_AMOUNT_FORMAT = "#,##0.00000000"
EXCEL_NUMBER_FORMAT = "0.00000000"
EXCEL_INTEGER_FORMAT = "0"


def _build_balance_workbook_payload(
    *,
    duckdb_path: str,
    governance_dir: str,
    report_date: str,
    position_scope: Literal["asset", "liability", "all"],
    currency_basis: Literal["native", "CNY"],
    cache_key: str,
    job_name: str,
    resolve_completed_formal_build_lineage_fn: Any,
    repo_cls: type[Any] = BalanceAnalysisRepository,
    import_module_fn: Any = importlib.import_module,
    reload_module_fn: Any = importlib.reload,
    workbook_module_name: str = "backend.app.core_finance.balance_analysis_workbook",
) -> tuple[dict[str, Any], dict[str, object] | None]:
    repo = repo_cls(duckdb_path)
    if report_date not in repo.list_report_dates():
        raise ValueError(f"No balance-analysis data found for report_date={report_date}.")

    zqtz_native_rows = [
        _to_formal_zqtz_fact_row(row)
        for row in repo.fetch_formal_zqtz_rows(
            report_date=report_date,
            position_scope=position_scope,
            currency_basis="native",
        )
    ]
    tyw_native_rows = [
        _to_formal_tyw_fact_row(row)
        for row in repo.fetch_formal_tyw_rows(
            report_date=report_date,
            position_scope=position_scope,
            currency_basis="native",
        )
    ]
    zqtz_currency_rows = [
        _to_formal_zqtz_fact_row(row)
        for row in repo.fetch_formal_zqtz_rows(
            report_date=report_date,
            position_scope=position_scope,
            currency_basis="CNY",
        )
    ]
    workbook_mod = import_module_fn(workbook_module_name)
    workbook_mod = reload_module_fn(workbook_mod)
    workbook = workbook_mod.build_balance_analysis_workbook_payload(
        report_date=zqtz_native_rows[0].report_date if zqtz_native_rows else tyw_native_rows[0].report_date,
        position_scope=position_scope,
        currency_basis=currency_basis,
        zqtz_rows=zqtz_native_rows,
        tyw_rows=tyw_native_rows,
        zqtz_currency_rows=zqtz_currency_rows,
    )
    return workbook, resolve_completed_formal_build_lineage_fn(
        governance_dir=governance_dir,
        cache_key=cache_key,
        job_name=job_name,
        report_date=report_date,
    )


def _extract_generated_decision_section(workbook: dict[str, Any]) -> dict[str, Any]:
    for section in workbook.get("tables", []):
        if str(section.get("section_kind")) == "decision_items":
            return dict(section)
    return {"columns": [], "rows": []}


def _to_formal_zqtz_fact_row(row: dict[str, object]) -> FormalZqtzBalanceFactRow:
    maturity_date = str(row.get("maturity_date") or "").strip()
    return FormalZqtzBalanceFactRow(
        report_date=_parse_date(str(row["report_date"])),
        instrument_code=str(row["instrument_code"]),
        instrument_name=str(row.get("instrument_name") or ""),
        portfolio_name=str(row.get("portfolio_name") or ""),
        cost_center=str(row.get("cost_center") or ""),
        account_category=str(row.get("account_category") or ""),
        asset_class=str(row.get("asset_class") or ""),
        bond_type=str(row.get("bond_type") or ""),
        issuer_name=str(row.get("issuer_name") or ""),
        industry_name=str(row.get("industry_name") or ""),
        rating=str(row.get("rating") or ""),
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        position_scope=str(row["position_scope"]),
        currency_basis=str(row["currency_basis"]),
        currency_code=str(row.get("currency_code") or ""),
        face_value_amount=_as_decimal(row["face_value_amount"]),
        market_value_amount=_as_decimal(row["market_value_amount"]),
        amortized_cost_amount=_as_decimal(row["amortized_cost_amount"]),
        accrued_interest_amount=_as_decimal(row["accrued_interest_amount"]),
        coupon_rate=_optional_decimal(row.get("coupon_rate")),
        ytm_value=_optional_decimal(row.get("ytm_value")),
        maturity_date=_parse_date(maturity_date) if maturity_date else None,
        interest_mode=str(row.get("interest_mode") or ""),
        is_issuance_like=bool(row["is_issuance_like"]),
        overdue_principal_days=_optional_nonnegative_int(row.get("overdue_principal_days")),
        overdue_interest_days=_optional_nonnegative_int(row.get("overdue_interest_days")),
        value_date=_parse_date(value_date_raw)
        if (value_date_raw := str(row.get("value_date") or "").strip())
        else None,
        customer_attribute=str(row.get("customer_attribute") or ""),
        source_version=str(row.get("source_version") or ""),
        rule_version=str(row.get("rule_version") or ""),
        ingest_batch_id=str(row.get("ingest_batch_id") or ""),
        trace_id=str(row.get("trace_id") or ""),
    )


def _to_formal_tyw_fact_row(row: dict[str, object]) -> FormalTywBalanceFactRow:
    maturity_date = str(row.get("maturity_date") or "").strip()
    return FormalTywBalanceFactRow(
        report_date=_parse_date(str(row["report_date"])),
        position_id=str(row["position_id"]),
        product_type=str(row.get("product_type") or ""),
        position_side=str(row.get("position_side") or ""),
        counterparty_name=str(row.get("counterparty_name") or ""),
        account_type=str(row.get("account_type") or ""),
        special_account_type=str(row.get("special_account_type") or ""),
        core_customer_type=str(row.get("core_customer_type") or ""),
        invest_type_std=str(row["invest_type_std"]),
        accounting_basis=str(row["accounting_basis"]),
        position_scope=str(row["position_scope"]),
        currency_basis=str(row["currency_basis"]),
        currency_code=str(row.get("currency_code") or ""),
        principal_amount=_as_decimal(row["principal_amount"]),
        accrued_interest_amount=_as_decimal(row["accrued_interest_amount"]),
        funding_cost_rate=_optional_decimal(row.get("funding_cost_rate")),
        maturity_date=_parse_date(maturity_date) if maturity_date else None,
        source_version=str(row.get("source_version") or ""),
        rule_version=str(row.get("rule_version") or ""),
        ingest_batch_id=str(row.get("ingest_batch_id") or ""),
        trace_id=str(row.get("trace_id") or ""),
    )


def _build_balance_analysis_workbook_xlsx_bytes(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    overview_sheet = workbook.active
    overview_sheet.title = "\u6982\u89c8"
    _write_workbook_cards_sheet(overview_sheet, payload.get("cards") or [])

    for sheet_name, candidate_keys in EXPORT_WORKBOOK_TABLES:
        table = _pick_workbook_export_table(payload, *candidate_keys)
        sheet = workbook.create_sheet(title=sheet_name)
        _write_workbook_table_sheet(sheet, table)

    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _pick_workbook_export_table(payload: dict[str, Any], *candidate_keys: str) -> dict[str, Any]:
    table_map = {
        str(table.get("key")): table
        for table in list(payload.get("tables") or [])
        if isinstance(table, dict)
    }
    for key in candidate_keys:
        if key in table_map:
            return table_map[key]
    raise RuntimeError(
        "Balance-analysis workbook export table unavailable. "
        f"Expected one of {candidate_keys!r}."
    )


def _write_workbook_cards_sheet(sheet: Worksheet, cards: list[dict[str, Any]]) -> None:
    sheet.append(["label", "value"])
    _style_header_row(sheet, column_count=2)
    for card in cards:
        value = _coerce_excel_value(card.get("value"))
        sheet.append([card.get("label"), value])
        _style_numeric_cell(sheet.cell(row=sheet.max_row, column=2), value, column_key="value")
    _autosize_sheet_columns(sheet)


def _write_workbook_table_sheet(sheet: Worksheet, table: dict[str, Any]) -> None:
    columns = list(table.get("columns") or [])
    rows = list(table.get("rows") or [])
    headers = [str(column.get("label") or "") for column in columns]
    column_keys = [str(column.get("key") or "") for column in columns]

    sheet.append(headers)
    _style_header_row(sheet, column_count=len(headers))

    for row in rows:
        values = []
        coerced_values: list[object] = []
        for column_key in column_keys:
            coerced = _coerce_excel_value(row.get(column_key))
            values.append(coerced)
            coerced_values.append(coerced)
        sheet.append(values)
        for column_index, (column_key, value) in enumerate(zip(column_keys, coerced_values), start=1):
            _style_numeric_cell(sheet.cell(row=sheet.max_row, column=column_index), value, column_key=column_key)

    _autosize_sheet_columns(sheet)


def _style_header_row(sheet: Worksheet, *, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        sheet.cell(row=1, column=column_index).font = EXCEL_HEADER_FONT


def _style_numeric_cell(cell: Any, value: object, *, column_key: str) -> None:
    if isinstance(value, bool) or not isinstance(value, (int, float, Decimal)):
        return
    cell.alignment = EXCEL_RIGHT_ALIGNMENT
    if _is_amount_column(column_key):
        cell.number_format = EXCEL_AMOUNT_FORMAT
        return
    if isinstance(value, int):
        cell.number_format = EXCEL_INTEGER_FORMAT
        return
    if isinstance(value, Decimal) and value == value.to_integral_value():
        cell.number_format = EXCEL_INTEGER_FORMAT
        return
    cell.number_format = EXCEL_NUMBER_FORMAT


def _is_amount_column(column_key: str) -> bool:
    normalized = column_key.lower()
    return normalized == "value" or normalized.endswith("_amount") or normalized.endswith("_value")


def _coerce_excel_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float, Decimal)):
        return value
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return ""
        try:
            return Decimal(stripped)
        except InvalidOperation:
            return value
    return value


def _autosize_sheet_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max(len(value) for value in values) if values else 0
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 40)


def _as_decimal(value: object) -> Decimal:
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _optional_decimal(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    return _as_decimal(value)


def _optional_nonnegative_int(value: object) -> int:
    if value in (None, ""):
        return 0
    try:
        return max(0, int(Decimal(str(value))))
    except (InvalidOperation, ValueError, TypeError):
        return 0


def _parse_date(raw_value: str):
    return datetime.strptime(raw_value, "%Y-%m-%d").date()
