from __future__ import annotations

import hashlib
import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from backend.app.core_finance.product_category_pnl import (
    CanonicalFactRow,
    ZERO,
    derive_monthly_pnl,
)


LEDGER_PREFIX = "\u603b\u8d26\u5bf9\u8d26"
AVG_PREFIX = "\u65e5\u5747"
RULE_VERSION = "rv_product_category_pnl_v1"


@dataclass(slots=True)
class SourcePair:
    month_key: str
    report_date: date
    ledger_path: Path
    avg_path: Path
    source_version: str


def discover_source_pairs(source_dir: Path) -> list[SourcePair]:
    source_dir = Path(source_dir)
    if not source_dir.exists():
        return []

    ledger_by_month: dict[str, Path] = {}
    avg_by_month: dict[str, Path] = {}
    for path in sorted(source_dir.glob("*.xlsx")):
        month_key = _extract_month_key(path.name)
        if month_key is None:
            continue
        if path.name.startswith(LEDGER_PREFIX):
            ledger_by_month[month_key] = path
        elif path.name.startswith(AVG_PREFIX):
            avg_by_month[month_key] = path

    pairs: list[SourcePair] = []
    for month_key in sorted(set(ledger_by_month) & set(avg_by_month)):
        year = int(month_key[:4])
        month = int(month_key[4:])
        pairs.append(
            SourcePair(
                month_key=month_key,
                report_date=date(year, month, monthrange(year, month)[1]),
                ledger_path=ledger_by_month[month_key],
                avg_path=avg_by_month[month_key],
                source_version=_build_source_version(
                    ledger_by_month[month_key],
                    avg_by_month[month_key],
                ),
            )
        )
    return pairs


def build_canonical_facts(pair: SourcePair) -> list[CanonicalFactRow]:
    ledger_rows = _parse_ledger_workbook(pair.ledger_path)
    annual_rows, monthly_rows = _parse_average_workbook(pair.avg_path)
    keys = set(ledger_rows) | set(annual_rows) | set(monthly_rows)

    facts: list[CanonicalFactRow] = []
    for account_code, currency in sorted(keys):
        ledger_row = ledger_rows.get((account_code, currency), {})
        facts.append(
            CanonicalFactRow(
                report_date=pair.report_date,
                account_code=account_code,
                currency=currency,
                account_name=str(ledger_row.get("account_name", "")),
                beginning_balance=Decimal(str(ledger_row.get("beginning_balance", ZERO))),
                ending_balance=Decimal(str(ledger_row.get("ending_balance", ZERO))),
                monthly_pnl=Decimal(str(ledger_row.get("monthly_pnl", ZERO))),
                daily_avg_balance=Decimal(str(monthly_rows.get((account_code, currency), ZERO))),
                annual_avg_balance=Decimal(str(annual_rows.get((account_code, currency), ZERO))),
                days_in_period=monthrange(pair.report_date.year, pair.report_date.month)[1],
            )
        )
    return facts


def _parse_ledger_workbook(path: Path) -> dict[tuple[str, str], dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: dict[tuple[str, str], dict[str, object]] = {}
    try:
        for worksheet in workbook.worksheets[:2]:
            for row in worksheet.iter_rows(min_row=7, values_only=True):
                if not row:
                    continue
                if _looks_like_currency(row[2] if len(row) > 2 else None):
                    account_index = 0
                    name_index = 1
                    currency_index = 2
                    amount_offset = 3
                elif _looks_like_currency(row[3] if len(row) > 3 else None):
                    account_index = 1
                    name_index = 2
                    currency_index = 3
                    amount_offset = 4
                else:
                    continue

                if row[account_index] is None:
                    continue
                currency = str(row[currency_index] or "").strip()
                if not currency:
                    continue

                account_code = _coerce_account_code(row[account_index])
                beginning_balance = _to_decimal(row[amount_offset])
                period_debit = _to_decimal(row[amount_offset + 1])
                period_credit = _to_decimal(row[amount_offset + 2])
                ending_balance = _to_decimal(row[amount_offset + 3])
                rows[(account_code, currency)] = {
                    "account_name": str(row[name_index] or "").strip(),
                    "beginning_balance": beginning_balance,
                    "ending_balance": ending_balance,
                    "monthly_pnl": derive_monthly_pnl(period_debit, period_credit),
                }
    finally:
        workbook.close()
    return rows


def _looks_like_currency(value: object) -> bool:
    text = str(value or "").strip()
    return bool(re.fullmatch(r"[A-Z]{3}", text))


def _parse_average_workbook(path: Path) -> tuple[dict[tuple[str, str], Decimal], dict[tuple[str, str], Decimal]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    annual_rows: dict[tuple[str, str], Decimal] = {}
    monthly_rows: dict[tuple[str, str], Decimal] = {}
    try:
        if workbook.worksheets:
            annual_rows = _parse_average_sheet(workbook.worksheets[0])
        if len(workbook.worksheets) > 1:
            monthly_rows = _parse_average_sheet(workbook.worksheets[1])
    finally:
        workbook.close()
    return annual_rows, monthly_rows


def _parse_average_sheet(worksheet) -> dict[tuple[str, str], Decimal]:
    rows: dict[tuple[str, str], Decimal] = {}
    for row in worksheet.iter_rows(min_row=4, values_only=True):
        for index in range(0, len(row), 4):
            currency, account_code, balance = row[index:index + 3]
            if currency is None or account_code is None or balance is None:
                continue
            key = (_coerce_account_code(account_code), str(currency).strip())
            rows[key] = rows.get(key, ZERO) + _to_decimal(balance)
    return rows


def _extract_month_key(file_name: str) -> str | None:
    match = re.search(r"(\d{6})(?=\.xlsx$)", file_name)
    if match is None:
        return None
    return match.group(1)


def _build_source_version(ledger_path: Path, avg_path: Path) -> str:
    parts = []
    for path in (ledger_path, avg_path):
        stat = path.stat()
        parts.append(f"{path.name}:{stat.st_size}:{stat.st_mtime_ns}")
    digest = hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:12]
    return f"sv_product_category_{digest}"


def _coerce_account_code(value: object) -> str:
    if isinstance(value, float):
        return str(int(value))
    return str(value).strip()


def _to_decimal(value: object) -> Decimal:
    if value in (None, ""):
        return ZERO
    return Decimal(str(value))
